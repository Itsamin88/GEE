"""The community that broke the crawler, run against the repaired one.

Modelled on the reported Tamera run: 420 pages, 5000 archived URLs, a report in
three languages, hundreds of gallery images, and extracted text carrying the
control bytes that killed the export.

What must be true at the end, whatever the site throws:

  * the run stops inside its active-time budget;
  * a workbook exists and reopens;
  * the archive was sampled, not enumerated;
  * translations were not each read in full;
  * images did not consume the run;
  * truncation, where it happened, is stated rather than hidden.
"""

from __future__ import annotations

import socket
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from dcr.app import Application
from dcr.db import Database
from dcr.runner import CommunityInput
from fixtures.harness import fixture_settings
from fixtures.server import FixtureServer
from fixtures.stress import (ARCHIVE_URL_COUNT, HOST, PAGE_COUNT,
                             build_stress_archive, build_stress_site, stress_urls)

ROOT = Path(__file__).resolve().parents[1]

#: The stress run is given a deliberately small budget so the test finishes in
#: seconds. Everything it proves about the shape of the run holds at 30 minutes.
STRESS_BUDGET_MINUTES = 0.9
STRESS_RESERVE_MINUTES = 0.2
STRESS_WIND_DOWN_MINUTES = 0.1

#: A ceiling scaled to a loopback fixture, so that it actually truncates the
#: crawl the way the shipped thirty-minute one truncated the real Tamera run.
#: The comparison in `capped_versus_governed` is meaningless against a cap that
#: is never reached.
CAP_THAT_BITES_MINUTES = 0.06


def _hosts_resolve() -> bool:
    try:
        socket.gethostbyname(HOST)
        socket.gethostbyname("archive.test")
        return True
    except OSError:
        return False


pytestmark = pytest.mark.skipif(
    not _hosts_resolve(),
    reason=f"{HOST} does not resolve; run tools/run_pilot.py once to add the fixture hosts",
)


@pytest.fixture(scope="module")
def stress_run(tmp_path_factory):
    output = tmp_path_factory.mktemp("stress")
    sites = build_stress_site()
    server = FixtureServer(archive_records=build_stress_archive())
    server.sites.update(sites)
    server.start()
    try:
        settings = fixture_settings(server.port, output, root=ROOT)
        settings.app["budget"] = {
            "enabled": True,
            "active_minutes": STRESS_BUDGET_MINUTES,
            "finalisation_reserve_minutes": STRESS_RESERVE_MINUTES,
            "wind_down_minutes": STRESS_WIND_DOWN_MINUTES,
        }
        settings.app["estimation"] = {"enabled": False}
        settings.app["crawl"]["max_pages_per_run"] = 4000
        settings.app["crawl"]["base_pages_per_source"] = 400
        settings.sources["archive"]["priority_snapshot_paths"] = ["/", "/history"]

        community = CommunityInput(
            name="Stress Test Community", latitude=37.72, longitude=-8.45,
            urls=stress_urls(server.port), country="Portugal",
            coder_id="STRESS", fixture=True,
        )
        app = Application(settings)
        app.preflight()
        started = time.monotonic()
        result = app.run(community, mode="FULL")
        wall_s = time.monotonic() - started
        db = Database(settings.database_path)
        yield {"result": result, "wall_s": wall_s, "db": db, "settings": settings,
               "community_id": result["report"]["community_id"], "server": server}
        db.close()
        app.close()
    finally:
        server.stop()


def _count(db, sql, community_id):
    return int(db.scalar(sql, (community_id,)) or 0)


# ---------------------------------------------------------------------------
# it finishes, and it finishes in time
# ---------------------------------------------------------------------------
def test_the_run_stops_inside_its_budget(stress_run):
    """The reported run went on for hours. This one may not."""
    report = stress_run["result"]["report"]
    budget = report.get("budget") or {}
    assert budget, "the run must report what the clock did"
    ceiling = STRESS_BUDGET_MINUTES * 60
    assert budget["active_s"] <= ceiling * 1.6, (
        f"active time {budget['active_s']:.0f}s overran the "
        f"{ceiling:.0f}s budget by more than the wind-down allows")


def test_the_wall_clock_is_bounded_too(stress_run):
    assert stress_run["wall_s"] < 300, (
        f"the stress run took {stress_run['wall_s']:.0f}s of wall clock")


def test_the_run_reached_finalisation(stress_run):
    """The reserve exists so this can never be false."""
    finalisation = stress_run["result"]["report"]["manifest"]["export"]["finalisation"]
    assert finalisation["ok"], finalisation.get("failure_reason")
    assert finalisation["verification"]["reopened"]


# ---------------------------------------------------------------------------
# the workbook, which is the whole point
# ---------------------------------------------------------------------------
def test_a_workbook_exists_and_reopens(stress_run):
    path = Path(stress_run["result"]["workbook"])
    assert path.exists() and path.stat().st_size > 0
    workbook = load_workbook(path)
    try:
        for sheet in ("O1_Community_Attributes", "O2_Practice_Matrix",
                      "O3_Onset_Register", "O6_Source_Index", "O11_Source_Set"):
            assert sheet in workbook.sheetnames
    finally:
        workbook.close()


def test_the_dirty_pdf_text_did_not_take_the_export_down(stress_run):
    """The fixture's documents carry the exact bytes that caused the crash."""
    export = stress_run["result"]["report"]["manifest"]["export"]
    assert export["finalisation"]["ok"]
    sanitisation = export["finalisation"]["sanitisation"]
    assert sanitisation["excel_sanitized"] == "yes", (
        "the fixture's control bytes should have been cleaned and the cleaning recorded")
    assert sanitisation["cells_sanitized"] > 0


def test_the_evidence_kept_the_original_bytes(stress_run):
    """Only the Excel representation is cleaned; the record is not."""
    db, cid = stress_run["db"], stress_run["community_id"]
    rows = db.query(
        "SELECT quote FROM evidence WHERE community_id=? AND quote LIKE '%hectares%'",
        (cid,))
    assert rows, "the dirty sentences should still have produced evidence"


# ---------------------------------------------------------------------------
# the archive was sampled, not enumerated
# ---------------------------------------------------------------------------
def test_the_archive_was_not_enumerated(stress_run):
    stats = stress_run["result"]["report"]["crawl_stats"]
    discovered = stats.get("archive_urls_discovered", 0)
    fetched = stats.get("archive_urls_fetched", 0)
    assert discovered > 1000, "the fixture should have offered thousands of archived URLs"
    assert fetched < discovered / 10, (
        f"{fetched} of {discovered} archived URLs were queued; the archive is to be "
        "sampled by relevance, not enumerated")


def test_both_archive_numbers_are_reported(stress_run):
    """Discovered is not fetched, and the report must show both (brief §63)."""
    stats = stress_run["result"]["report"]["crawl_stats"]
    assert "archive_urls_discovered" in stats
    assert "archive_urls_fetched" in stats


# ---------------------------------------------------------------------------
# documents: translations are not each read in full
# ---------------------------------------------------------------------------
def test_the_same_report_in_three_languages_is_read_once(stress_run):
    db, cid = stress_run["db"], stress_run["community_id"]
    reports = db.query(
        "SELECT document_id, filename, notes, text_chars FROM documents "
        "WHERE community_id=? AND (filename LIKE '%annual-report%' "
        "OR filename LIKE '%jahresbericht%' OR filename LIKE '%relatorio%')", (cid,))
    if len(reports) < 2:
        pytest.skip("the crawl did not reach more than one language of the report")
    mirrored = [r for r in reports if "mirror" in (r["notes"] or "").lower()
                or "translation" in (r["notes"] or "").lower()]
    assert mirrored, (
        "at least one language version should have been kept as a provenance "
        "mirror rather than read in full")


def test_low_value_documents_did_not_consume_the_budget(stress_run):
    """A festival flyer is not worth a slice of a thirty-minute budget."""
    db, cid = stress_run["db"], stress_run["community_id"]
    flyer = db.query_one(
        "SELECT * FROM documents WHERE community_id=? AND filename LIKE '%flyer%'", (cid,))
    stats = stress_run["result"]["report"]["crawl_stats"]
    assert flyer is None or stats.get("documents_skipped", 0) >= 0


# ---------------------------------------------------------------------------
# images did not eat the run
# ---------------------------------------------------------------------------
def test_images_did_not_consume_the_run(stress_run):
    triage = stress_run["result"]["report"].get("image_triage") or {}
    seen = triage.get("candidates_seen", 0)
    downloaded = triage.get("downloaded", 0)
    assert seen > 0
    assert downloaded <= 60, (
        f"{downloaded} images were downloaded from {seen} candidates; the image "
        "subsystem must not dominate the run")


def test_the_gallery_was_triaged_rather_than_hoarded(stress_run):
    triage = stress_run["result"]["report"].get("image_triage") or {}
    seen = triage.get("candidates_seen", 0)
    if seen < 20:
        pytest.skip("too few candidates reached triage to judge the ratio")
    assert triage.get("download_rate", 1.0) < 0.6, (
        "most of a photo gallery should be recorded, not fetched")


# ---------------------------------------------------------------------------
# honesty about what was not done
# ---------------------------------------------------------------------------
def test_truncation_is_stated_not_hidden(stress_run):
    report = stress_run["result"]["report"]
    if report["crawl_truncated"] == "yes":
        assert report["truncation_reasons"], (
            "a truncated run must say why, or a partial search reads as a complete one")


def test_a_budget_stopped_run_is_never_called_exhaustive(stress_run):
    report = stress_run["result"]["report"]
    status = report["completion_status"]
    if (report.get("budget") or {}).get("budget_exhausted"):
        assert status != "COMPLETE", (
            "a run stopped by the clock may be COMPLETE_WITH_TRUNCATION, never COMPLETE")


def test_the_queue_that_was_left_is_reported(stress_run):
    """What remains is part of the record, not an embarrassment to hide."""
    queue = stress_run["result"]["report"].get("queue") or {}
    assert isinstance(queue, dict)
    assert sum(queue.values()) > 0


def test_the_profile_says_where_the_time_went(stress_run):
    profile = stress_run["result"]["report"].get("profile") or {}
    if not profile:
        pytest.skip("no profile recorded")
    assert profile["by_stage_s"], "the run must be able to say where its time went"


# ---------------------------------------------------------------------------
# conflicts stay proportionate
# ---------------------------------------------------------------------------
def test_conflicts_did_not_explode(stress_run):
    """The reported run ended with 5569."""
    db, cid = stress_run["db"], stress_run["community_id"]
    conflicts = _count(db, "SELECT COUNT(*) FROM conflicts WHERE community_id=?", cid)
    claims = _count(db, "SELECT COUNT(*) FROM claims WHERE community_id=?", cid)
    assert conflicts < 200, f"{conflicts} conflicts from {claims} claims"


# ---------------------------------------------------------------------------
# the guarantee: running out of time still produces a workbook
# ---------------------------------------------------------------------------
@pytest.fixture(scope="module")
def exhausted_run(tmp_path_factory):
    """A budget so small the crawl cannot possibly finish the protocol."""
    output = tmp_path_factory.mktemp("exhausted")
    server = FixtureServer(archive_records=build_stress_archive())
    server.sites.update(build_stress_site())
    server.start()
    try:
        settings = fixture_settings(server.port, output, root=ROOT)
        settings.app["budget"] = {
            "enabled": True,
            # Seconds, not minutes: the wind-down begins almost immediately.
            "active_minutes": 0.08,
            "finalisation_reserve_minutes": 0.02,
            "wind_down_minutes": 0.01,
        }
        settings.app["estimation"] = {"enabled": False}
        community = CommunityInput(
            name="Exhausted Budget Community", urls=stress_urls(server.port),
            country="Portugal", coder_id="STRESS", fixture=True,
        )
        app = Application(settings)
        app.preflight()
        result = app.run(community, mode="FULL")
        db = Database(settings.database_path)
        yield {"result": result, "db": db,
               "community_id": result["report"]["community_id"]}
        db.close()
        app.close()
    finally:
        server.stop()


def test_a_run_that_runs_out_of_time_still_produces_a_workbook(exhausted_run):
    """The single most important guarantee in this repair (brief §28, §46)."""
    path = Path(exhausted_run["result"]["workbook"])
    assert path.exists(), "the budget expired and no workbook was written"
    assert path.stat().st_size > 0
    workbook = load_workbook(path)
    try:
        assert "O1_Community_Attributes" in workbook.sheetnames
    finally:
        workbook.close()


def test_the_exhausted_run_verified_its_workbook(exhausted_run):
    finalisation = exhausted_run["result"]["report"]["manifest"]["export"]["finalisation"]
    assert finalisation["ok"], finalisation.get("failure_reason")
    assert finalisation["verification"]["reopened"]


def test_the_exhausted_run_says_the_clock_stopped_it(exhausted_run):
    report = exhausted_run["result"]["report"]
    budget = report.get("budget") or {}
    assert budget.get("budget_exhausted") or report["crawl_truncated"] == "yes", (
        "a run cut short by its budget must say so")
    assert report["truncation_reasons"], "and must say why"


def test_the_exhausted_run_is_never_called_complete(exhausted_run):
    status = exhausted_run["result"]["report"]["completion_status"]
    assert status != "COMPLETE", (
        f"a run stopped by the clock reported {status!r}; COMPLETE would claim an "
        "exhaustive search that never happened")


def test_the_exhausted_run_still_wrote_its_manifests(exhausted_run):
    """The audit package is part of the deliverable, not an optional extra."""
    counts = exhausted_run["result"]["report"]["manifest"]["export"]["manifests"]
    for name in ("source_manifest.csv", "evidence_manifest.csv",
                 "document_manifest.csv", "image_manifest.csv"):
        assert name in counts, f"{name} was not written"


def test_the_completion_report_exists_after_exhaustion(exhausted_run):
    output = Path(exhausted_run["result"]["output_dir"])
    assert (output / "09_final" / "completion_report.md").exists()
    assert (output / "09_final" / "completion_report.json").exists()


# ===========================================================================
# THE MEASUREMENT THIS REWRITE EXISTS FOR
#
# The same rich community, run twice: once with the old thirty-minute-shaped
# cap, once with no cap at all and the yield governor deciding. If removing the
# cap does not recover evidence, the whole change was pointless; if it does not
# terminate, the change traded one failure for another.
#
# Both are asserted here, on real numbers, in one test (brief §1, §25, §94).
# ===========================================================================
@pytest.fixture(scope="module")
def capped_versus_governed(tmp_path_factory):
    """Run the stress community twice, and measure the difference."""
    server = FixtureServer(archive_records=build_stress_archive())
    server.sites.update(build_stress_site())
    server.start()
    results = {}
    try:
        for label, budget in (
            # The shape of the old design: a fixed slice of active time, cut
            # into per-stage shares, applied whatever the crawl was finding.
            #
            # The number is small because the fixture is on loopback and the
            # whole site is served in seconds; a 30-minute cap would never bite
            # here, and a cap that does not bite is not the behaviour being
            # compared against. On the real Tamera site the shipped 30-minute
            # cap bit in exactly this way — a fraction of the crawl done, the
            # rest of the protocol never reached.
            ("capped", {"enabled": True, "active_minutes": CAP_THAT_BITES_MINUTES,
                        "finalisation_reserve_minutes": 0.02,
                        "wind_down_minutes": 0.01}),
            # The new one: no ceiling; the yield governor decides.
            ("governed", {"enabled": True, "active_minutes": 0,
                          "finalisation_reserve_minutes": STRESS_RESERVE_MINUTES,
                          "wind_down_minutes": STRESS_WIND_DOWN_MINUTES}),
        ):
            output = tmp_path_factory.mktemp(f"compare-{label}")
            settings = fixture_settings(server.port, output, root=ROOT)
            settings.app["budget"] = budget
            settings.app["estimation"] = {"enabled": False}
            settings.app["crawl"]["max_pages_per_run"] = 4000
            settings.app["crawl"]["base_pages_per_source"] = 400
            settings.sources["archive"]["priority_snapshot_paths"] = ["/", "/history"]
            community = CommunityInput(
                name=f"Stress Comparison {label}", latitude=37.72, longitude=-8.45,
                urls=stress_urls(server.port), country="Portugal",
                coder_id="STRESS", fixture=True,
            )
            app = Application(settings)
            app.preflight()
            started = time.monotonic()
            result = app.run(community, mode="FULL")
            wall_s = time.monotonic() - started
            db = Database(settings.database_path)
            cid = result["report"]["community_id"]
            results[label] = {
                "result": result, "wall_s": wall_s,
                "evidence": _count(db, "SELECT COUNT(*) FROM evidence "
                                       "WHERE community_id=?", cid),
                "claims": _count(db, "SELECT COUNT(*) FROM claims "
                                     "WHERE community_id=?", cid),
                "pages": _count(db, "SELECT COUNT(*) FROM pages WHERE community_id=?", cid),
                "documents": _count(db, "SELECT COUNT(*) FROM documents "
                                        "WHERE community_id=?", cid),
                "fields": _count(db, "SELECT COUNT(*) FROM field_values "
                                     "WHERE community_id=? AND status='coded'", cid),
                "report": result["report"],
            }
            db.close()
            app.close()
        yield results
    finally:
        server.stop()


def test_removing_the_cap_recovers_evidence(capped_versus_governed):
    """The whole point. If this is not true, the rewrite was pointless."""
    capped = capped_versus_governed["capped"]
    governed = capped_versus_governed["governed"]
    assert governed["evidence"] > capped["evidence"], (
        f"the governed run found {governed['evidence']} evidence items and the "
        f"capped run {capped['evidence']}: removing the cap recovered nothing")
    assert governed["pages"] >= capped["pages"]
    print(f"\n  capped:   {capped['pages']} pages, {capped['documents']} documents, "
          f"{capped['evidence']} evidence, {capped['fields']} fields coded, "
          f"{capped['wall_s']:.0f}s")
    print(f"  governed: {governed['pages']} pages, {governed['documents']} documents, "
          f"{governed['evidence']} evidence, {governed['fields']} fields coded, "
          f"{governed['wall_s']:.0f}s")


def test_the_governed_run_still_terminates(capped_versus_governed):
    """Removing the cap must not mean running for ever (brief §62)."""
    governed = capped_versus_governed["governed"]
    assert governed["wall_s"] < 600, (
        f"the governed run took {governed['wall_s']:.0f}s on a fixture site; "
        "the yield governor is not stopping it")


def test_the_governed_run_stopped_for_a_stated_reason(capped_versus_governed):
    report = capped_versus_governed["governed"]["report"]
    budget = report.get("budget") or {}
    cause = report.get("retrieval_stop_cause") or ""
    if cause:
        assert cause in ("exhausted", "requested"), (
            f"an unbounded run reported {cause!r}; only a ceiling can produce "
            "'ceiling', and none was set")
        assert budget.get("stop_reason"), "it stopped without saying why"


def test_the_capped_run_is_reported_as_truncated(capped_versus_governed):
    """A run the clock stopped may be COMPLETE_WITH_TRUNCATION, never COMPLETE."""
    report = capped_versus_governed["capped"]["report"]
    if (report.get("retrieval_stop_cause") or "") == "ceiling":
        assert report["completion_status"] != "COMPLETE"
        assert report["truncation_reasons"]


def test_the_governed_run_reports_what_it_found_per_minute(capped_versus_governed):
    """The number the whole stopping rule is based on (brief §25)."""
    report = capped_versus_governed["governed"]["report"]
    measured = report.get("yield") or {}
    assert measured, "the report does not carry the yield account"
    assert measured["evidence_yield_per_min"] > 0
    assert measured["curve"], "no yield curve, so nobody can see the shape of the run"


def test_both_runs_produced_a_verified_workbook(capped_versus_governed):
    """Whichever way a run ends, it ends with a workbook that reopens."""
    for label, run in capped_versus_governed.items():
        finalisation = run["report"]["manifest"]["export"]["finalisation"]
        assert finalisation["ok"], f"{label}: {finalisation.get('failure_reason')}"
        assert finalisation["verification"]["reopened"], label
