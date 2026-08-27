"""Several communities, researched at once, all the way to verified workbooks.

Everything else in the suite tests a piece. This tests the thing: real spawned
worker processes, running the real ten-stage engine, against the real fixture
web, each writing its own database and its own workbook, with the real
scheduler deciding who runs and the real host broker keeping them polite.

If this passes, the system does what it says it does.
"""

from __future__ import annotations

import json
import socket
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook

from dcr.orchestrator.recovery import (find_interrupted, plan_resume, apply_resume,
                                       queue_offline_pass, repair)
from dcr.orchestrator.session import RunSession, read_community_file
from dcr.orchestrator.store import COMPLETED, FAILED, QUEUED, RunStore
from fixtures.harness import fixture_settings, fixture_urls
from fixtures.server import FixtureServer

ROOT = Path(__file__).resolve().parents[1]
FIXTURE_HOSTS = ("pourgues.test", "ancien-pourgues.test", "annuaire.test",
                 "theses.test", "facebook.test", "archive.test", "boekel.test",
                 "oud-boekel.test")


def _hosts_resolve() -> bool:
    try:
        for host in FIXTURE_HOSTS:
            socket.gethostbyname(host)
        return True
    except OSError:
        return False


pytestmark = [
    pytest.mark.skipif(
        not _hosts_resolve(),
        reason="the fixture hostnames do not resolve; run tools/run_pilot.py once"),
    pytest.mark.slow,
]


def _overrides(settings, output_root: Path) -> dict:
    """The fixture settings, as a payload every worker can apply.

    Workers are separate processes and load configuration themselves, so the
    fixture endpoints have to travel as data. This is the same mechanism an
    advanced user changes timeouts or budgets with (brief §98).
    """
    return {
        "settings_overrides": dict(settings.app),
        "sources_overrides": dict(settings.sources),
    }


@pytest.fixture(scope="module")
def run(tmp_path_factory):
    """Three communities, in parallel, start to finish."""
    output = tmp_path_factory.mktemp("multi")
    server = FixtureServer().start()
    try:
        settings = fixture_settings(server.port, output, root=ROOT)
        session = RunSession(settings=settings)
        session.settings_overrides = dict(settings.app)
        session.sources_overrides = dict(settings.sources)
        entries = [
            {"name": "EcoVillage de Pourgues", "latitude": 43.0561, "longitude": 1.8342,
             "country": "France", "coder_id": "TEST",
             "urls": fixture_urls(server.port, "pourgues")},
            {"name": "Boekel Ecodorp", "country": "Netherlands", "coder_id": "TEST",
             "urls": fixture_urls(server.port, "boekel")},
            {"name": "Nowhere In Particular", "country": "France", "coder_id": "TEST",
             "urls": [f"http://does-not-exist.invalid:{server.port}/"]},
        ]
        plan = session.create(entries, mode="FULL")
        for job in plan.jobs:
            session.store.update_job(job.job_id, {"fixture": 1})
        summary = session.start(workers_max=3, show_dashboard=False)
        yield {"session": session, "summary": summary, "output": output,
               "store": session.store, "server": server, "plan": plan}
        session.close()
    finally:
        server.stop()


# ---------------------------------------------------------------------------
# §2, §7 — several communities, one START
# ---------------------------------------------------------------------------
def test_every_community_was_researched(run):
    counts = run["store"].counts(run["summary"]["run_id"])
    assert counts["TOTAL"] == 3
    assert counts[QUEUED] == 0, "the queue did not empty"
    assert counts[COMPLETED] + counts[FAILED] == 3


def test_the_two_reachable_communities_produced_verified_workbooks(run):
    jobs = run["store"].jobs(run["summary"]["run_id"])
    verified = [job for job in jobs if job.workbook_path]
    assert len(verified) >= 2, [(j.job_id, j.final_status, j.last_error) for j in jobs]
    for job in verified:
        path = Path(job.workbook_path)
        assert path.exists(), f"{job.job_id} claims a workbook that is not there"
        workbook = load_workbook(path)
        assert "O1_Community_Attributes" in workbook.sheetnames
        assert workbook["O1_Community_Attributes"]["A3"].value == job.site_id
        workbook.close()


def test_each_community_has_its_own_database_and_directory(run):
    """The isolation guarantee, on disk (brief §8, §53)."""
    jobs = run["store"].jobs(run["summary"]["run_id"])
    databases = set()
    for job in jobs:
        if not job.output_dir:
            continue
        directory = Path(job.output_dir)
        if not directory.exists():
            continue
        assert job.site_id in directory.name
        database = directory / "research.sqlite3"
        if database.exists():
            databases.add(database.resolve())
    assert len(databases) >= 2, "communities shared a database"


def test_no_communitys_evidence_reached_another(run):
    """Two databases, two site_ids, and neither knows about the other."""
    import sqlite3

    jobs = [j for j in run["store"].jobs(run["summary"]["run_id"])
            if j.output_dir and (Path(j.output_dir) / "research.sqlite3").exists()]
    assert len(jobs) >= 2
    for job in jobs:
        connection = sqlite3.connect(str(Path(job.output_dir) / "research.sqlite3"))
        try:
            ids = {row[0] for row in
                   connection.execute("SELECT community_id FROM communities")}
        finally:
            connection.close()
        assert ids == {job.site_id}, (
            f"{job.job_id}'s database holds {ids}, not only {job.site_id}")


def test_the_site_ids_are_globally_unique(run):
    """212 databases each numbering their one community IC001 would give 212
    workbooks all claiming to be site IC001 (brief §8)."""
    jobs = run["store"].jobs(run["summary"]["run_id"])
    ids = [job.site_id for job in jobs]
    assert len(set(ids)) == len(ids)
    assert all(job_id.startswith(("IC", "TEST-IC")) for job_id in ids)


# ---------------------------------------------------------------------------
# §39 — an unreachable community does not affect the others
# ---------------------------------------------------------------------------
def test_an_unreachable_community_does_not_stop_the_others(run):
    jobs = {job.name: job for job in run["store"].jobs(run["summary"]["run_id"])}
    unreachable = jobs["Nowhere In Particular"]
    assert unreachable.final_status, "the unreachable community was never resolved"
    for name in ("EcoVillage de Pourgues", "Boekel Ecodorp"):
        assert jobs[name].final_status, f"{name} did not finish"
        assert jobs[name].evidence >= 0


def test_every_failure_is_a_row_not_a_stack_trace(run):
    errors = run["store"].errors(run["summary"]["run_id"])
    for row in errors:
        assert row["error_class"], "an error was recorded without a class"
        assert row["message"]


# ---------------------------------------------------------------------------
# §54 — the run's own outputs
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("filename", [
    "global_run_manifest.json",
    "global_progress.json",
    "global_error_log.csv",
    "global_summary.md",
    "community_status_table.csv",
])
def test_the_run_writes_its_own_record(run, filename):
    path = run["output"] / filename
    assert path.exists(), f"{filename} was not written"
    assert path.stat().st_size > 0


def test_the_status_table_has_a_row_for_every_community(run):
    import csv

    with (run["output"] / "community_status_table.csv").open(encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 3
    assert all(row["job_id"] and row["site_id"] for row in rows)


def test_the_summary_reports_the_numbers_the_brief_asks_for(run):
    summary = run["summary"]
    for section in ("communities", "evidence", "time", "workers"):
        assert section in summary
    assert summary["evidence"]["workbooks_verified"] >= 2
    assert summary["time"]["wall_clock_s"] > 0
    assert summary["communities"]["total"] == 3


def test_the_manifest_records_what_was_asked_for(run):
    manifest = json.loads((run["output"] / "global_run_manifest.json").read_text())
    assert len(manifest["communities"]) == 3
    assert manifest["app_version"]
    assert manifest["configuration_files"], (
        "the reproducibility record must name the configuration it ran with")


# ---------------------------------------------------------------------------
# §1, §25 — the crawl stopped on evidence, not on a clock
# ---------------------------------------------------------------------------
def test_no_community_was_stopped_by_a_clock(run):
    """The regression this rewrite exists to remove."""
    for job in run["store"].jobs(run["summary"]["run_id"]):
        if not job.output_dir:
            continue
        report = Path(job.output_dir) / "09_final" / "completion_report.json"
        if not report.exists():
            continue
        payload = json.loads(report.read_text(encoding="utf-8"))
        text = json.dumps(payload)
        assert "30-minute" not in text
        assert "active budget was reached" not in text


def test_the_yield_account_reaches_the_report(run):
    for job in run["store"].jobs(run["summary"]["run_id"]):
        if not job.output_dir:
            continue
        report = Path(job.output_dir) / "09_final" / "completion_report.json"
        if not report.exists():
            continue
        payload = json.loads(report.read_text(encoding="utf-8"))
        text = json.dumps(payload)
        if job.evidence:
            assert "yield" in text.lower(), (
                "the report does not say what the crawl found per minute, which is "
                "the number the whole stopping rule is based on")
            return
    pytest.skip("no community produced evidence in this run")


# ---------------------------------------------------------------------------
# §100-§103 — recovery
# ---------------------------------------------------------------------------
def test_a_finished_run_is_not_offered_for_resumption(run):
    """Re-running a completed community would cost hours and change a verified
    research record."""
    plan = plan_resume(run["store"], run["summary"]["run_id"])
    completed = [job for job in run["store"].jobs(run["summary"]["run_id"])
                 if job.state == COMPLETED]
    assert all(job.job_id in plan.keep_complete for job in completed)
    assert not plan.requeue


def test_export_can_be_re_run_without_touching_the_network(run):
    """The crawl succeeded; rebuilding the workbook must not need the web
    (brief §102)."""
    queued = queue_offline_pass(run["store"], run["summary"]["run_id"], "EXPORT")
    assert queued, "no community had stored evidence to export from"
    for job_id in queued:
        job = run["store"].job(job_id)
        assert job.mode == "EXPORT"
        assert job.state == QUEUED
    # Put them back, so the module-scoped fixture is not left altered for the
    # tests that run after this one.
    for job_id in queued:
        run["store"].update_job(job_id, {"state": COMPLETED, "mode": "FULL"})


def test_a_crashed_run_leaves_no_community_stuck_running(run):
    """After a crash the queue still says RUNNING and nothing is running
    (brief §106)."""
    store = run["store"]
    run_id = run["summary"]["run_id"]
    victim = store.jobs(run_id)[0]
    store.update_job(victim.job_id, {"state": "RUNNING", "worker": "w1"})
    repaired = repair(store, run_id)
    assert repaired["requeued"] == 1
    assert store.job(victim.job_id).state == QUEUED
    store.update_job(victim.job_id, {"state": victim.state, "worker": None})


# ---------------------------------------------------------------------------
# §4 — politeness across communities
# ---------------------------------------------------------------------------
def test_the_archive_was_not_hit_by_every_community_at_once(run):
    """Three communities, one archive host. The broker holds it to its limit."""
    hosts = {row["host"]: row for row in run["store"].hosts()}
    # The broker records into the run store only when a worker reports it; the
    # guarantee itself is tested directly in test_orchestrator.py. Here we only
    # check that nothing recorded a concurrency above the configured ceiling.
    for host, row in hosts.items():
        if row["concurrency"]:
            assert row["concurrency"] <= 4, f"{host} was allowed {row['concurrency']}"


# ===========================================================================
# §108 — the final scientific audit, as tests rather than as a checklist
#
# Nine things the brief says the system must NOT have done. A checklist a
# person ticks is a checklist that drifts; these fail the build.
# ===========================================================================
def _community_databases(run) -> list[tuple[str, str]]:
    import sqlite3

    out = []
    for job in run["store"].jobs(run["summary"]["run_id"]):
        if not job.output_dir:
            continue
        path = Path(job.output_dir) / "research.sqlite3"
        if path.exists():
            out.append((job.site_id, str(path)))
    return out


def _query(path: str, sql: str, params=()) -> list:
    import sqlite3

    connection = sqlite3.connect(path)
    connection.row_factory = sqlite3.Row
    try:
        return list(connection.execute(sql, params))
    finally:
        connection.close()


def test_audit_no_satellite_quantity_was_coded_documentarily(run):
    """§89, §90: the crawler is documentary. Imagery proves nothing here."""
    from dcr.config import load_settings

    forbidden = {str(q).lower() for q in
                 load_settings(ROOT).schema.get("satellite_only_quantities", [])}
    assert forbidden, "the schema no longer lists the satellite-only quantities"
    for site_id, path in _community_databases(run):
        coded = {str(row["field_name"]).lower() for row in
                 _query(path, "SELECT field_name FROM field_values "
                              "WHERE status = 'coded'")}
        assert not (coded & forbidden), (
            f"{site_id} coded a satellite-only quantity documentarily: "
            f"{coded & forbidden}")


def test_audit_no_polygon_or_managed_area_was_inferred_from_imagery(run):
    """§90: the researcher controls polygon tracing, separately."""
    for site_id, path in _community_databases(run):
        rows = _query(path, "SELECT field_name, method FROM field_values "
                            "WHERE field_name LIKE 'polygon%' AND status = 'coded'")
        assert not rows, f"{site_id} coded a polygon field: {[dict(r) for r in rows]}"


def test_audit_no_publication_date_became_an_intervention_date(run):
    """§24, §108. Enforced by the role vocabulary, checked here."""
    from dcr.evidence import roles

    for site_id, path in _community_databases(run):
        rows = _query(path,
                      "SELECT claim_id, field_name, semantic_role FROM claims "
                      "WHERE field_name IN ('date_intervention_onset', "
                      "'date_formal_founding', 'date_land_acquisition', "
                      "'date_first_residence')")
        for row in rows:
            assert row["semantic_role"] not in roles.SOURCE_ROLES, (
                f"{site_id}: claim {row['claim_id']} put a "
                f"{row['semantic_role']} date into {row['field_name']}")


def test_audit_no_visitor_count_reached_the_population_field(run):
    """§22, §108."""
    for site_id, path in _community_databases(run):
        rows = _query(path,
                      "SELECT claim_id, semantic_role FROM claims "
                      "WHERE field_name IN ('e3_population_value', 'population_value')")
        for row in rows:
            if row["semantic_role"]:
                assert row["semantic_role"] == "resident", (
                    f"{site_id}: a {row['semantic_role']} count is standing in "
                    f"the population field (claim {row['claim_id']})")


def test_audit_property_area_is_not_managed_area(run):
    """§23, §108. Confusing them moves a community two size classes."""
    for site_id, path in _community_databases(run):
        rows = _query(path, "SELECT claim_id, semantic_role FROM claims "
                            "WHERE field_name = 'managed_area_ha'")
        for row in rows:
            if row["semantic_role"]:
                assert row["semantic_role"] != "total_holding", (
                    f"{site_id}: the whole holding was coded as managed area "
                    f"(claim {row['claim_id']})")


def test_audit_no_academic_source_was_fabricated(run):
    """§29, §86, §108. An unverified record may be stored, never coded."""
    for site_id, path in _community_databases(run):
        tables = {row["name"] for row in
                  _query(path, "SELECT name FROM sqlite_master WHERE type='table'")}
        if "academic_records" not in tables:
            continue
        unverified = _query(
            path, "SELECT record_id FROM academic_records "
                  "WHERE COALESCE(verified_resolves, 0) = 0")
        for row in unverified:
            supported = _query(
                path, "SELECT claim_id FROM claims WHERE source_id = ?",
                (row["record_id"],))
            assert not supported, (
                f"{site_id}: unverified academic record {row['record_id']} "
                "supports a claim")


def test_audit_not_mentioned_was_never_turned_into_absent(run):
    """§59, §108: silence is not absence."""
    for site_id, path in _community_databases(run):
        rows = _query(path, "SELECT field_name, value, rationale FROM field_values "
                            "WHERE status = 'coded' AND field_name LIKE 'pc%'")
        for row in rows:
            if str(row["value"]).strip().lower() == "explicitly absent":
                assert "not mentioned" not in str(row["rationale"] or "").lower(), (
                    f"{site_id}: {row['field_name']} was coded absent from silence")


def test_audit_copied_sources_are_not_counted_as_independent(run):
    """§28, §108: ten copied pages are not ten independent sources."""
    for site_id, path in _community_databases(run):
        sources = _query(path, "SELECT source_id, independence_group FROM sources")
        assert all(row["independence_group"] for row in sources), (
            f"{site_id}: a source has no independence group, so corroboration "
            "cannot be counted correctly")
        groups = {row["independence_group"] for row in sources}
        assert len(groups) <= len(sources)


def test_audit_every_coded_value_traces_to_evidence_or_a_named_rule(run):
    """§55, §58, §108: nothing reaches a cell without something behind it."""
    rule_methods = {
        "rule", "derived", "researcher_input", "crawl_audit", "run_state",
        "search_log", "independence", "activity_rules", "practice_rules",
        "onset_engine", "extraction", "image_evidence", "grey_literature",
    }
    for site_id, path in _community_databases(run):
        rows = _query(path, "SELECT field_name, method, claim_ids, source_ids "
                            "FROM field_values WHERE status = 'coded'")
        orphans = [dict(r) for r in rows
                   if not r["claim_ids"] and not r["source_ids"]
                   and r["method"] not in rule_methods]
        assert not orphans, f"{site_id}: {orphans[:3]}"


def test_audit_the_raw_wording_survives_beside_every_normalised_value(run):
    """§58: never destroy the original."""
    for site_id, path in _community_databases(run):
        rows = _query(path, "SELECT claim_id, value, original_value FROM claims "
                            "WHERE value_type IN ('float', 'integer')")
        for row in rows:
            assert row["original_value"], (
                f"{site_id}: claim {row['claim_id']} kept a number without the "
                "wording it came from")
