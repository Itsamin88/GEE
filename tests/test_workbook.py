"""Workbook audit and export: the study's destination must not be damaged."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from dcr.db import utcnow
from dcr.export.workbook import WorkbookExporter
from dcr.workbook_audit import audit, profile_workbook


def test_template_matches_the_configured_schema(settings, schema):
    result = audit(settings.workbook_template, schema)
    assert result.ok, "\n".join(result.errors)


def test_register_declares_88_documentary_fields(schema):
    count = sum(
        1
        for block in schema["blocks"].values()
        for field in block["fields"]
        if field.get("route") == "documentary"
    )
    assert count == 88


def test_thirteen_practice_codes(schema):
    codes = [f["name"] for f in schema["blocks"]["F"]["fields"] if f["name"].startswith("pc")]
    assert len(codes) == 13
    assert codes[-1] == "pc13_restoration"


def test_satellite_quantities_are_blocked(schema):
    blocked = {q.lower() for q in schema["satellite_only_quantities"]}
    for name in ("polygon_area_ha", "elevation_m", "koppen_group", "size_class_documentary",
                 "control_distance_km", "reference_circle"):
        assert name in blocked


@pytest.fixture()
def exported(tmp_path, db, community, settings, schema):
    """A minimal but complete community, exported to a workbook."""
    now = utcnow()
    db.insert("sources", {
        "source_id": "IC001-S001", "community_id": community, "address_id": "IC001-01",
        "url": "https://x.org/", "registrable_domain": "x.org", "platform_type": "own website",
        "source_class": "S4", "supplied_or_discovered": "supplied", "independence_group": "G1",
        "crawl_status": "crawled", "access_status": "ok", "pages_opened": 12,
        "first_discovered_utc": now, "last_crawled_utc": now,
    })
    db.insert("searches", {
        "search_id": "Q1", "community_id": community, "database_name": "OpenAlex",
        "database_type": "academic", "query": '"Test"', "language": "en",
        "hits_returned": 0, "result": "none found", "searched_utc": now, "stage": 5,
    })
    for name, value, status in (
        ("community_name_official", "Test Community", "coded"),
        ("managed_area_ha", "15", "coded"),
        ("managed_area_basis", "stated", "coded"),
        ("e5_active_currently", "yes", "coded"),
        ("crawl_truncated", "no", "coded"),
        ("date_intervention_onset", "1992", "coded"),
        ("onset_evidence_rank", "3", "coded"),
        ("pc07_tree_planting", "documented", "coded"),
        ("pc01_rainwater", "not mentioned", "coded"),
        ("tenure_type", None, "not_found"),
    ):
        db.upsert("field_values", {
            "community_id": community, "field_name": name, "value": value, "status": status,
            "method": "rule", "updated_utc": now,
        }, ["community_id", "field_name"])

    exporter = WorkbookExporter(settings.workbook_template, schema, db, coder_id="TEST")
    path = tmp_path / "out.xlsx"
    result = exporter.export(community, path, manifest={"run_mode": "FULL"})
    return path, result


def test_export_opens_and_keeps_every_canonical_sheet(exported):
    path, _ = exported
    workbook = load_workbook(path)
    for sheet in ("O1_Community_Attributes", "O2_Practice_Matrix", "O2b_Practice_Evidence",
                  "O3_Onset_Register", "O5_Disagreement_Log", "O6_Source_Index",
                  "O7_Search_Log", "O10_Polygon_And_Area", "O11_Source_Set",
                  "R1_Codebook", "Reference_Codes", "Cohort_Tracker"):
        assert sheet in workbook.sheetnames
    workbook.close()


def test_export_writes_values_to_the_right_columns(exported, schema):
    path, _ = exported
    workbook = load_workbook(path)
    sheet = workbook["O1_Community_Attributes"]
    assert sheet["A3"].value == "IC001"
    assert sheet["V3"].value == 15          # managed_area_ha
    assert sheet["Y3"].value == "stated"    # managed_area_basis
    assert sheet["O3"].value == "yes"       # e5_active_currently
    assert sheet["BN3"].value == "no"       # crawl_truncated
    assert sheet["BF3"].value == "TEST"     # coder_id, machine identity
    assert sheet["A2"].value is None        # the example row is emptied
    onset = workbook["O3_Onset_Register"]
    assert onset["F3"].value == 1992
    practices = workbook["O2_Practice_Matrix"]
    assert practices["I3"].value == "documented"    # pc07
    assert practices["C3"].value == "not mentioned"  # pc01
    workbook.close()


def test_formula_cells_are_never_overwritten(exported, settings, schema):
    path, result = exported
    template = profile_workbook(settings.workbook_template)
    workbook = load_workbook(path, data_only=False)
    for sheet_name, columns in schema["formula_columns"].items():
        sheet = workbook[sheet_name]
        for column in columns:
            for row in range(3, 7):
                value = sheet[f"{column}{row}"].value
                if value is not None:
                    assert str(value).startswith("="), (
                        f"{sheet_name}!{column}{row} lost its formula")
    workbook.close()


def test_polygon_columns_are_left_to_the_researcher(exported):
    path, _ = exported
    workbook = load_workbook(path)
    sheet = workbook["O10_Polygon_And_Area"]
    assert sheet["A3"].value == "IC001"
    for column in ("D", "E", "F", "G"):     # polygon file, imagery, confidence
        assert sheet[f"{column}3"].value is None
    workbook.close()


def test_supplementary_sheets_are_added_with_a_purpose(exported):
    path, _ = exported
    workbook = load_workbook(path)
    for sheet in ("X1_Evidence_Register", "X3_Image_Evidence", "X4_Document_Register",
                  "X6_Failure_Log", "X8_Review_Queue", "X10_Field_Provenance",
                  "X11_Run_Manifest"):
        assert sheet in workbook.sheetnames
        assert workbook[sheet]["A1"].value      # a header row exists
    workbook.close()


def test_a_value_outside_the_dropdown_is_refused(tmp_path, db, community, settings, schema):
    db.upsert("field_values", {
        "community_id": community, "field_name": "e5_active_currently",
        "value": "perhaps", "status": "coded", "method": "rule", "updated_utc": utcnow(),
    }, ["community_id", "field_name"])
    exporter = WorkbookExporter(settings.workbook_template, schema, db)
    result = exporter.export(community, tmp_path / "o.xlsx", manifest={})
    assert any("perhaps" in refusal for refusal in result.refusals)
    workbook = load_workbook(tmp_path / "o.xlsx")
    assert workbook["O1_Community_Attributes"]["O3"].value is None
    workbook.close()


def test_the_example_row_is_removed(exported):
    path, _ = exported
    workbook = load_workbook(path)
    for sheet_name in ("O1_Community_Attributes", "O3_Onset_Register"):
        values = [c.value for c in workbook[sheet_name][2]]
        assert not any(isinstance(v, str) and "EXAMPLE ROW" in v for v in values)
    workbook.close()


def test_fixture_runs_are_marked_in_the_notes(tmp_path, db, settings, schema):
    now = utcnow()
    db.insert("communities", {
        "community_id": "TEST-IC009", "site_id": "TEST-IC009", "name_input": "Fixture",
        "safe_name": "Fixture", "provenance_mode": "FIXTURE",
        "created_utc": now, "updated_utc": now,
    })
    exporter = WorkbookExporter(settings.workbook_template, schema, db)
    exporter.export("TEST-IC009", tmp_path / "f.xlsx", manifest={})
    workbook = load_workbook(tmp_path / "f.xlsx")
    assert "FIXTURE" in (workbook["O1_Community_Attributes"]["BJ3"].value or "")
    workbook.close()
