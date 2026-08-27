"""The export must not be able to lose a run's work.

The reported failure was real and release-blocking: a run that had gathered its
evidence died at the last step with

    openpyxl.utils.exceptions.IllegalCharacterError

because one piece of extracted PDF text contained a control byte. These tests
reproduce that exactly, and the other ways the same step can fail.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError

from dcr.export.finalise import finalise_workbook, verify_workbook
from dcr.export.sanitize import (MAX_CELL_CHARS, Sanitisation, clean_cell,
                                 clean_text, find_illegal, sanitise_row,
                                 sanitise_workbook)


# ---------------------------------------------------------------------------
# the exact reported crash
# ---------------------------------------------------------------------------
#: Every shape of broken text a PDF extractor has been seen to produce.
BROKEN_TEXT = {
    "null byte": "Nous cultivons 4\x00 hectares.",
    "vertical tab": "planted\x0bin 2016",
    "form feed": "page one\x0cpage two",
    "bell": "restoration\x07report",
    "escape": "swales\x1bdug in 2015",
    "low controls": "text \x01\x02\x03\x04\x05 more",
    "unit separator": "field\x1fvalue",
    "delete": "caption\x7fhere",
    "mixed binary": "PDF\x00\x01 garbage \x0c mixed \x1b with text",
    "unpaired surrogate": "broken \ud800 encoding",
}


def test_openpyxl_really_does_reject_these(  ):
    """The premise. If openpyxl ever stopped raising, these tests would be moot."""
    sheet = Workbook().active
    raised = 0
    for value in BROKEN_TEXT.values():
        try:
            sheet["A1"] = value
        except IllegalCharacterError:
            raised += 1
    assert raised >= 8, "openpyxl no longer rejects control characters"


@pytest.mark.parametrize("label", sorted(BROKEN_TEXT))
def test_every_broken_string_becomes_writable(label):
    value = BROKEN_TEXT[label]
    cleaned, removed, _ = clean_cell(value)
    sheet = Workbook().active
    sheet["A1"] = cleaned            # must not raise
    assert removed > 0
    assert not find_illegal(cleaned)


@pytest.mark.parametrize("label", sorted(BROKEN_TEXT))
def test_every_broken_string_survives_being_saved(label, tmp_path):
    """The surrogate case only fails at save(), after every sheet is written."""
    cleaned, _, _ = clean_cell(BROKEN_TEXT[label])
    workbook = Workbook()
    workbook.active["A1"] = cleaned
    path = tmp_path / "out.xlsx"
    workbook.save(path)              # must not raise UnicodeEncodeError
    assert load_workbook(path).active["A1"].value is not None


# ---------------------------------------------------------------------------
# the evidence must survive the cleaning
# ---------------------------------------------------------------------------
def test_the_readable_words_are_kept():
    cleaned, _, _ = clean_cell("Nous cultivons 4\x00 hectares en maraichage.")
    assert "Nous cultivons 4" in cleaned
    assert "hectares en maraichage." in cleaned


def test_a_control_character_becomes_a_space_not_a_join():
    """A form feed separated two words; deleting it would run them together."""
    cleaned, _, _ = clean_cell("page one\x0cpage two")
    assert "one page" in cleaned or "one  page" in cleaned
    assert "onepage" not in cleaned


def test_tabs_and_newlines_are_left_alone():
    """A multi-line quotation must stay multi-line."""
    value = "line one\nline two\tcolumn"
    cleaned, removed, _ = clean_cell(value)
    assert cleaned == value
    assert removed == 0


def test_accented_and_non_latin_text_is_untouched():
    for value in ("forêt-jardin", "Boekel Ecodorp", "Tamera — Portugal",
                  "восстановление", "农业", "café"):
        cleaned, removed, _ = clean_cell(value)
        assert cleaned == value, value
        assert removed == 0


def test_numbers_are_not_turned_into_text():
    """Coercion to text would silently drop the value out of O4's calculations."""
    for value in (4, 4.2, 2016, True, None):
        cleaned, removed, _ = clean_cell(value)
        assert cleaned is value
        assert removed == 0


def test_a_giant_cell_is_truncated_not_refused():
    cleaned, _, truncated = clean_cell("x" * 60000)
    assert truncated
    assert len(cleaned) <= MAX_CELL_CHARS
    Workbook().active["A1"] = cleaned


def test_aggressive_mode_is_more_destructive_and_so_is_not_the_default():
    value = "text with an astral char \U0001F600 and more"
    normal, normal_removed = clean_text(value)
    harsh, harsh_removed = clean_text(value, aggressive=True)
    assert normal == value and normal_removed == 0
    assert harsh_removed > 0


# ---------------------------------------------------------------------------
# sanitisation is recorded, never silent
# ---------------------------------------------------------------------------
def test_sanitisation_is_counted_and_reported():
    log = Sanitisation()
    sanitise_row(["clean", "dirty\x00text", 42], sheet="X1_Evidence_Register", log=log)
    assert log.occurred
    assert log.cells == 1
    assert log.characters == 1
    assert log.by_sheet["X1_Evidence_Register"] == 1
    assert log.as_dict()["excel_sanitized"] == "yes"
    assert "database" in log.summary(), "the summary must say where the raw text still is"


def test_a_clean_run_reports_no_sanitisation():
    log = Sanitisation()
    sanitise_row(["clean", "also clean", 7], sheet="X1", log=log)
    assert not log.occurred
    assert log.as_dict()["excel_sanitized"] == "no"


def test_the_whole_workbook_sweep_catches_what_a_writer_missed(tmp_path):
    """The safety net: it does not care which code path wrote the cell."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "fine"
    sheet["A2"] = "=SUM(B1:B9)"
    # Bypass openpyxl's own guard the way a direct _cells write would.
    from openpyxl.cell.cell import Cell
    cell = Cell(sheet, row=3, column=1)
    cell._value = "sneaky\x00value"
    cell.data_type = "s"
    sheet._cells[(3, 1)] = cell

    log = sanitise_workbook(workbook)
    assert log.occurred
    assert sheet["A2"].value == "=SUM(B1:B9)", "a formula must not be treated as text"
    path = tmp_path / "swept.xlsx"
    workbook.save(path)
    assert load_workbook(path)["Sheet"]["A3"].value == "sneaky value"


# ---------------------------------------------------------------------------
# verification: writing a file is not evidence the file is good
# ---------------------------------------------------------------------------
def test_a_missing_workbook_does_not_verify(tmp_path):
    result = verify_workbook(tmp_path / "never-written.xlsx")
    assert not result.ok
    assert "not written" in result.problems[0]


def test_a_zero_byte_workbook_does_not_verify(tmp_path):
    path = tmp_path / "empty.xlsx"
    path.write_bytes(b"")
    result = verify_workbook(path)
    assert not result.ok
    assert "zero bytes" in result.problems[0]


def test_a_corrupt_workbook_does_not_verify(tmp_path):
    path = tmp_path / "corrupt.xlsx"
    path.write_bytes(b"this is not a spreadsheet")
    result = verify_workbook(path)
    assert not result.ok
    assert "could not be reopened" in result.problems[0]


def test_a_workbook_without_the_core_sheets_does_not_verify(tmp_path):
    path = tmp_path / "wrong.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "=1+1"
    workbook.save(path)
    result = verify_workbook(path)
    assert not result.ok
    assert any("missing core sheet" in problem for problem in result.problems)


# ---------------------------------------------------------------------------
# the retry ladder
# ---------------------------------------------------------------------------
class _FakeExporter:
    """An exporter that fails in a chosen way until a chosen rung is reached."""

    def __init__(self, *, fail_until: str, aggressive: bool, core_only: bool,
                 error: Exception | None = None):
        self.fail_until = fail_until
        self.aggressive = aggressive
        self.core_only = core_only
        self.error = error or IllegalCharacterError()
        self.sanitisation = Sanitisation()
        self.omitted_sheets: dict[str, str] = {}
        self.rows_written = {"O1_Community_Attributes": 1}
        self.refusals: list[str] = []
        self.warnings: list[str] = []

    def _reached(self) -> bool:
        """Has the ladder climbed far enough for this exporter to succeed?"""
        if self.fail_until == "never":
            return False
        if self.fail_until == "normal":
            return True
        if self.fail_until == "aggressive":
            return self.aggressive
        return self.core_only

    def export(self, community_id, destination, manifest=None):
        if not self._reached():
            raise self.error
        workbook = Workbook()
        first = workbook.active
        first.title = "O1_Community_Attributes"
        first["A1"] = "site_id"
        first["A3"] = community_id
        first["B3"] = "=1+1"
        for name in ("O2_Practice_Matrix", "O3_Onset_Register", "O6_Source_Index",
                     "O11_Source_Set"):
            sheet = workbook.create_sheet(name)
            sheet["A3"] = community_id
        if self.core_only:
            self.omitted_sheets["X1_Evidence_Register"] = "omitted deliberately"
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return self


def _factory(fail_until: str, error: Exception | None = None):
    made: list[_FakeExporter] = []

    def build(*, aggressive: bool = False, core_only: bool = False):
        exporter = _FakeExporter(fail_until=fail_until, aggressive=aggressive,
                                 core_only=core_only, error=error)
        made.append(exporter)
        return exporter

    return build, made


def test_a_clean_export_needs_only_one_attempt(tmp_path):
    build, _ = _factory("normal")
    result = finalise_workbook(exporter_factory=build, community_id="IC001",
                               destination=tmp_path / "wb.xlsx")
    assert result.ok
    assert result.attempts == 1
    assert result.strategy == "normal"


def test_an_illegal_character_crash_is_recovered_by_retrying(tmp_path):
    """The exact reported failure, at the exact point it happened."""
    build, made = _factory("aggressive")
    result = finalise_workbook(exporter_factory=build, community_id="IC001",
                               destination=tmp_path / "wb.xlsx")
    assert result.ok, "the run must not end without a workbook"
    assert result.attempts == 2
    assert result.strategy == "aggressive-sanitisation"
    assert made[1].aggressive
    assert result.repairs, "the repair must be recorded, not silent"
    assert load_workbook(result.verification.path) is not None


def test_a_malformed_supplementary_dataset_still_yields_the_core_workbook(tmp_path):
    build, made = _factory("core_only")
    result = finalise_workbook(exporter_factory=build, community_id="IC001",
                               destination=tmp_path / "wb.xlsx")
    assert result.ok
    assert result.attempts == 3
    assert result.strategy == "core-workbook-only"
    assert made[2].core_only
    workbook = load_workbook(result.verification.path)
    assert "O1_Community_Attributes" in workbook.sheetnames
    assert result.as_dict()["omitted_sheets"]


def test_a_run_that_cannot_export_at_all_says_so_rather_than_crashing(tmp_path):
    build, _ = _factory("never")
    result = finalise_workbook(exporter_factory=build, community_id="IC001",
                               destination=tmp_path / "wb.xlsx")
    assert not result.ok
    assert result.failed
    assert result.failure_reason
    assert result.attempts == 3          # the ladder was tried in full


def test_a_failed_attempt_leaves_no_half_written_file_behind(tmp_path):
    """Otherwise the next attempt could 'verify' the corpse of the last one."""
    destination = tmp_path / "wb.xlsx"
    destination.write_bytes(b"stale partial file")
    build, _ = _factory("never")
    finalise_workbook(exporter_factory=build, community_id="IC001",
                      destination=destination)
    assert not destination.exists()


@pytest.mark.parametrize("error", [
    IllegalCharacterError(),
    UnicodeEncodeError("utf-8", "x", 0, 1, "surrogates not allowed"),
    ValueError("cannot convert"),
])
def test_the_ladder_recovers_from_any_export_exception(tmp_path, error):
    build, _ = _factory("aggressive", error=error)
    result = finalise_workbook(exporter_factory=build, community_id="IC001",
                               destination=tmp_path / "wb.xlsx")
    assert result.ok
