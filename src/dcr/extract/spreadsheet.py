"""Spreadsheet and CSV extraction.

Spreadsheets are not flattened into text (brief §21). Every worksheet, header
row and cell coordinate is preserved, so a value extracted from a workbook can
be cited as ``Sheet1!B7`` rather than as "somewhere in the file".
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from ..logging_setup import get_logger

log = get_logger("spreadsheet")

MAX_ROWS = 5000
MAX_COLS = 200


@dataclass
class SheetTable:
    sheet_name: str
    rows: list[list[str]]
    header: list[str] = field(default_factory=list)
    first_row: int = 1
    first_col: int = 1
    hidden: bool = False

    @property
    def cell_range(self) -> str:
        if not self.rows:
            return ""
        return f"{_col_letter(self.first_col)}{self.first_row}:" \
               f"{_col_letter(self.first_col + max(len(r) for r in self.rows) - 1)}" \
               f"{self.first_row + len(self.rows) - 1}"

    def cell_reference(self, row_index: int, col_index: int) -> str:
        return f"{self.sheet_name}!{_col_letter(self.first_col + col_index)}{self.first_row + row_index}"


@dataclass
class SpreadsheetResult:
    ok: bool = False
    tables: list[SheetTable] = field(default_factory=list)
    text: str = ""
    metadata: dict[str, str] = field(default_factory=dict)
    parser: str = ""
    parser_status: str = "not_attempted"
    text_status: str = "not_attempted"
    table_status: str = "not_attempted"
    detail: str = ""
    sheet_names: list[str] = field(default_factory=list)
    hidden_sheets: list[str] = field(default_factory=list)


def _col_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"


def extract_xlsx(data: bytes) -> SpreadsheetResult:
    result = SpreadsheetResult(parser="openpyxl")
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        result.parser_status = "unsupported_format"
        result.detail = "openpyxl is not installed"
        return result
    try:
        # data_only reads the cached results of formulas, which is what a
        # published workbook's numbers actually are.
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        result.parser_status = "corrupt"
        result.detail = f"{type(exc).__name__}: {exc}"
        return result

    try:
        properties = workbook.properties
        for key in ("title", "creator", "created", "modified", "subject", "description"):
            value = getattr(properties, key, None)
            if value:
                result.metadata[key] = str(value)[:500]
    except Exception:
        pass

    try:
        # Hidden sheets are inspected where technically accessible; a hidden
        # sheet often holds the working figures behind a published summary.
        for worksheet in workbook.worksheets:
            result.sheet_names.append(worksheet.title)
            hidden = getattr(worksheet, "sheet_state", "visible") != "visible"
            if hidden:
                result.hidden_sheets.append(worksheet.title)
            rows: list[list[str]] = []
            for row in worksheet.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True):
                cells = _trim(["" if cell is None else str(cell).strip() for cell in row])
                if cells:
                    rows.append(cells)
            if not rows:
                continue
            header = rows[0] if _looks_like_header(rows) else []
            result.tables.append(
                SheetTable(sheet_name=worksheet.title, rows=rows, header=header, hidden=hidden)
            )
    except Exception as exc:
        result.parser_status = "corrupt"
        result.detail = f"{type(exc).__name__}: {exc}"
        return result
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return _finish(result)


def extract_xls(data: bytes) -> SpreadsheetResult:
    result = SpreadsheetResult(parser="xlrd")
    try:
        import xlrd  # type: ignore
    except ImportError:
        result.parser_status = "unsupported_format"
        result.detail = "xlrd is not installed; the original .xls is preserved unparsed"
        return result
    try:
        book = xlrd.open_workbook(file_contents=data)
    except Exception as exc:
        result.parser_status = "corrupt"
        result.detail = f"{type(exc).__name__}: {exc}"
        return result
    for sheet in book.sheets():
        result.sheet_names.append(sheet.name)
        rows: list[list[str]] = []
        for index in range(min(sheet.nrows, MAX_ROWS)):
            cells = ["" if v is None else str(v).strip()
                     for v in sheet.row_values(index)[:MAX_COLS]]
            if any(cells):
                rows.append(cells)
        if rows:
            header = rows[0] if _looks_like_header(rows) else []
            result.tables.append(SheetTable(sheet_name=sheet.name, rows=rows, header=header))
    return _finish(result)


def extract_csv(data: bytes, *, filename: str = "data.csv") -> SpreadsheetResult:
    result = SpreadsheetResult(parser="csv")
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        result.parser_status = "corrupt"
        result.detail = "could not decode the file in UTF-8 or Latin-1"
        return result

    sample = text[:8000]
    try:
        dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
        if text.count(";") > text.count(","):
            dialect.delimiter = ";"
        elif text.count("\t") > text.count(","):
            dialect.delimiter = "\t"
    rows: list[list[str]] = []
    try:
        for row in csv.reader(io.StringIO(text), dialect):
            cells = [c.strip() for c in row[:MAX_COLS]]
            if any(cells):
                rows.append(cells)
            if len(rows) >= MAX_ROWS:
                break
    except csv.Error as exc:
        result.parser_status = "corrupt"
        result.detail = f"malformed CSV: {exc}"
        return result

    if rows:
        sheet = filename.rsplit("/", 1)[-1]
        header = rows[0] if _looks_like_header(rows) else []
        result.tables.append(SheetTable(sheet_name=sheet, rows=rows, header=header))
        result.sheet_names.append(sheet)
    return _finish(result)


def _trim(cells: list[str]) -> list[str]:
    """Drop the trailing empty cells a fixed column window pads a row with."""
    end = len(cells)
    while end > 0 and not cells[end - 1]:
        end -= 1
    return cells[:end]


def _looks_like_header(rows: list[list[str]]) -> bool:
    """A first row of labels above rows of values."""
    if len(rows) < 2:
        return False
    first, second = rows[0], rows[1]
    if not first or not any(first):
        return False
    numeric_first = sum(1 for c in first if _is_number(c))
    numeric_second = sum(1 for c in second if _is_number(c))
    return numeric_first == 0 and numeric_second > 0


def _is_number(value: str) -> bool:
    return bool(re.fullmatch(r"[-+]?\d{1,3}(?:[ .,]\d{3})*(?:[.,]\d+)?%?", (value or "").strip()))


def _finish(result: SpreadsheetResult) -> SpreadsheetResult:
    if result.tables:
        result.parser_status = "parsed"
        result.table_status = "extracted"
        pieces = []
        for table in result.tables:
            pieces.append(f"[sheet {table.sheet_name}{' (hidden)' if table.hidden else ''}]")
            for row in table.rows[:500]:
                pieces.append(" | ".join(row))
        result.text = "\n".join(pieces)
        result.text_status = "extracted"
        result.ok = True
    else:
        result.parser_status = result.parser_status if result.parser_status != "not_attempted" else "parsed"
        result.table_status = "none_found"
        result.text_status = "empty"
        result.ok = result.parser_status == "parsed"
    return result
