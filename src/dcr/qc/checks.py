"""Quality control before a community may be called complete.

The eighteen checks in brief §55, plus the coverage matrix of §57. Each returns
a verdict with the evidence behind it, and no community reaches ``COMPLETE``
with a failing check.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence

from ..db import Database

PASS = "pass"
FAIL = "fail"
WARN = "warn"


@dataclass
class CheckResult:
    number: int
    name: str
    verdict: str
    detail: str
    evidence: str = ""


@dataclass
class QcReport:
    results: list[CheckResult] = field(default_factory=list)
    coverage: list[dict[str, Any]] = field(default_factory=list)

    @property
    def failures(self) -> list[CheckResult]:
        return [r for r in self.results if r.verdict == FAIL]

    @property
    def warnings(self) -> list[CheckResult]:
        return [r for r in self.results if r.verdict == WARN]

    @property
    def ok(self) -> bool:
        return not self.failures


class QualityControl:
    def __init__(self, db: Database, community_id: str, schema: Mapping[str, Any],
                 storage_root: Path | None = None):
        self.db = db
        self.community_id = community_id
        self.schema = schema
        self.storage_root = storage_root

    def run(self, *, workbook_path: Path | None = None) -> QcReport:
        report = QcReport()
        checks: list[tuple[int, str, Callable[[], CheckResult]]] = [
            (1, "Every supplied URL was processed or explicitly marked failed", self._c1),
            (2, "Discovered sources are recorded", self._c2),
            (3, "Every source has a source class", self._c3),
            (4, "Every source has an independence group", self._c4),
            (5, "No field contains an invented value", self._c5),
            (6, "All citations point to retrieved material", self._c6),
            (7, "Academic records are individually verified", self._c7),
            (8, "All documents have hashes", self._c8),
            (9, "All image files have provenance", self._c9),
            (10, "Search stages are logged", self._c10),
            (11, "Negative consultations are recorded", self._c11),
            (12, "Truncation status is explicit", self._c12),
            (13, "Conflicts are not silently erased", self._c13),
            (14, "Workbook values validate against allowed values", self._c14),
            (15, "Source IDs are internally consistent", self._c15),
            (16, "All files can be traced back to a source", self._c16),
            (17, "The final workbook opens successfully", lambda: self._c17(workbook_path)),
            (18, "The exported evidence manifest is internally consistent", self._c18),
        ]
        for number, name, check in checks:
            try:
                result = check()
            except Exception as exc:  # a broken check must not hide the others
                result = CheckResult(number, name, FAIL, f"the check itself failed: {exc}")
            result.number, result.name = number, name
            report.results.append(result)
        report.coverage = self.coverage_matrix()
        return report

    # -- individual checks -------------------------------------------------
    def _c1(self) -> CheckResult:
        rows = self.db.query(
            "SELECT source_id, url, crawl_status, access_status FROM sources "
            "WHERE community_id=? AND supplied_or_discovered='supplied'", (self.community_id,))
        unhandled = [r for r in rows
                     if (r["crawl_status"] or "not attempted") == "not attempted"
                     and (r["access_status"] or "not_attempted") == "not_attempted"]
        if not rows:
            return CheckResult(1, "", WARN, "no addresses were supplied by the researcher")
        if unhandled:
            return CheckResult(
                1, "", FAIL,
                f"{len(unhandled)} supplied address(es) end at 'not attempted'",
                "; ".join(r["url"] for r in unhandled[:5]))
        return CheckResult(1, "", PASS,
                           f"all {len(rows)} supplied addresses reached a recorded outcome")

    def _c2(self) -> CheckResult:
        count = self.db.scalar(
            "SELECT COUNT(*) FROM sources WHERE community_id=? AND supplied_or_discovered='discovered'",
            (self.community_id,)) or 0
        attempts = self.db.scalar(
            "SELECT COUNT(*) FROM discovery_log WHERE community_id=?", (self.community_id,)) or 0
        if count == 0 and attempts == 0:
            return CheckResult(2, "", FAIL,
                               "no discovery was attempted, so 'none found' cannot be trusted")
        return CheckResult(2, "", PASS,
                           f"{count} discovered addresses recorded from {attempts} discovery events")

    def _c3(self) -> CheckResult:
        missing = self.db.query(
            "SELECT source_id FROM sources WHERE community_id=? AND (source_class IS NULL "
            "OR source_class='')", (self.community_id,))
        if missing:
            return CheckResult(3, "", FAIL, f"{len(missing)} sources have no class",
                               "; ".join(r["source_id"] for r in missing[:5]))
        return CheckResult(3, "", PASS, "every source carries an S1-S8 class")

    def _c4(self) -> CheckResult:
        missing = self.db.query(
            "SELECT source_id FROM sources WHERE community_id=? AND (independence_group IS NULL "
            "OR independence_group='')", (self.community_id,))
        if missing:
            return CheckResult(4, "", FAIL, f"{len(missing)} sources have no independence group",
                               "; ".join(r["source_id"] for r in missing[:5]))
        groups = self.db.scalar(
            "SELECT COUNT(DISTINCT independence_group) FROM sources WHERE community_id=?",
            (self.community_id,)) or 0
        total = self.db.scalar(
            "SELECT COUNT(*) FROM sources WHERE community_id=?", (self.community_id,)) or 0
        return CheckResult(4, "", PASS,
                           f"{total} addresses fall into {groups} independence groups")

    def _c5(self) -> CheckResult:
        """Every coded value must trace to a claim, a rule or a researcher input."""
        rows = self.db.query(
            "SELECT field_name, value, method, claim_ids, source_ids FROM field_values "
            "WHERE community_id=? AND status='coded'", (self.community_id,))
        rule_methods = {
            "rule", "derived", "researcher_input", "crawl_audit", "run_state", "search_log",
            "independence", "activity_rules", "practice_rules", "onset_engine", "extraction",
            "image_evidence", "grey_literature",
        }
        orphans = [r for r in rows
                   if not r["claim_ids"] and not r["source_ids"] and r["method"] not in rule_methods]
        if orphans:
            return CheckResult(5, "", FAIL,
                               f"{len(orphans)} coded values have neither a claim nor a rule behind them",
                               "; ".join(r["field_name"] for r in orphans[:5]))
        forbidden = set(str(q).lower() for q in self.schema.get("satellite_only_quantities", []))
        offenders = [r["field_name"] for r in rows if r["field_name"].lower() in forbidden]
        if offenders:
            return CheckResult(5, "", FAIL,
                               "a satellite-only quantity was coded documentarily",
                               "; ".join(offenders))
        return CheckResult(5, "", PASS,
                           f"all {len(rows)} coded values trace to a claim or a stated rule")

    def _c6(self) -> CheckResult:
        claims = self.db.query(
            "SELECT claim_id, source_id FROM claims WHERE community_id=? AND source_id IS NOT NULL",
            (self.community_id,))
        known = {r["source_id"] for r in self.db.query(
            "SELECT source_id FROM sources WHERE community_id=?", (self.community_id,))}
        dangling = [c for c in claims if c["source_id"] not in known]
        if dangling:
            return CheckResult(6, "", FAIL,
                               f"{len(dangling)} claims cite a source that is not in the source index",
                               "; ".join(c["claim_id"] for c in dangling[:5]))
        unopened = self.db.query(
            "SELECT s.source_id FROM sources s WHERE s.community_id=? AND s.pages_opened = 0 "
            "AND EXISTS (SELECT 1 FROM claims c WHERE c.source_id = s.source_id "
            "AND c.field_name NOT LIKE 'pc%')", (self.community_id,))
        if unopened:
            return CheckResult(6, "", WARN,
                               f"{len(unopened)} sources support a claim but recorded no opened page",
                               "; ".join(r["source_id"] for r in unopened[:5]))
        return CheckResult(6, "", PASS, f"all {len(claims)} attributed claims cite a real source")

    def _c7(self) -> CheckResult:
        total = self.db.scalar(
            "SELECT COUNT(*) FROM academic_records WHERE community_id=?", (self.community_id,)) or 0
        if total == 0:
            searched = self.db.scalar(
                "SELECT COUNT(*) FROM searches WHERE community_id=? AND database_type IN "
                "('academic','thesis portal')", (self.community_id,)) or 0
            if searched == 0:
                return CheckResult(7, "", FAIL,
                                   "no academic database was consulted, so 'none found' is not a finding")
            return CheckResult(7, "", PASS,
                               f"no academic records found after {searched} database consultations "
                               "— the expected result for most communities")
        unverified = self.db.query(
            "SELECT record_id, title FROM academic_records WHERE community_id=? "
            "AND verified_resolves != 'yes'", (self.community_id,))
        supporting = self.db.scalar(
            "SELECT COUNT(*) FROM claims WHERE community_id=? AND source_class='S1'",
            (self.community_id,)) or 0
        if unverified and supporting:
            return CheckResult(7, "", FAIL,
                               f"{len(unverified)} unverified academic records exist while "
                               f"{supporting} claims rest on S1 sources",
                               "; ".join(r["title"][:60] for r in unverified[:3]))
        if unverified:
            return CheckResult(7, "", WARN,
                               f"{len(unverified)} of {total} academic records are unverified and "
                               "are barred from supporting any value")
        return CheckResult(7, "", PASS, f"all {total} academic records verified")

    def _c8(self) -> CheckResult:
        missing = self.db.query(
            "SELECT document_id FROM documents WHERE community_id=? AND (sha256 IS NULL OR sha256='')",
            (self.community_id,))
        total = self.db.scalar(
            "SELECT COUNT(*) FROM documents WHERE community_id=?", (self.community_id,)) or 0
        if missing:
            return CheckResult(8, "", FAIL, f"{len(missing)} documents have no hash",
                               "; ".join(r["document_id"] for r in missing[:5]))
        return CheckResult(8, "", PASS, f"all {total} documents carry a sha256")

    def _c9(self) -> CheckResult:
        orphans = self.db.query(
            "SELECT image_id FROM images WHERE community_id=? AND source_id IS NULL "
            "AND document_id IS NULL AND page_id IS NULL", (self.community_id,))
        total = self.db.scalar(
            "SELECT COUNT(*) FROM images WHERE community_id=?", (self.community_id,)) or 0
        no_url = self.db.query(
            "SELECT image_id FROM images WHERE community_id=? AND (original_url IS NULL "
            "OR original_url='')", (self.community_id,))
        if orphans or no_url:
            return CheckResult(9, "", FAIL,
                               f"{len(orphans)} images have no source and {len(no_url)} have no URL",
                               "; ".join(r["image_id"] for r in (orphans + no_url)[:5]))
        return CheckResult(9, "", PASS, f"all {total} images trace to a source, page or document")

    def _c10(self) -> CheckResult:
        stages = self.db.query(
            "SELECT DISTINCT stage FROM searches WHERE community_id=?", (self.community_id,))
        total = self.db.scalar(
            "SELECT COUNT(*) FROM searches WHERE community_id=?", (self.community_id,)) or 0
        if total == 0:
            return CheckResult(10, "", FAIL, "no search was logged at all")
        return CheckResult(10, "", PASS,
                           f"{total} consultations logged across stages "
                           f"{sorted(r['stage'] for r in stages if r['stage'] is not None)}")

    def _c11(self) -> CheckResult:
        negatives = self.db.scalar(
            "SELECT COUNT(*) FROM searches WHERE community_id=? AND result IN "
            "('none found','unreachable')", (self.community_id,)) or 0
        field_row = self.db.query_one(
            "SELECT value FROM field_values WHERE community_id=? AND field_name='negative_consultations'",
            (self.community_id,))
        if negatives and not (field_row and field_row["value"]):
            return CheckResult(11, "", FAIL,
                               f"{negatives} empty or unreachable consultations were not carried "
                               "into negative_consultations")
        return CheckResult(11, "", PASS,
                           f"{negatives} negative consultations recorded and carried to the workbook")

    def _c12(self) -> CheckResult:
        row = self.db.query_one(
            "SELECT value, rationale FROM field_values WHERE community_id=? "
            "AND field_name='crawl_truncated'", (self.community_id,))
        if row is None or row["value"] not in ("yes", "no"):
            return CheckResult(12, "", FAIL, "crawl_truncated is not set to yes or no")
        return CheckResult(12, "", PASS,
                           f"crawl_truncated = {row['value']}", (row["rationale"] or "")[:300])

    def _c13(self) -> CheckResult:
        conflicts = self.db.scalar(
            "SELECT COUNT(*) FROM conflicts WHERE community_id=?", (self.community_id,)) or 0
        unresolved = self.db.scalar(
            "SELECT COUNT(*) FROM conflicts WHERE community_id=? AND resolution_type='unresolved'",
            (self.community_id,)) or 0
        contradicted = self.db.query(
            "SELECT field_name, COUNT(DISTINCT value) AS n FROM claims WHERE community_id=? "
            "GROUP BY field_name HAVING n > 1", (self.community_id,))
        recorded_fields = {r["field_name"] for r in self.db.query(
            "SELECT DISTINCT field_name FROM conflicts WHERE community_id=?", (self.community_id,))}
        missing = [r["field_name"] for r in contradicted
                   if r["field_name"] not in recorded_fields
                   and not r["field_name"].startswith(("pc", "notable_context",
                                                       "e1_self_identification",
                                                       "documentary_area_note",
                                                       "external_funding_or_programme"))]
        if missing:
            return CheckResult(13, "", WARN,
                               f"{len(missing)} fields hold differing claims with no conflict row",
                               "; ".join(missing[:5]))
        return CheckResult(13, "", PASS,
                           f"{conflicts} conflicts recorded, {unresolved} left for a human")

    def _c14(self) -> CheckResult:
        problems: list[str] = []
        for block in self.schema["blocks"].values():
            for field_def in block["fields"]:
                allowed = field_def.get("allowed_values")
                if not allowed:
                    continue
                row = self.db.query_one(
                    "SELECT value FROM field_values WHERE community_id=? AND field_name=? "
                    "AND status='coded'", (self.community_id, field_def["name"]))
                if row and row["value"] and str(row["value"]) not in [str(a) for a in allowed]:
                    problems.append(f"{field_def['name']}={row['value']!r}")
        if problems:
            return CheckResult(14, "", FAIL,
                               f"{len(problems)} values fall outside their allowed set",
                               "; ".join(problems[:5]))
        return CheckResult(14, "", PASS, "every controlled value is inside its allowed set")

    def _c15(self) -> CheckResult:
        rows = self.db.query(
            "SELECT source_id, address_id FROM sources WHERE community_id=?", (self.community_id,))
        bad = [r["source_id"] for r in rows
               if not str(r["source_id"]).startswith(self.community_id)
               or not str(r["address_id"]).startswith(self.community_id)]
        if bad:
            return CheckResult(15, "", FAIL, f"{len(bad)} identifiers do not match the community id",
                               "; ".join(bad[:5]))
        duplicates = self.db.query(
            "SELECT source_id, COUNT(*) AS n FROM sources WHERE community_id=? "
            "GROUP BY source_id HAVING n > 1", (self.community_id,))
        if duplicates:
            return CheckResult(15, "", FAIL, f"{len(duplicates)} duplicate source ids")
        return CheckResult(15, "", PASS, f"{len(rows)} identifiers are consistent and unique")

    def _c16(self) -> CheckResult:
        if self.storage_root is None:
            return CheckResult(16, "", WARN, "no storage root supplied to the check")
        missing: list[str] = []
        for row in self.db.query(
            "SELECT document_id, storage_path FROM documents WHERE community_id=?",
            (self.community_id,)
        ):
            if not row["storage_path"] or not (self.storage_root / row["storage_path"]).exists():
                missing.append(row["document_id"])
        for row in self.db.query(
            "SELECT image_id, local_path FROM images WHERE community_id=?", (self.community_id,)
        ):
            if not row["local_path"] or not (self.storage_root / row["local_path"]).exists():
                missing.append(row["image_id"])
        if missing:
            return CheckResult(16, "", FAIL,
                               f"{len(missing)} recorded artefacts are missing from disk",
                               "; ".join(missing[:5]))
        return CheckResult(16, "", PASS, "every recorded artefact is present on disk")

    def _c17(self, workbook_path: Path | None) -> CheckResult:
        if workbook_path is None:
            return CheckResult(17, "", WARN, "no workbook was exported in this run")
        if not workbook_path.exists():
            return CheckResult(17, "", FAIL, f"{workbook_path} does not exist")
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path, data_only=False)
            sheets = list(workbook.sheetnames)
            workbook.close()
        except Exception as exc:
            return CheckResult(17, "", FAIL, f"the workbook does not open: {exc}")
        required = ["O1_Community_Attributes", "O2_Practice_Matrix", "O3_Onset_Register",
                    "O6_Source_Index", "O7_Search_Log", "O11_Source_Set"]
        missing = [s for s in required if s not in sheets]
        if missing:
            return CheckResult(17, "", FAIL, f"the workbook is missing {missing}")
        return CheckResult(17, "", PASS, f"the workbook opens with {len(sheets)} sheets")

    def _c18(self) -> CheckResult:
        evidence_ids = {r["evidence_id"] for r in self.db.query(
            "SELECT evidence_id FROM evidence WHERE community_id=?", (self.community_id,))}
        dangling = [r["claim_id"] for r in self.db.query(
            "SELECT claim_id, evidence_id FROM claims WHERE community_id=? "
            "AND evidence_id IS NOT NULL", (self.community_id,))
            if r["evidence_id"] not in evidence_ids]
        empty_quotes = self.db.scalar(
            "SELECT COUNT(*) FROM evidence WHERE community_id=? AND (quote IS NULL OR quote='')",
            (self.community_id,)) or 0
        if dangling or empty_quotes:
            return CheckResult(18, "", FAIL,
                               f"{len(dangling)} claims point at a missing evidence row and "
                               f"{empty_quotes} evidence rows carry no wording")
        return CheckResult(18, "", PASS,
                           f"{len(evidence_ids)} evidence rows, all with wording, all referenced "
                           "consistently")

    # -- coverage ----------------------------------------------------------
    def coverage_matrix(self) -> list[dict[str, Any]]:
        """Source class × searched / found / opened / yielded / failed (brief §57)."""
        rows: list[dict[str, Any]] = []
        for source_class in self.schema.get("source_classes", {}):
            found = self.db.scalar(
                "SELECT COUNT(*) FROM sources WHERE community_id=? AND source_class=?",
                (self.community_id, source_class)) or 0
            opened = self.db.scalar(
                "SELECT COUNT(*) FROM sources WHERE community_id=? AND source_class=? "
                "AND pages_opened > 0", (self.community_id, source_class)) or 0
            yielded = self.db.scalar(
                "SELECT COUNT(DISTINCT source_id) FROM claims WHERE community_id=? "
                "AND source_class=?", (self.community_id, source_class)) or 0
            failed = self.db.scalar(
                "SELECT COUNT(*) FROM sources WHERE community_id=? AND source_class=? "
                "AND crawl_status IN ('blocked','dead link')",
                (self.community_id, source_class)) or 0
            searched = found
            if source_class == "S1":
                searched = self.db.scalar(
                    "SELECT COUNT(*) FROM searches WHERE community_id=? AND database_type IN "
                    "('academic','thesis portal')", (self.community_id,)) or 0
                found = self.db.scalar(
                    "SELECT COUNT(*) FROM academic_records WHERE community_id=?",
                    (self.community_id,)) or 0
                opened = self.db.scalar(
                    "SELECT COUNT(*) FROM academic_records WHERE community_id=? "
                    "AND full_text_status='full text'", (self.community_id,)) or 0
                yielded = self.db.scalar(
                    "SELECT COUNT(*) FROM academic_records WHERE community_id=? "
                    "AND verified_resolves='yes'", (self.community_id,)) or 0
            elif source_class == "S2":
                searched = self.db.scalar(
                    "SELECT COUNT(*) FROM searches WHERE community_id=? AND database_type LIKE "
                    "'grey%' OR (community_id=? AND database_type='registry')",
                    (self.community_id, self.community_id)) or 0
            elif source_class == "S5":
                searched = self.db.scalar(
                    "SELECT COUNT(*) FROM searches WHERE community_id=? AND database_type='archive'",
                    (self.community_id,)) or 0
                found = self.db.scalar(
                    "SELECT COUNT(*) FROM pages WHERE community_id=? AND archive_timestamp IS NOT NULL",
                    (self.community_id,)) or 0
                opened = found
            rows.append({
                "source_class": source_class,
                "label": self.schema["source_classes"][source_class]["label"],
                "searched": searched,
                "found": found,
                "opened": opened,
                "evidence_yielded": yielded,
                "failed_or_blocked": failed,
            })
        return rows


def completion_status(report: QcReport, *, truncated: bool, blocking_review: int,
                      pages_opened: int, min_pages: int,
                      workbook_verified: bool = True,
                      budget_exhausted: bool = False,
                      retrieval_stop_cause: str = "",
                      blocked_sources: int = 0,
                      reachable_sources: int = 0) -> str:
    """Exactly one status, from the six the brief names, and never a flattering
    one (brief §92).

        COMPLETE                    every stage ran, nothing outstanding
        COMPLETE_WITH_UNCERTAINTY   quality warnings a coder should read
        COMPLETE_WITH_TRUNCATION    usable, with parts deliberately not reached
        PARTIAL_BLOCKED             sources refused access
        REQUIRES_HUMAN_REVIEW       something a machine must not decide
        FAILED_TECHNICALLY          no verified workbook

    Three distinctions this has to keep straight.

    **A workbook that cannot be reopened is not an output.** However much
    evidence was gathered, if finalisation could not produce a verified file the
    run failed technically and says so (brief §4, §92).

    **Complete is not exhaustive — but exhausted is not truncated either.**
    This is the change the yield governor brings. A run that stopped because
    every source had gone quiet followed the protocol to its end on the
    evidence, and is COMPLETE. A run stopped by a configured ceiling, or by the
    researcher, left work undone and is COMPLETE_WITH_TRUNCATION: a usable
    record with its limits stated, never presentable as an exhaustive search
    (brief §30, §61).

    **Refused is not absent.** A community whose OWN sources refused the crawler
    is PARTIAL_BLOCKED, which says the evidence exists and could not be reached
    — quite different from a community about which little is published (§60).

    The signal for that is the source records, not the truncation text. An
    earlier version matched words like "unreachable" in the truncation reasons,
    which made every community PARTIAL_BLOCKED the moment a third-party academic
    index was down — an infrastructure gap, not a refusal, and nothing to do
    with whether this community's evidence could be reached.
    """
    if not workbook_verified:
        return "FAILED_TECHNICALLY"
    if report.failures:
        critical = {1, 5, 6, 8, 9, 15, 17, 18}
        if any(r.number in critical for r in report.failures):
            return "FAILED_TECHNICALLY"
        return "REQUIRES_HUMAN_REVIEW"
    if blocking_review:
        return "REQUIRES_HUMAN_REVIEW"

    # Blocked only when the community's own addresses refused us, and only when
    # that is most of them: one login-walled Facebook page beside a fully
    # crawled website is an ordinary result, not a blocked community.
    if blocked_sources and blocked_sources >= max(1, reachable_sources):
        return "PARTIAL_BLOCKED"

    if retrieval_stop_cause == "exhausted":
        # The community was worked out: the protocol finished on the evidence
        # rather than on a clock. That is a complete search.
        return "COMPLETE_WITH_UNCERTAINTY" if report.warnings else "COMPLETE"
    if retrieval_stop_cause in ("ceiling", "requested"):
        return "COMPLETE_WITH_TRUNCATION"
    if budget_exhausted:
        return "COMPLETE_WITH_TRUNCATION"
    if truncated:
        return "COMPLETE_WITH_TRUNCATION"
    if pages_opened < min_pages:
        # Too little was opened to call this a search of the community, but the
        # workbook is real and its limits are stated.
        return "COMPLETE_WITH_TRUNCATION"
    if report.warnings:
        return "COMPLETE_WITH_UNCERTAINTY"
    return "COMPLETE"


#: The six statuses a community run may end with (brief §92). Nothing else may
#: be written to `completion_status`.
COMPLETION_STATUSES = (
    "COMPLETE",
    "COMPLETE_WITH_UNCERTAINTY",
    "COMPLETE_WITH_TRUNCATION",
    "PARTIAL_BLOCKED",
    "REQUIRES_HUMAN_REVIEW",
    "FAILED_TECHNICALLY",
)
