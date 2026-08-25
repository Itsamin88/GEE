"""Phase A — read the actual research workbook and audit the field schema.

This runs before any crawling. It opens Stage_1_Documentary_Coding_Workbook_v6,
reads its real headers, dropdown vocabularies, formula cells and merged ranges,
and checks them against ``config/field_schema.yaml``. A mismatch is a hard error:
the study's destination has moved and the configuration must be updated
deliberately rather than the exporter silently writing to the wrong column.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# Sheets the exporter writes to, and the row the data starts on. Row 2 holds
# the template's worked example and is emptied rather than written over.
DATA_SHEETS = {
    "O1_Community_Attributes": 3,
    "O2_Practice_Matrix": 3,
    "O2b_Practice_Evidence": 3,
    "O3_Onset_Register": 3,
    "O5_Disagreement_Log": 3,
    "O6_Source_Index": 3,
    "O7_Search_Log": 3,
    "O11_Source_Set": 3,
}

# Row 2 of most sheets is an italic worked example the README tells the
# researcher to delete. The exporter deletes it in its working copy.
EXAMPLE_ROW_MARKER = "EXAMPLE ROW"


@dataclass
class SheetProfile:
    """What a sheet actually contains, read from the file rather than assumed."""

    title: str
    headers: dict[str, str]              # column letter -> header text
    header_index: dict[str, str]         # header text -> column letter
    validations: dict[str, list[str]]    # column letter -> allowed values
    formula_columns: set[str]
    merged_ranges: list[str]
    max_row: int
    max_column: int
    has_example_row: bool
    trailing_note_column: str | None


@dataclass
class AuditResult:
    workbook_path: Path
    sheets: dict[str, SheetProfile]
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors

    def raise_if_failed(self) -> None:
        if self.errors:
            joined = "\n  - ".join(self.errors)
            raise WorkbookAuditError(
                "The workbook template does not match config/field_schema.yaml:\n  - "
                + joined
                + "\n\nThe study's destination has changed. Update config/field_schema.yaml "
                  "deliberately (and log the change in config/decisions.yaml) before running again."
            )


class WorkbookAuditError(RuntimeError):
    """Raised when the workbook and the configured schema disagree."""


_LIST_SPLIT = re.compile(r",(?=(?:[^\"]*\"[^\"]*\")*[^\"]*$)")


def _parse_validation_formula(formula: str | None) -> list[str]:
    """Turn openpyxl's dropdown formula into the list of allowed values."""
    if not formula:
        return []
    text = formula.strip()
    if text.startswith('"') and text.endswith('"'):
        text = text[1:-1]
        return [part.strip() for part in text.split(",") if part.strip()]
    return []  # a range reference, not an inline list


def profile_workbook(path: Path) -> dict[str, SheetProfile]:
    """Read the real structure of every sheet."""
    wb = load_workbook(path, data_only=False, read_only=False)
    profiles: dict[str, SheetProfile] = {}
    try:
        for ws in wb.worksheets:
            headers: dict[str, str] = {}
            header_index: dict[str, str] = {}
            trailing_note: str | None = None
            for col in range(1, ws.max_column + 1):
                letter = get_column_letter(col)
                value = ws.cell(1, col).value
                if value is None:
                    continue
                text = str(value).strip()
                headers[letter] = text
                # Long trailing prose in row 1 is the sheet's guidance note,
                # not a column header.
                if len(text) > 80:
                    trailing_note = letter
                    continue
                header_index.setdefault(text, letter)

            validations: dict[str, list[str]] = {}
            for dv in ws.data_validations.dataValidation:
                values = _parse_validation_formula(dv.formula1)
                if not values:
                    continue
                for rng in str(dv.sqref).split():
                    letters = re.findall(r"([A-Z]+)\d+", rng)
                    if not letters:
                        continue
                    start, end = letters[0], letters[-1]
                    for idx in range(column_index_from_string(start),
                                     column_index_from_string(end) + 1):
                        validations[get_column_letter(idx)] = values

            formula_columns: set[str] = set()
            scan_max = min(ws.max_row, 410)
            for row in ws.iter_rows(min_row=2, max_row=scan_max):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_columns.add(cell.column_letter)

            has_example = False
            for col in range(1, min(ws.max_column, 8) + 1):
                value = ws.cell(2, col).value
                if isinstance(value, str) and EXAMPLE_ROW_MARKER in value.upper():
                    has_example = True
                    break

            profiles[ws.title] = SheetProfile(
                title=ws.title,
                headers=headers,
                header_index=header_index,
                validations=validations,
                formula_columns=formula_columns,
                merged_ranges=[str(r) for r in ws.merged_cells.ranges],
                max_row=ws.max_row,
                max_column=ws.max_column,
                has_example_row=has_example,
                trailing_note_column=trailing_note,
            )
    finally:
        wb.close()
    return profiles


def _iter_schema_fields(schema: dict[str, Any]) -> Iterable[tuple[str, dict[str, Any], str]]:
    for block_key, block in schema["blocks"].items():
        default_sheet = block.get("sheet")
        for fld in block["fields"]:
            sheet = fld.get("sheet") or default_sheet
            yield block_key, fld, sheet


def audit(path: Path, schema: dict[str, Any]) -> AuditResult:
    """Cross-check every configured field against the workbook it must land in."""
    profiles = profile_workbook(path)
    result = AuditResult(workbook_path=path, sheets=profiles)

    for sheet_name in DATA_SHEETS:
        if sheet_name not in profiles:
            result.errors.append(f"Sheet '{sheet_name}' is missing from the workbook.")

    for block_key, fld, sheet_name in _iter_schema_fields(schema):
        name = fld["name"]
        column = fld.get("column")
        if fld.get("unpacked_to_rows"):
            continue
        if column is None:
            continue
        profile = profiles.get(sheet_name)
        if profile is None:
            result.errors.append(
                f"Block {block_key} field '{name}' targets sheet '{sheet_name}', which does not exist."
            )
            continue

        expected_header = fld.get("workbook_header", name)
        actual = profile.headers.get(column)
        if actual is None:
            result.errors.append(
                f"{sheet_name}!{column} is empty but block {block_key} expects header '{expected_header}'."
            )
            continue
        if actual != expected_header:
            result.errors.append(
                f"{sheet_name}!{column} header is '{actual}' but the schema calls it "
                f"'{expected_header}'."
            )
            continue

        allowed = fld.get("allowed_values")
        actual_values = profile.validations.get(column)
        if allowed and actual_values:
            want = [str(v) for v in allowed]
            if want != actual_values:
                result.errors.append(
                    f"{sheet_name}!{column} ({name}) dropdown is {actual_values!r} "
                    f"but the schema allows {want!r}."
                )
        elif allowed and not actual_values:
            result.warnings.append(
                f"{sheet_name}!{column} ({name}) has configured values but no dropdown in the workbook."
            )
        elif actual_values and not allowed and fld.get("route") == "documentary":
            result.warnings.append(
                f"{sheet_name}!{column} ({name}) has a workbook dropdown {actual_values!r} "
                "that the schema does not constrain."
            )

        route = fld.get("route")
        is_formula = column in profile.formula_columns
        if is_formula and route not in {"derived", "researcher"}:
            result.errors.append(
                f"{sheet_name}!{column} ({name}) holds a formula in the template but the schema "
                f"marks it route={route}. Writing it would destroy the formula."
            )
        if route == "derived" and not is_formula:
            result.warnings.append(
                f"{sheet_name}!{column} ({name}) is marked derived but the template has no formula there."
            )

    # The formula map in the schema must match the file.
    for sheet_name, columns in schema.get("formula_columns", {}).items():
        profile = profiles.get(sheet_name)
        if profile is None:
            result.warnings.append(f"formula_columns names sheet '{sheet_name}', which does not exist.")
            continue
        configured = set(columns)
        if configured != profile.formula_columns:
            missing = sorted(profile.formula_columns - configured)
            extra = sorted(configured - profile.formula_columns)
            detail = []
            if missing:
                detail.append(f"present in the file but not configured: {missing}")
            if extra:
                detail.append(f"configured but not in the file: {extra}")
            result.errors.append(
                f"{sheet_name} formula columns disagree — " + "; ".join(detail)
            )

    # Practice matrix: all thirteen codes, in the workbook's own order.
    practice_profile = profiles.get("O2_Practice_Matrix")
    if practice_profile:
        codes = [f["name"] for f in schema["blocks"]["F"]["fields"] if f["name"].startswith("pc")]
        if len(codes) != 13:
            result.errors.append(f"The schema carries {len(codes)} practice codes; the codebook has 13.")
        for code in codes:
            if code not in practice_profile.header_index:
                result.errors.append(f"Practice code '{code}' has no column in O2_Practice_Matrix.")

    # Coding-level vocabulary must match both practice sheets.
    levels = schema["blocks"]["F"].get("coding_levels", [])
    for sheet_name, col in (("O2_Practice_Matrix", "C"), ("O2b_Practice_Evidence", "C")):
        profile = profiles.get(sheet_name)
        if profile and profile.validations.get(col) and profile.validations[col] != levels:
            result.errors.append(
                f"{sheet_name}!{col} coding levels are {profile.validations[col]!r}, "
                f"schema says {levels!r}."
            )

    # Source class vocabulary must match everywhere it appears.
    classes = list(schema["source_classes"].keys())
    for sheet_name, col in (("O1_Community_Attributes", "Z"), ("O2b_Practice_Evidence", "E"),
                            ("O6_Source_Index", "C"), ("O8_Enquiry_Record", "O")):
        profile = profiles.get(sheet_name)
        if profile and profile.validations.get(col) and profile.validations[col] != classes:
            result.errors.append(
                f"{sheet_name}!{col} source classes are {profile.validations[col]!r}, "
                f"schema says {classes!r}."
            )

    # Machine-identity columns must exist and must not be formulas.
    for sheet_name, mapping in schema.get("machine_identity_columns", {}).items():
        profile = profiles.get(sheet_name)
        if profile is None:
            result.errors.append(f"machine_identity_columns names missing sheet '{sheet_name}'.")
            continue
        for label, col in mapping.items():
            if profile.headers.get(col) != label:
                result.errors.append(
                    f"{sheet_name}!{col} should be '{label}' but is '{profile.headers.get(col)}'."
                )
            if col in profile.formula_columns:
                result.errors.append(f"{sheet_name}!{col} ({label}) is a formula cell; refusing to write it.")

    documentary = sum(
        1 for _, fld, _ in _iter_schema_fields(schema)
        if fld.get("route") == "documentary"
    )
    result.notes.append(f"{documentary} documentary fields configured (register v2.4 declares 88).")
    if documentary != 88:
        result.warnings.append(
            f"The schema carries {documentary} documentary fields; register v2.4 declares 88. "
            "If this is deliberate, record it in config/decisions.yaml."
        )
    result.notes.append(
        f"{len(profiles)} sheets read; "
        f"{sum(len(p.formula_columns) for p in profiles.values())} formula columns protected."
    )
    return result
