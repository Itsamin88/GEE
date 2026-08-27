"""Finalisation: getting a usable workbook out, whatever happened during the run.

A community run that gathered evidence for half an hour and then died writing the
spreadsheet has produced nothing. That is the failure this module exists to make
impossible (brief §4).

    export  ->  reopen and verify  ->  (repair and retry)  ->  manifests  ->  status

The retry ladder is deliberately ordered from least to most destructive, and
each rung is recorded:

  1. normal export — values cleaned as they are written
  2. aggressive sanitisation — additionally strips anything outside the range
     Excel reliably handles; can alter legitimate text, so never first
  3. core workbook only — supplementary evidence sheets dropped, each replaced
     by a notice saying where its rows still are

Rung 3 still produces the coded research record, which is the part the study
needs. Nothing here can end a run without a workbook unless the template itself
is unreadable, and that is reported as FAILED_TECHNICALLY rather than hidden.
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import load_workbook

from .. import profiling
from ..logging_setup import event, get_logger
from .sanitize import Sanitisation

log = get_logger("finalise")

#: Sheets whose absence means the workbook is not a research record. These come
#: from the researcher's own template; the X-prefixed ones are supplementary.
REQUIRED_SHEETS = ("O1_Community_Attributes", "O2_Practice_Matrix", "O3_Onset_Register",
                   "O6_Source_Index", "O11_Source_Set")


@dataclass
class VerificationResult:
    """Did the workbook we just wrote actually survive being written?"""

    ok: bool = False
    path: Path | None = None
    bytes: int = 0
    sheets: list[str] = field(default_factory=list)
    missing_sheets: list[str] = field(default_factory=list)
    core_rows: int = 0
    formulas_intact: bool = True
    problems: list[str] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "reopened": self.ok,
            "path": str(self.path) if self.path else None,
            "bytes": self.bytes,
            "sheet_count": len(self.sheets),
            "missing_sheets": list(self.missing_sheets),
            "core_rows": self.core_rows,
            "formulas_intact": self.formulas_intact,
            "problems": list(self.problems),
        }


@dataclass
class FinalisationResult:
    """The whole finalisation, including how many attempts it took."""

    export: Any = None
    verification: VerificationResult = field(default_factory=VerificationResult)
    attempts: int = 0
    strategy: str = ""
    repairs: list[str] = field(default_factory=list)
    elapsed_s: float = 0.0
    failed: bool = False
    failure_reason: str = ""

    @property
    def ok(self) -> bool:
        return self.verification.ok and not self.failed

    @property
    def sanitisation(self) -> Sanitisation:
        return getattr(self.export, "sanitisation", Sanitisation())

    def as_dict(self) -> dict[str, Any]:
        return {
            "ok": self.ok,
            "attempts": self.attempts,
            "strategy": self.strategy,
            "repairs": list(self.repairs),
            "elapsed_s": round(self.elapsed_s, 2),
            "verification": self.verification.as_dict(),
            "sanitisation": self.sanitisation.as_dict(),
            "omitted_sheets": dict(getattr(self.export, "omitted_sheets", {}) or {}),
            "failure_reason": self.failure_reason,
        }


def verify_workbook(path: Path, *, required: Sequence[str] = REQUIRED_SHEETS,
                    expect_formulas: bool = True) -> VerificationResult:
    """Reopen the workbook we just wrote and check it is really usable.

    Writing a file is not evidence that the file is good. This opens it again
    from disk, the way the researcher will (brief §47).
    """
    result = VerificationResult(path=path)
    if not path.exists():
        result.problems.append("the workbook was not written to disk")
        return result
    result.bytes = path.stat().st_size
    if result.bytes == 0:
        result.problems.append("the workbook is zero bytes")
        return result

    try:
        workbook = load_workbook(path, data_only=False)
    except Exception as exc:
        result.problems.append(f"the workbook could not be reopened: "
                               f"{type(exc).__name__}: {exc}")
        return result

    try:
        result.sheets = list(workbook.sheetnames)
        result.missing_sheets = [name for name in required if name not in result.sheets]
        if result.missing_sheets:
            result.problems.append(
                "missing core sheet(s): " + ", ".join(result.missing_sheets))

        # Row 3 is where coded rows begin; row 2 is the template's worked example.
        for name in ("O1_Community_Attributes", "O11_Source_Set"):
            if name in result.sheets:
                sheet = workbook[name]
                rows = sum(1 for row in range(3, min(sheet.max_row, 500) + 1)
                           if sheet.cell(row=row, column=1).value not in (None, ""))
                result.core_rows += rows

        if expect_formulas:
            found = any(
                isinstance(cell.value, str) and cell.value.startswith("=")
                for sheet in workbook.worksheets
                for row in sheet.iter_rows(max_row=min(sheet.max_row, 40))
                for cell in row
            )
            result.formulas_intact = found
            if not found:
                result.problems.append(
                    "no formula survived anywhere in the workbook; the template's "
                    "calculations would be dead")

        if result.core_rows == 0:
            result.problems.append("no coded rows reached the core sheets")
    finally:
        workbook.close()

    result.ok = not result.problems
    return result


def finalise_workbook(
    *,
    exporter_factory: Any,
    community_id: str,
    destination: Path,
    manifest: Mapping[str, Any] | None = None,
    max_attempts: int = 3,
) -> FinalisationResult:
    """Export, verify, and repair until a usable workbook exists.

    ``exporter_factory(aggressive, core_only)`` returns a fresh exporter. A new
    one is built for each attempt so no state from a failed attempt leaks into
    the next.
    """
    started = time.monotonic()
    outcome = FinalisationResult()
    ladder = [
        ("normal", dict(aggressive=False, core_only=False)),
        ("aggressive-sanitisation", dict(aggressive=True, core_only=False)),
        ("core-workbook-only", dict(aggressive=True, core_only=True)),
    ][:max(1, max_attempts)]

    last_error = ""
    for name, options in ladder:
        outcome.attempts += 1
        outcome.strategy = name
        if outcome.attempts > 1:
            event(log, "EXPORT",
                  f"attempt {outcome.attempts}: retrying as {name} after — {last_error}")
            outcome.repairs.append(f"attempt {outcome.attempts}: {name} ({last_error})")

        try:
            exporter = exporter_factory(**options)
            with profiling.timing("export"):
                export_result = exporter.export(community_id, destination,
                                                manifest=dict(manifest or {}))
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            log.error("export attempt %d (%s) failed: %s", outcome.attempts, name,
                      last_error, exc_info=True)
            # A partly written file is worse than none: the next attempt should
            # not be able to "verify" a corpse from the previous one.
            _remove(destination)
            continue

        verification = verify_workbook(destination)
        outcome.export = export_result
        outcome.verification = verification
        if verification.ok:
            if outcome.attempts > 1:
                event(log, "EXPORT",
                      f"workbook recovered on attempt {outcome.attempts} ({name})")
            outcome.elapsed_s = time.monotonic() - started
            return outcome

        last_error = "; ".join(verification.problems) or "verification failed"
        log.warning("[EXPORT] attempt %d (%s) produced a workbook that did not verify: %s",
                    outcome.attempts, name, last_error)
        _remove(destination)

    outcome.failed = True
    outcome.failure_reason = last_error or "the workbook could not be produced"
    outcome.elapsed_s = time.monotonic() - started
    log.error("[EXPORT] no usable workbook after %d attempt(s): %s",
              outcome.attempts, outcome.failure_reason)
    return outcome


def _remove(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass
    except OSError as exc:
        log.debug("could not remove %s: %s", path, exc)
