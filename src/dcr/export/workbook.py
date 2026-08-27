"""Workbook export.

A clean copy of Stage_1_Documentary_Coding_Workbook_v6 is filled in place. Its
sheet names, columns, dropdowns and formulas are preserved exactly: the exporter
refuses to write into any cell whose template value is a formula, and refuses
any value a dropdown does not allow (brief §42, §71; decision DCR-D020).

Supplementary sheets are appended for the evidence the canonical workbook has
nowhere to put. Each has a stated methodological purpose (brief §43).
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .. import __version__
from ..db import Database
from .sanitize import (SafeSheet, Sanitisation, clean_cell, safe_sheet_title,
                       sanitise_row,
                       sanitise_workbook, _sample as _sample_of)
from ..logging_setup import event, get_logger
from ..workbook_audit import EXAMPLE_ROW_MARKER, profile_workbook

log = get_logger("export")

# Data starts at row 3. Row 2 is the template's worked example: it is emptied,
# not deleted, because deleting it would shift every formula in the sheet, and
# the workbook's own Cohort_Tracker counts from A3 (decision DCR-D023).
DATA_START_ROW = 3

HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
HEADER_FONT = Font(color="FFFFFF", bold=True)

# Supplementary sheets, each with the reason it exists.
SUPPLEMENTARY_SHEETS: dict[str, str] = {
    "X1_Evidence_Register":
        "Every passage, table cell, figure and metadata item behind a coded value, with its "
        "exact wording. The canonical workbook records the source; this records the sentence.",
    "X2_Claim_Register":
        "One row per claim, before reconciliation. Contradictory claims are preserved here even "
        "where only one value reached O1, so nothing is silently discarded.",
    "X3_Image_Evidence":
        "The image manifest. Every saved image, why it was kept, what it may support, and what "
        "documentary text would license a claim from it.",
    "X3b_Image_Triage":
        "Every image the crawl SAW, downloaded or not, with the metadata read before deciding "
        "and the reason for the decision. A gallery caption often carries a date no text on the "
        "page gives, so a skipped candidate is still research material — and the ledger is what "
        "makes the triage auditable rather than something to be taken on trust.",
    "X4_Document_Register":
        "Every file retrieved, its hash, parser status and the sources it was reached from. A "
        "document downloaded but not parsed is visible here rather than counted as read.",
    "X5_Crawl_Audit":
        "What was actually opened, per source and per stage: pages, documents, images, budget "
        "spent and why a source stopped being crawled.",
    "X6_Failure_Log":
        "Every failed or blocked retrieval, with its cause. A reported block is data; a silent "
        "gap is not.",
    "X7_Source_Graph":
        "Relationships between sources — which copied which, and the similarity that established "
        "it. This is the evidence behind the independence groups.",
    "X8_Review_Queue":
        "Cases where a machine decision would be a bad decision: conflicting evidence, uncertain "
        "ownership, unverified citations, ambiguous area.",
    "X9_Discovery_Log":
        "How each address and URL was found, so a thin result can be distinguished from a "
        "narrow search.",
    "X10_Field_Provenance":
        "One row per coded field: the value, how it was resolved, which sources and independence "
        "groups support it, and the residual uncertainty.",
    "X11_Run_Manifest":
        "The reproducibility record: versions, configuration hashes, research-document hashes, "
        "optional features available, and the decisions the program applied.",
}


class WorkbookWriteError(RuntimeError):
    pass


@dataclass
class ExportResult:
    path: Path
    rows_written: dict[str, int] = field(default_factory=dict)
    refusals: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    #: What had to be cleaned to make the text storable in Excel. The raw text
    #: is untouched in the database; only this workbook was cleaned.
    sanitisation: Sanitisation = field(default_factory=Sanitisation)
    #: Supplementary sheets deliberately left out because they could not be
    #: written. The core workbook is never sacrificed for one of these.
    omitted_sheets: dict[str, str] = field(default_factory=dict)

    @property
    def excel_sanitized(self) -> str:
        return "yes" if self.sanitisation.occurred else "no"


class WorkbookExporter:
    """Fills a copy of the canonical workbook without damaging it."""

    def __init__(self, template: Path, schema: Mapping[str, Any], db: Database,
                 *, coder_id: str = "", decisions: Mapping[str, Any] | None = None,
                 aggressive_sanitize: bool = False, core_only: bool = False):
        #: The last rung of the retry ladder: write the coded research record
        #: and skip the supplementary evidence sheets entirely. Their rows are
        #: still in the database and in the CSV manifests.
        self.core_only = bool(core_only)
        #: Set only by the finalisation retry ladder, after a normal export has
        #: already failed. It can alter legitimate text, so it is never first.
        self.aggressive_sanitize = bool(aggressive_sanitize)
        self.sanitisation = Sanitisation()
        #: Supplementary sheets that could not be built, and why.
        self.omitted_sheets: dict[str, str] = {}
        self.template = Path(template)
        self.schema = schema
        self.db = db
        self.coder_id = coder_id or f"DCR/{__version__}"
        self.decisions = dict(decisions or {})
        self.profiles = profile_workbook(self.template)
        self.refusals: list[str] = []
        self.warnings: list[str] = []

    # -- entry -------------------------------------------------------------
    def sheet(self, workbook: Any, name: str) -> Any:
        """The ONE way this exporter reaches a worksheet to write to it.

        Wrapping every sheet means there is no route by which a value can reach
        a cell without being made Excel-safe first. That matters more than the
        cleaning itself: the reported failure did not happen because nobody had
        thought about illegal characters — it happened because one writer did
        not go through the code that had (brief §11).
        """
        return SafeSheet(workbook[name], log=self.sanitisation,
                         aggressive=self.aggressive_sanitize)

    def export(self, community_id: str, destination: Path,
               manifest: Mapping[str, Any] | None = None) -> ExportResult:
        workbook = load_workbook(self.template, data_only=False)
        result = ExportResult(path=destination)
        today = date.today().isoformat()

        community = self.db.query_one(
            "SELECT * FROM communities WHERE community_id = ?", (community_id,))
        if community is None:
            raise WorkbookWriteError(f"no community {community_id!r} in the database")
        site_id = community["site_id"]
        name = community["name_input"]

        for sheet_name in ("O1_Community_Attributes", "O2_Practice_Matrix", "O2b_Practice_Evidence",
                           "O3_Onset_Register", "O5_Disagreement_Log", "O6_Source_Index",
                           "O7_Search_Log", "O11_Source_Set", "O10_Polygon_And_Area"):
            self._clear_example_row(self.sheet(workbook, sheet_name))

        result.rows_written["O1_Community_Attributes"] = self._write_o1(
            workbook, community_id, site_id, name, today)
        result.rows_written["O2_Practice_Matrix"] = self._write_o2(
            workbook, community_id, site_id, name, today)
        result.rows_written["O2b_Practice_Evidence"] = self._write_o2b(
            workbook, community_id, site_id)
        result.rows_written["O3_Onset_Register"] = self._write_o3(
            workbook, community_id, site_id, name, today)
        result.rows_written["O5_Disagreement_Log"] = self._write_o5(
            workbook, community_id, site_id, today)
        result.rows_written["O6_Source_Index"] = self._write_o6(workbook, community_id, site_id)
        result.rows_written["O7_Search_Log"] = self._write_o7(workbook, community_id, site_id)
        result.rows_written["O11_Source_Set"] = self._write_o11(workbook, community_id, site_id)
        result.rows_written["O10_Polygon_And_Area"] = self._write_o10_key(
            workbook, community_id, site_id, name)

        self._write_supplementary(workbook, community_id, manifest or {})

        # The safety net. The writers above clean as they go, which is what
        # makes the per-sheet counts meaningful; this does not care which code
        # path wrote a cell, so a value arriving by a route nobody remembered
        # still cannot take the run down after an hour of work.
        sanitise_workbook(workbook, log=self.sanitisation,
                          aggressive=self.aggressive_sanitize)

        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        workbook.close()
        result.refusals = list(self.refusals)
        result.warnings = list(self.warnings)
        result.sanitisation = self.sanitisation
        result.omitted_sheets = dict(self.omitted_sheets)
        if self.sanitisation.occurred:
            result.warnings.append(f"excel_sanitized=yes: {self.sanitisation.summary()}")
        return result

    # -- guarded cell writing ---------------------------------------------
    def _set(self, sheet: Worksheet, row: int, column: str, value: Any, *,
             field_name: str = "", datatype: str | None = None) -> bool:
        """Write one cell, refusing anything that would damage the workbook."""
        sheet_name = sheet.title
        profile = self.profiles.get(sheet_name)
        if profile is None:
            self.refusals.append(f"{sheet_name} is not part of the audited template")
            return False

        if column in profile.formula_columns:
            self.refusals.append(
                f"refused to write {field_name or column} to {sheet_name}!{column}{row}: "
                "the template holds a formula there"
            )
            return False

        researcher_only = self.schema.get("researcher_only_columns", {}).get(sheet_name, [])
        if column in researcher_only:
            self.refusals.append(
                f"refused to write {field_name or column} to {sheet_name}!{column}{row}: "
                "that column belongs to the researcher"
            )
            return False

        if value is None or value == "":
            return False

        allowed = profile.validations.get(column)
        if allowed and str(value) not in allowed:
            self.refusals.append(
                f"refused to write {value!r} to {sheet_name}!{column}{row} "
                f"({field_name or column}): the dropdown allows {allowed}"
            )
            return False

        target = sheet.cell(row=row, column=column_index_from_string(column))
        if isinstance(target.value, str) and target.value.startswith("="):
            self.refusals.append(
                f"refused to overwrite a formula at {sheet_name}!{column}{row}")
            return False
        cleaned, removed, truncated = clean_cell(
            _coerce(value, datatype), aggressive=self.aggressive_sanitize)
        if removed or truncated:
            self.sanitisation.note(sheet_name, removed, truncated=truncated,
                                   sample=_sample_of(value))
        target.value = cleaned
        return True

    def _clear_example_row(self, sheet: Worksheet) -> None:
        """Empty row 2, the worked example the README tells the researcher to delete.

        The row is emptied rather than deleted: deleting it would shift every
        formula in the sheet by one row. Its cells are cleared completely,
        including the static values it carries in columns that hold formulas
        from row 3 down — leaving those in place would show a coded community
        the example's polygon area.
        """
        profile = self.profiles.get(sheet.title)
        if profile is None or not profile.has_example_row:
            return
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=2, column=column)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None

    def _next_row(self, sheet: Worksheet, start: int = DATA_START_ROW) -> int:
        row = start
        while sheet.cell(row=row, column=1).value not in (None, ""):
            row += 1
        return row

    # -- canonical sheets --------------------------------------------------
    def _write_o1(self, workbook: Any, community_id: str, site_id: str, name: str,
                  today: str) -> int:
        sheet = self.sheet(workbook, "O1_Community_Attributes")
        row = self._next_row(sheet)
        sheet.cell(row=row, column=1).value = site_id

        values = self._field_values(community_id)
        for block in self.schema["blocks"].values():
            if block.get("sheet") != "O1_Community_Attributes":
                continue
            for field_def in block["fields"]:
                column = field_def.get("column")
                if not column or field_def.get("route") in ("derived", "researcher"):
                    continue
                record = values.get(field_def["name"])
                if record is None or record["status"] != "coded":
                    continue
                self._set(sheet, row, column, record["value"], field_name=field_def["name"],
                          datatype=field_def.get("datatype"))

        # The name is the paste target's identity even when nothing else coded.
        self._set(sheet, row, "B", name, field_name="community_name_official")
        # Machine identity (decision DCR-D004).
        identity = self.schema.get("machine_identity_columns", {}).get(
            "O1_Community_Attributes", {})
        self._set(sheet, row, identity.get("coder_id", "BF"), self.coder_id, field_name="coder_id")
        self._set(sheet, row, identity.get("coding_date", "BG"), today, field_name="coding_date")

        notes = self._notes_for(community_id)
        self._set(sheet, row, "BJ", notes, field_name="notes")
        return 1

    def _write_o2(self, workbook: Any, community_id: str, site_id: str, name: str,
                  today: str) -> int:
        sheet = self.sheet(workbook, "O2_Practice_Matrix")
        row = self._next_row(sheet)
        sheet.cell(row=row, column=1).value = site_id
        sheet.cell(row=row, column=2).value = name
        values = self._field_values(community_id)
        for field_def in self.schema["blocks"]["F"]["fields"]:
            name_key = field_def["name"]
            if not name_key.startswith("pc"):
                continue
            record = values.get(name_key)
            level = record["value"] if record and record["status"] == "coded" else "not mentioned"
            self._set(sheet, row, field_def["column"], level, field_name=name_key)
        identity = self.schema.get("machine_identity_columns", {}).get("O2_Practice_Matrix", {})
        self._set(sheet, row, identity.get("coder_id", "P"), self.coder_id)
        self._set(sheet, row, identity.get("coding_date", "Q"), today)
        notes = values.get("practice_evidence_notes")
        if notes and notes["status"] == "coded":
            self._set(sheet, row, "R", str(notes["value"])[:2000],
                      field_name="practice_evidence_notes")
        return 1

    def _write_o2b(self, workbook: Any, community_id: str, site_id: str) -> int:
        sheet = self.sheet(workbook, "O2b_Practice_Evidence")
        rows = self.db.query(
            "SELECT c.*, e.quote, e.locator FROM claims c "
            "LEFT JOIN evidence e ON c.evidence_id = e.evidence_id "
            "WHERE c.community_id = ? AND c.coding_level IS NOT NULL "
            "AND c.coding_level != 'not mentioned' ORDER BY c.field_name, c.claim_id",
            (community_id,))
        written = 0
        seen: set[tuple[str, str]] = set()
        for record in rows:
            key = (record["field_name"], record["source_id"] or "")
            if key in seen:
                continue
            seen.add(key)
            row = self._next_row(sheet)
            sheet.cell(row=row, column=1).value = site_id
            self._set(sheet, row, "B", record["field_name"])
            self._set(sheet, row, "C", record["coding_level"])
            self._set(sheet, row, "D", record["source_id"])
            self._set(sheet, row, "E", record["source_class"])
            self._set(sheet, row, "F", self._full_text_status(record["document_id"]))
            self._set(sheet, row, "G", _year_of(record["publication_date"]))
            self._set(sheet, row, "H", record["reference_year"])
            self._set(sheet, row, "J", (record["exact_wording"] or record["quote"] or "")[:1500])
            identity = self.schema.get("machine_identity_columns", {}).get(
                "O2b_Practice_Evidence", {})
            self._set(sheet, row, identity.get("coder_id", "K"), self.coder_id)
            self._set(sheet, row, "L", (record["rationale"] or "")[:1000])
            written += 1
        return written

    def _write_o3(self, workbook: Any, community_id: str, site_id: str, name: str,
                  today: str) -> int:
        sheet = self.sheet(workbook, "O3_Onset_Register")
        row = self._next_row(sheet)
        sheet.cell(row=row, column=1).value = site_id
        sheet.cell(row=row, column=2).value = name
        values = self._field_values(community_id)
        for field_def in self.schema["blocks"]["C"]["fields"]:
            column = field_def.get("column")
            if not column or field_def.get("route") in ("derived", "researcher"):
                continue
            record = values.get(field_def["name"])
            if record is None or record["status"] != "coded":
                continue
            self._set(sheet, row, column, record["value"], field_name=field_def["name"],
                      datatype=field_def.get("datatype"))
        source_ids = (values.get("date_intervention_onset") or {}).get("source_ids", "")
        self._set(sheet, row, "S", source_ids)
        identity = self.schema.get("machine_identity_columns", {}).get("O3_Onset_Register", {})
        self._set(sheet, row, identity.get("coder_id", "T"), self.coder_id)
        self._set(sheet, row, identity.get("coding_date", "U"), today)
        onset = values.get("date_intervention_onset")
        if onset:
            self._set(sheet, row, "V", (onset.get("rationale") or "")[:1500])
        return 1

    def _write_o5(self, workbook: Any, community_id: str, site_id: str, today: str) -> int:
        sheet = self.sheet(workbook, "O5_Disagreement_Log")
        written = 0
        for record in self.db.query(
            "SELECT * FROM conflicts WHERE community_id = ? ORDER BY conflict_id", (community_id,)
        ):
            row = self._next_row(sheet)
            sheet.cell(row=row, column=1).value = record["conflict_id"]
            self._set(sheet, row, "B", site_id)
            self._set(sheet, row, "C", record["field_name"])
            self._set(sheet, row, "D", record["value_a"])
            self._set(sheet, row, "E", record["value_b"])
            self._set(sheet, row, "F",
                      f"A: {record['source_a'] or 'unattributed'} "
                      f"(group {record['group_a'] or '?'}, rank {record['rank_a'] or '-'}) | "
                      f"B: {record['source_b'] or 'unattributed'} "
                      f"(group {record['group_b'] or '?'}, rank {record['rank_b'] or '-'})")
            self._set(sheet, row, "G", (record["rule_invoked"] or "")[:1000])
            self._set(sheet, row, "H", record["resolution_type"] or "unresolved")
            self._set(sheet, row, "J", record["final_value"])
            self._set(sheet, row, "K", "no")
            self._set(sheet, row, "L", "no")
            self._set(sheet, row, "M", today)
            self._set(sheet, row, "N",
                      ("machine resolution; " if not record["human_review"]
                       else "REQUIRES A HUMAN DECISION; ")
                      + f"residual uncertainty: {record['residual_uncertainty'] or 'none'}")
            written += 1
        return written

    def _write_o6(self, workbook: Any, community_id: str, site_id: str) -> int:
        sheet = self.sheet(workbook, "O6_Source_Index")
        written = 0
        for record in self.db.query(
            "SELECT * FROM sources WHERE community_id = ? ORDER BY source_id", (community_id,)
        ):
            row = self._next_row(sheet)
            sheet.cell(row=row, column=1).value = record["source_id"]
            self._set(sheet, row, "B", site_id)
            self._set(sheet, row, "C", record["source_class"])
            self._set(sheet, row, "D", record["platform_type"])
            self._set(sheet, row, "E", self._source_title(record["source_id"]) or record["url"])
            self._set(sheet, row, "F", record["url"])
            self._set(sheet, row, "G", record["archive_earliest_snapshot"] or "")
            self._set(sheet, row, "H", (record["last_crawled_utc"] or "")[:10])
            self._set(sheet, row, "I", self._source_access(record))
            self._set(sheet, row, "J", f"{record['pages_opened'] or 0} pages opened")
            self._set(sheet, row, "K", self._fields_from_source(community_id, record["source_id"]))
            self._set(sheet, row, "L", record["language"])
            self._set(sheet, row, "M", "no")
            self._set(sheet, row, "O", "no")
            self._set(sheet, row, "P", (record["independence_reason"] or "")[:800])
            self._set(sheet, row, "Q", record["independence_group"])
            written += 1

        # Verified academic and institutional records are sources in their own right.
        for record in self.db.query(
            "SELECT * FROM academic_records WHERE community_id = ? ORDER BY record_id",
            (community_id,)
        ):
            row = self._next_row(sheet)
            sheet.cell(row=row, column=1).value = record["record_id"]
            self._set(sheet, row, "B", site_id)
            self._set(sheet, row, "C", "S1")
            self._set(sheet, row, "D", record["record_type"])
            self._set(sheet, row, "E", record["title"])
            self._set(sheet, row, "F", record["url"] or record["doi"])
            self._set(sheet, row, "G", record["year"])
            self._set(sheet, row, "H", (record["created_utc"] or "")[:10])
            self._set(sheet, row, "I", record["full_text_status"])
            self._set(sheet, row, "K", record["relevance_reason"])
            self._set(sheet, row, "N", record["doi"])
            self._set(sheet, row, "O", record["verified_resolves"])
            self._set(sheet, row, "P", (record["verification_detail"] or "")[:800])
            self._set(sheet, row, "Q", "G-ACAD")
            written += 1
        return written

    def _write_o7(self, workbook: Any, community_id: str, site_id: str) -> int:
        sheet = self.sheet(workbook, "O7_Search_Log")
        written = 0
        rows = self.db.query(
            "SELECT database_name, database_type, "
            "       GROUP_CONCAT(DISTINCT query) AS queries, "
            "       GROUP_CONCAT(DISTINCT language) AS languages, "
            "       SUM(hits_returned) AS hits, SUM(full_text_opened) AS full_text, "
            "       SUM(abstract_only) AS abstracts, MAX(searched_utc) AS searched, "
            "       GROUP_CONCAT(DISTINCT result) AS results, "
            "       GROUP_CONCAT(DISTINCT detail) AS details "
            "FROM searches WHERE community_id = ? GROUP BY database_name, database_type "
            "ORDER BY database_type, database_name",
            (community_id,))
        identity = self.schema.get("machine_identity_columns", {}).get("O7_Search_Log", {})
        for record in rows:
            row = self._next_row(sheet)
            sheet.cell(row=row, column=1).value = site_id
            self._set(sheet, row, "B", record["database_name"])
            self._set(sheet, row, "C", _search_type(record["database_type"]))
            self._set(sheet, row, "D", (record["queries"] or "")[:1200])
            self._set(sheet, row, "E", (record["languages"] or "")[:200])
            self._set(sheet, row, "F", int(record["hits"] or 0))
            self._set(sheet, row, "G", int(record["full_text"] or 0))
            self._set(sheet, row, "H", int(record["abstracts"] or 0))
            self._set(sheet, row, "I", _worst_result(record["results"]))
            self._set(sheet, row, "J", (record["searched"] or "")[:10])
            self._set(sheet, row, identity.get("coder_id", "K"), self.coder_id)
            self._set(sheet, row, "L", (record["details"] or "")[:900])
            written += 1
        return written

    def _write_o11(self, workbook: Any, community_id: str, site_id: str) -> int:
        sheet = self.sheet(workbook, "O11_Source_Set")
        written = 0
        identity = self.schema.get("machine_identity_columns", {}).get("O11_Source_Set", {})
        for record in self.db.query(
            "SELECT * FROM sources WHERE community_id = ? ORDER BY address_id", (community_id,)
        ):
            row = self._next_row(sheet)
            sheet.cell(row=row, column=1).value = site_id
            self._set(sheet, row, "B", record["address_id"])
            self._set(sheet, row, "C", record["url"])
            self._set(sheet, row, "D", record["platform_type"])
            self._set(sheet, row, "E", record["supplied_or_discovered"])
            self._set(sheet, row, "F", record["independence_group"])
            self._set(sheet, row, "G", record["crawl_status"] or "not attempted")
            self._set(sheet, row, "H", record["pages_opened"] or 0)
            self._set(sheet, row, "I", "yes" if record["archive_checked"] else "no")
            self._set(sheet, row, "J", record["archive_earliest_snapshot"])
            self._set(sheet, row, "K", record["earliest_dated_item"])
            self._set(sheet, row, "L", record["latest_dated_item"])
            self._set(sheet, row, "M", self._fields_from_source(community_id, record["source_id"]))
            self._set(sheet, row, "N", record["language"])
            self._set(sheet, row, "O", (record["last_crawled_utc"] or "")[:10])
            self._set(sheet, row, identity.get("coder_id", "P"), self.coder_id)
            self._set(sheet, row, "Q",
                      f"{record['discovery_method'] or 'supplied'}; "
                      f"{record['independence_reason'] or ''}"[:900])
            written += 1
        return written

    def _write_o10_key(self, workbook: Any, community_id: str, site_id: str, name: str) -> int:
        """O10 gets only its identity columns. Everything else is the researcher's
        drawing or a formula reading O1 (decision DCR-D020)."""
        sheet = self.sheet(workbook, "O10_Polygon_And_Area")
        row = self._next_row(sheet)
        sheet.cell(row=row, column=1).value = site_id
        self._set(sheet, row, "B", name)
        return 1

    # -- supplementary sheets ----------------------------------------------
    def _write_supplementary(self, workbook: Any, community_id: str,
                             manifest: Mapping[str, Any]) -> None:
        builders = {
            "X1_Evidence_Register": self._sheet_evidence,
            "X2_Claim_Register": self._sheet_claims,
            "X3_Image_Evidence": self._sheet_images,
            "X3b_Image_Triage": self._sheet_image_triage,
            "X4_Document_Register": self._sheet_documents,
            "X5_Crawl_Audit": self._sheet_crawl,
            "X6_Failure_Log": self._sheet_failures,
            "X7_Source_Graph": self._sheet_source_graph,
            "X8_Review_Queue": self._sheet_review,
            "X9_Discovery_Log": self._sheet_discovery,
            "X10_Field_Provenance": self._sheet_field_provenance,
        }
        if self.core_only:
            for title in list(builders) + ["X11_Run_Manifest"]:
                self.omitted_sheets[title] = (
                    "omitted deliberately: an earlier export attempt failed, so the "
                    "core workbook was written without the supplementary sheets")
                self._add_omission_notice(workbook, title, self.omitted_sheets[title])
            self.warnings.append(
                "the supplementary evidence sheets were omitted so that the coded "
                "workbook could be produced; every row is still in the database and "
                "in the CSV manifests in 09_final/")
            return

        for title, builder in builders.items():
            # A supplementary sheet is evidence ABOUT the run, not the coded
            # research record. If one cannot be built, the workbook still ships
            # without it and the omission is written into the audit sheet —
            # losing the whole workbook to a malformed evidence row would be a
            # far worse outcome than losing that row (brief §4, §48).
            try:
                headers, rows = builder(community_id)
                self._add_sheet(workbook, title, headers, rows)
            except Exception as exc:
                reason = f"{type(exc).__name__}: {exc}"
                log.error("supplementary sheet %s could not be built: %s", title, reason,
                          exc_info=True)
                self.omitted_sheets[title] = reason
                self.warnings.append(
                    f"{title} was omitted from the workbook ({reason}); the underlying "
                    "rows are still in the database and in the CSV manifests")
                self._add_omission_notice(workbook, title, reason)
        headers, rows = self._sheet_manifest(manifest)
        self._add_sheet(workbook, "X11_Run_Manifest", headers, rows)

    def _add_omission_notice(self, workbook: Any, title: str, reason: str) -> None:
        """Leave a sheet saying what is missing, rather than a silent absence."""
        try:
            self._add_sheet(
                workbook, title,
                ["sheet", "status", "reason", "where_the_data_still_is"],
                [[title, "OMITTED", reason,
                  "the database, and the CSV manifests in 09_final/"]],
            )
        except Exception:                     # even the notice must not throw
            log.debug("could not write an omission notice for %s", title, exc_info=True)

    def _add_sheet(self, workbook: Any, title: str, headers: Sequence[str],
                   rows: Sequence[Sequence[Any]]) -> None:
        title = safe_sheet_title(title)
        if title in workbook.sheetnames:
            del workbook[title]
        sheet = SafeSheet(workbook.create_sheet(title), log=self.sanitisation,
                          aggressive=self.aggressive_sanitize)
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index)
            cell.value = header
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        note_column = len(headers) + 2
        note = sheet.cell(row=1, column=note_column)
        note.value = SUPPLEMENTARY_SHEETS.get(title, "")
        note.font = Font(italic=True, color="555555")
        for row_index, row in enumerate(rows, start=2):
            cleaned = sanitise_row([_coerce(v) for v in row], sheet=title,
                                   log=self.sanitisation,
                                   aggressive=self.aggressive_sanitize)
            for col_index, value in enumerate(cleaned, start=1):
                sheet.cell(row=row_index, column=col_index).value = value
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = min(
                48, max(12, len(str(header)) + 4))

    def _sheet_evidence(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["evidence_id", "source_id", "document_id", "page_id", "image_id",
                   "evidence_type", "locator", "page_number", "source_class",
                   "publication_date", "retrieval_date", "language", "quote", "context"]
        rows = [
            [r["evidence_id"], r["source_id"], r["document_id"], r["page_id"], r["image_id"],
             r["evidence_type"], r["locator"], r["page_number"], r["source_class"],
             r["publication_date"], r["retrieval_date"], r["language"],
             (r["quote"] or "")[:2000], (r["context"] or "")[:1000]]
            for r in self.db.query(
                "SELECT * FROM evidence WHERE community_id=? ORDER BY evidence_id", (community_id,))
        ]
        return headers, rows

    def _sheet_claims(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["claim_id", "field_name", "value", "original_value", "normalized_value",
                   "value_type", "reference_year", "publication_date", "retrieval_date",
                   "source_id", "source_class", "independence_group", "evidence_id",
                   "document_id", "locator", "evidence_rank", "coding_level", "confidence",
                   "conflict_status", "extractor", "model_name", "prompt_version",
                   "verified_passage", "exact_wording", "rationale"]
        rows = [
            [r["claim_id"], r["field_name"], r["value"], r["original_value"],
             r["normalized_value"], r["value_type"], r["reference_year"], r["publication_date"],
             r["retrieval_date"], r["source_id"], r["source_class"], r["independence_group"],
             r["evidence_id"], r["document_id"], r["locator"], r["evidence_rank"],
             r["coding_level"], r["confidence"], r["conflict_status"], r["extractor"],
             r["model_name"], r["prompt_version"], r["verified_passage"],
             (r["exact_wording"] or "")[:1500], (r["rationale"] or "")[:800]]
            for r in self.db.query(
                "SELECT * FROM claims WHERE community_id=? ORDER BY field_name, claim_id",
                (community_id,))
        ]
        return headers, rows

    def _sheet_images(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["image_id", "community_id", "candidate_id", "source_id", "document_id",
                   "filename", "original_filename", "local_path", "original_url", "page_url",
                   "archive_url", "page_number", "figure_number", "extraction_method",
                   "source_title", "publication_date", "image_type", "research_topic",
                   "caption", "surrounding_text_summary", "evidence_subject",
                   "possible_relevant_fields", "visual_evidence_allowed",
                   "documentary_text_support", "image_date_if_known",
                   "image_date_confidence", "OCR_text_if_used", "relevance_class",
                   "retrieval_priority", "relevance_reason", "confidence", "width", "height",
                   "bytes", "sha256", "retrieved_utc", "notes"]
        rows = [
            [r["image_id"], community_id, r["candidate_id"], r["source_id"], r["document_id"],
             r["filename"], r["original_filename"], r["local_path"], r["original_url"],
             r["page_url"], r["archive_url"], r["page_number"], r["figure_number"],
             r["extraction_method"], r["source_title"],
             r["publication_date"], r["image_type"], r["research_topic"],
             (r["caption"] or "")[:800], r["surrounding_summary"], r["evidence_subject"],
             r["possible_fields"], r["visual_evidence_allowed"], r["documentary_text_support"],
             r["image_date"], r["image_date_confidence"], (r["ocr_text"] or "")[:800],
             r["relevance_class"], r["priority"], r["relevance_reason"], r["confidence"],
             r["width"], r["height"], r["bytes"], r["sha256"],
             r["retrieval_utc"] or r["created_utc"], r["notes"]]
            for r in self.db.query(
                "SELECT * FROM images WHERE community_id=? ORDER BY image_id", (community_id,))
        ]
        return headers, rows

    def _sheet_image_triage(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        """The triage ledger: what was seen, what was taken, and why (brief §6, §10)."""
        headers = ["candidate_id", "image_id", "decision", "decision_reason",
                   "retrieval_priority", "priority_rank", "relevance_class",
                   "relevance_score", "relevance_reason", "image_type", "research_topic",
                   "origin", "source_id", "document_id", "page_id", "original_url",
                   "page_url", "archive_url", "filename", "alt_text", "title_text",
                   "caption", "page_heading", "document_title", "page_number",
                   "figure_number", "extraction_method", "width", "height", "bytes",
                   "mime_type", "publication_date", "image_date", "source_class",
                   "independence_group", "possible_relevant_fields",
                   "documentary_text_support", "sha256", "stage", "seen_utc"]
        rows = [
            [r["candidate_id"], r["image_id"], r["decision"], r["decision_reason"],
             r["priority"], r["priority_rank"], r["relevance_class"],
             r["relevance_score"], r["relevance_reason"], r["image_type"],
             r["research_topic"], r["origin"], r["source_id"], r["document_id"],
             r["page_id"], r["original_url"], r["page_url"], r["archive_url"],
             r["filename"], (r["alt_text"] or "")[:500], (r["title_text"] or "")[:500],
             (r["caption"] or "")[:800], r["page_heading"], r["document_title"],
             r["page_number"], r["figure_number"], r["extraction_method"],
             r["width"], r["height"], r["bytes"], r["mime_type"],
             r["publication_date"], r["image_date"], r["source_class"],
             r["independence_group"], r["possible_fields"],
             r["documentary_text_support"], r["sha256"], r["stage"], r["seen_utc"]]
            for r in self.db.query(
                "SELECT * FROM image_candidates WHERE community_id=? "
                "ORDER BY CASE priority WHEN 'HIGH' THEN 0 WHEN 'MEDIUM' THEN 1 "
                "WHEN 'LOW' THEN 2 ELSE 3 END, priority_rank DESC, candidate_id",
                (community_id,))
        ]
        return headers, rows

    def _sheet_documents(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["document_id", "title", "filename", "extension", "mime_sniffed", "bytes",
                   "sha256", "storage_path", "page_count", "publication_date", "doc_kind",
                   "parser", "parser_status", "text_status", "table_status", "image_status",
                   "text_chars", "source_ids", "original_urls", "retrieved", "notes"]
        rows = []
        for r in self.db.query(
            "SELECT * FROM documents WHERE community_id=? ORDER BY document_id", (community_id,)
        ):
            links = self.db.query(
                "SELECT source_id, original_url, retrieved_utc FROM document_sources "
                "WHERE document_id=?", (r["document_id"],))
            rows.append([
                r["document_id"], r["title"], r["filename"], r["extension"], r["mime_sniffed"],
                r["bytes"], r["sha256"], r["storage_path"], r["page_count"],
                r["publication_date"], r["doc_kind"], r["parser"], r["parser_status"],
                r["text_status"], r["table_status"], r["image_status"], r["text_chars"],
                "; ".join(str(l["source_id"]) for l in links),
                "; ".join(str(l["original_url"]) for l in links)[:1500],
                "; ".join(str(l["retrieved_utc"])[:10] for l in links),
                r["notes"],
            ])
        return headers, rows

    def _sheet_crawl(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["source_id", "address_id", "url", "platform_type", "source_class",
                   "supplied_or_discovered", "independence_group", "independence_reason",
                   "retrieval_priority", "access_status", "crawl_status", "http_status",
                   "pages_opened", "documents_found", "images_found", "evidence_count",
                   "budget_pages", "budget_spent", "exhausted", "archive_checked",
                   "archive_snapshots", "archive_earliest", "discovery_method", "notes"]
        rows = [
            [r["source_id"], r["address_id"], r["url"], r["platform_type"], r["source_class"],
             r["supplied_or_discovered"], r["independence_group"], r["independence_reason"],
             r["retrieval_priority"], r["access_status"], r["crawl_status"], r["http_status"],
             r["pages_opened"], r["documents_found"], r["images_found"], r["evidence_count"],
             r["budget_pages"], r["budget_spent"], r["exhausted"], r["archive_checked"],
             r["archive_snapshot_count"], r["archive_earliest_snapshot"], r["discovery_method"],
             r["notes"]]
            for r in self.db.query(
                "SELECT * FROM sources WHERE community_id=? ORDER BY source_id", (community_id,))
        ]
        return headers, rows

    def _sheet_failures(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["error_id", "timestamp", "run_id", "stage", "source_id", "url", "error_type",
                   "http_status", "retry_count", "resolution", "unresolved",
                   "human_review_required", "detail"]
        rows = [
            [r["error_id"], r["ts_utc"], r["run_id"], r["stage"], r["source_id"], r["url"],
             r["error_type"], r["http_status"], r["retry_count"], r["resolution"],
             "yes" if r["unresolved"] else "no", "yes" if r["human_review"] else "no",
             (r["detail"] or "")[:1000]]
            for r in self.db.query(
                "SELECT * FROM errors WHERE community_id=? ORDER BY ts_utc", (community_id,))
        ]
        return headers, rows

    def _sheet_source_graph(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["source_a", "source_b", "relation", "similarity", "evidence", "created"]
        rows = [
            [r["source_a"], r["source_b"], r["relation"], r["similarity"], r["evidence"],
             r["created_utc"]]
            for r in self.db.query(
                "SELECT sr.* FROM source_relations sr JOIN sources s ON sr.source_a = s.source_id "
                "WHERE s.community_id = ?", (community_id,))
        ]
        if not rows:
            for r in self.db.query(
                "SELECT source_id, independence_group, independence_reason FROM sources "
                "WHERE community_id=? AND independence_group IS NOT NULL ORDER BY "
                "independence_group, source_id", (community_id,)
            ):
                rows.append([r["source_id"], r["independence_group"], "assigned_to_group", "",
                             r["independence_reason"], ""])
        return headers, rows

    def _sheet_review(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["item_id", "category", "severity", "subject", "detail", "related_ids",
                   "suggested_action", "created"]
        rows = [
            [r["item_id"], r["category"], r["severity"], r["subject"], r["detail"],
             r["related_ids"], r["suggested_action"], r["created_utc"]]
            for r in self.db.query(
                "SELECT * FROM review_queue WHERE community_id=? ORDER BY "
                "CASE severity WHEN 'blocking' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END, item_id",
                (community_id,))
        ]
        return headers, rows

    def _sheet_discovery(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["discovery_id", "stage", "method", "query", "found_url", "outcome", "detail",
                   "timestamp"]
        rows = [
            [r["discovery_id"], r["stage"], r["method"], r["query"], r["found_url"],
             r["outcome"], r["detail"], r["ts_utc"]]
            for r in self.db.query(
                "SELECT * FROM discovery_log WHERE community_id=? ORDER BY ts_utc LIMIT 5000",
                (community_id,))
        ]
        return headers, rows

    def _sheet_field_provenance(self, community_id: str) -> tuple[list[str], list[list[Any]]]:
        headers = ["field_name", "value", "status", "method", "independence_groups",
                   "group_count", "source_ids", "claim_ids", "residual_uncertainty",
                   "rationale", "updated"]
        rows = [
            [r["field_name"], r["value"], r["status"], r["method"], r["independence_groups"],
             r["group_count"], r["source_ids"], r["claim_ids"], r["residual_uncertainty"],
             r["rationale"], r["updated_utc"]]
            for r in self.db.query(
                "SELECT * FROM field_values WHERE community_id=? ORDER BY field_name",
                (community_id,))
        ]
        return headers, rows

    def _sheet_manifest(self, manifest: Mapping[str, Any]) -> tuple[list[str], list[list[Any]]]:
        headers = ["key", "value"]
        rows: list[list[Any]] = []
        for key, value in manifest.items():
            if isinstance(value, (dict, list)):
                rows.append([key, json.dumps(value, ensure_ascii=False, default=str)[:30000]])
            else:
                rows.append([key, value])
        for decision in self.decisions.get("decisions", []):
            rows.append([f"decision:{decision['id']}",
                         f"{decision.get('issue', '')} -> {decision.get('resolution', '')}"[:2000]])
        return headers, rows

    # -- lookups -----------------------------------------------------------
    def _field_values(self, community_id: str) -> dict[str, dict[str, Any]]:
        return {
            r["field_name"]: dict(r)
            for r in self.db.query(
                "SELECT * FROM field_values WHERE community_id = ?", (community_id,))
        }

    def _notes_for(self, community_id: str) -> str:
        pieces: list[str] = []
        truncated = self.db.query_one(
            "SELECT value, rationale FROM field_values WHERE community_id=? "
            "AND field_name='crawl_truncated'", (community_id,))
        if truncated and truncated["value"] == "yes":
            pieces.append(f"Crawl truncated: {truncated['rationale']}")
        review = self.db.scalar(
            "SELECT COUNT(*) FROM review_queue WHERE community_id=? AND severity='blocking'",
            (community_id,)) or 0
        if review:
            pieces.append(f"{review} blocking review items — see X8_Review_Queue")
        provenance = self.db.query_one(
            "SELECT provenance_mode FROM communities WHERE community_id=?", (community_id,))
        if provenance and provenance["provenance_mode"] == "FIXTURE":
            pieces.append("FIXTURE RUN — synthetic test data, not research evidence")
        return " | ".join(pieces)[:2000]

    def _source_title(self, source_id: str) -> str | None:
        row = self.db.query_one(
            "SELECT title FROM pages WHERE source_id=? AND title IS NOT NULL "
            "ORDER BY text_chars DESC LIMIT 1", (source_id,))
        return row["title"] if row else None

    def _source_access(self, record: Mapping[str, Any]) -> str:
        status = record["access_status"]
        if status in ("blocked", "login_required", "robots_denied"):
            return "unreachable"
        if status == "dead":
            return "unreachable"
        return "full text"

    def _full_text_status(self, document_id: str | None) -> str:
        if not document_id:
            return "full text"
        row = self.db.query_one(
            "SELECT text_status FROM documents WHERE document_id=?", (document_id,))
        if row is None:
            return "record only"
        return {"extracted": "full text", "ocr_used": "full text",
                "empty": "record only", "failed": "unreachable",
                "ocr_unavailable": "record only"}.get(row["text_status"], "record only")

    def _fields_from_source(self, community_id: str, source_id: str) -> str:
        rows = self.db.query(
            "SELECT DISTINCT field_name FROM claims WHERE community_id=? AND source_id=? "
            "ORDER BY field_name LIMIT 30", (community_id, source_id))
        names = [r["field_name"] for r in rows]
        return "; ".join(names) if names else "nothing"


def _coerce(value: Any, datatype: str | None = None) -> Any:
    """Write numbers as numbers.

    Values are stored as text in the database, but a column the workbook averages
    or correlates (managed area, population, a year) must arrive in Excel as a
    number: as text it is silently excluded from every calculation in O4.
    """
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value).strip()
    if datatype in ("float", "integer", "year", "year_or_na"):
        cleaned = text.replace(" ", "").replace(",", ".")
        try:
            number = float(cleaned)
        except ValueError:
            return text          # "NOT FOUND", "n/a": keep the wording
        if datatype in ("integer", "year", "year_or_na") or number.is_integer():
            return int(number)
        return number
    if len(text) > 32000:
        return text[:32000] + " […truncated]"
    return text


def _year_of(value: Any) -> int | None:
    import re

    match = re.search(r"\b(19\d{2}|20\d{2})\b", str(value or ""))
    return int(match.group(1)) if match else None


def _search_type(value: str | None) -> str:
    allowed = {"academic", "thesis portal", "grey - funding", "grey - government", "grey - NGO",
               "registry", "directory", "news", "archive"}
    return value if value in allowed else "directory"


def _worst_result(values: str | None) -> str:
    """One row per database: report the least favourable outcome it produced."""
    seen = {v.strip() for v in (values or "").split(",") if v.strip()}
    for candidate in ("hits found", "paywalled", "none found", "unreachable"):
        if candidate in seen:
            return candidate
    return "none found"
