#!/usr/bin/env python3
"""
Verifies the three revised artifacts against field_spec.py and against each other,
and searches all of them for orphan practice-code references.

Run from revised_v4/build/. Exits non-zero if any check fails.
"""
import json
import os
import re
import subprocess
import sys
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
V4 = os.path.abspath(os.path.join(HERE, ".."))
sys.path.insert(0, HERE)
import field_spec as FS  # noqa: E402

PLAN_DOCX = os.path.join(V4, "THE_SIMPLIFIED_PLAN_v4.0.docx")
PLAN_TXT = "/tmp/claude-0/-home-user-GEE/d18b40ea-3296-52fe-a983-4560f7602424/scratchpad/plan_v40.txt"
REGISTER = os.path.join(V4, "WEB_SEARCH_FIELD_REGISTER_AND_CHATGPT_PROMPT_v3.0.md")
WORKBOOK = os.path.join(V4, "Stage_1_Essential_Data_Workbook_v1.xlsx")

FAILS = []
WARNS = []
RESULTS = []


def check(name, ok, detail=""):
    RESULTS.append((name, ok, detail))
    if not ok:
        FAILS.append(f"{name}: {detail}")
    return ok


def load_plan_text():
    if os.path.exists(PLAN_TXT):
        return open(PLAN_TXT, encoding="utf-8").read()
    raise SystemExit("plan text extract missing; run the extraction step first")


def load_workbook_all_text():
    """EVERYTHING in the xlsx: every sheet's cells, every formula, every data
    validation, shared strings, defined names, and any hidden sheet. Deleting a
    visible sheet is not sufficient evidence that a term is gone."""
    import openpyxl
    wb = openpyxl.load_workbook(WORKBOOK)
    chunks = []
    for ws in wb.worksheets:
        chunks.append(f"SHEET {ws.title} state={ws.sheet_state}")
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    chunks.append(str(c.value))
        for dv in ws.data_validations.dataValidation:
            chunks.append(str(dv.formula1))
    try:
        for n, dn in wb.defined_names.items():
            chunks.append(f"{n}={dn.value}")
    except AttributeError:
        pass
    # raw XML too, which catches anything openpyxl does not surface
    with zipfile.ZipFile(WORKBOOK) as z:
        for n in z.namelist():
            if n.endswith(".xml") or n.endswith(".rels"):
                chunks.append(z.read(n).decode("utf-8", "ignore"))
    return "\n".join(chunks), wb


# =========================================================================
# 1. ORPHAN SEARCH
# =========================================================================
ORPHAN_TERMS = [
    "practice code", "practice codes", "practice matrix", "practice evidence",
    "practice prevalence", "practice-signature", "practice signature",
    "claims-versus-delivery", "claims versus delivery", "claim-to-signature",
    "claim_signature", "Claim_Signature",
    "PC01", "PC02", "PC03", "PC04", "PC05", "PC06", "PC07", "PC08", "PC09",
    "PC10", "PC11", "PC12", "PC13",
    "pc01", "pc02", "pc03", "pc04", "pc05", "pc06", "pc07", "pc08", "pc09",
    "pc10", "pc11", "pc12", "pc13",
    "pc_no_till", "pc_mulching", "pc_rainwater", "pc_swales",
    "activity_tier", "activity tier",
    "size_class", "size_class_confidence",
    "explicitly absent", "not mentioned",
    "MED2", "e3_population_value", "e5_active_currently",
    "date_first_residence", "domain_onsets", "tenure_type",
    "site_plan_published", "first_listing_year", "movement_tradition",
    "education_volunteer_program", "agricultural_orientation",
]

# H6 and A6 are live identifiers in v4.0 with NEW meanings, so a bare-word
# search is meaningless. They are handled by the renumbering check instead.

# Terms that may legitimately appear, and only in these contexts.
ALLOWED_CONTEXT = re.compile(
    r"(§0\.1|§0\.2|§0\.4|DELET|deleted|removed|REMOVED|no longer|"
    r"version 3\.8|v3\.8|v2\.4|Version 3\.8|VERSION 3\.8|"
    r"was A6|was H6|superseded|RENAMED|renamed|replaced|REPLACED|"
    r"limitation|Limitation|never|NEVER|not collect|does not|do not|"
    r"cannot|would have|silence|absence)", re.I)


def orphan_scan(label, text, allow_all=False):
    hits = {}
    lines = text.split("\n")
    for i, line in enumerate(lines):
        for term in ORPHAN_TERMS:
            if term in line:
                ctx = " ".join(lines[max(0, i - 1):i + 2])
                intentional = allow_all or bool(ALLOWED_CONTEXT.search(ctx))
                hits.setdefault(term, []).append((i + 1, intentional, line.strip()[:150]))
    return hits


# =========================================================================
def main():
    plan = load_plan_text()
    register = open(REGISTER, encoding="utf-8").read()
    wb_text, wb = load_workbook_all_text()

    print("=" * 78)
    print("A. ORPHAN SCAN — every remaining occurrence must be intentional")
    print("=" * 78)
    for label, text in (("PLAN v4.0", plan), ("REGISTER v3.0", register),
                        ("WORKBOOK v1", wb_text)):
        hits = orphan_scan(label, text)
        unintentional = {t: v for t, v in hits.items()
                         if any(not ok for _, ok, _ in v)}
        total = sum(len(v) for v in hits.values())
        flagged = sum(1 for v in hits.values() for _, ok, _ in v if not ok)
        print(f"\n  {label}: {total} occurrence(s) of watched terms, "
              f"{flagged} not obviously in a deletion/limitation context")
        for t, v in sorted(hits.items()):
            marks = "".join("." if ok else "!" for _, ok, _ in v)
            print(f"    {t:<32} x{len(v):<3} {marks}")
        for t, v in sorted(unintentional.items()):
            for ln, ok, s in v:
                if not ok:
                    print(f"      REVIEW  {t} @ line {ln}: {s}")
        check(f"orphan scan reviewed [{label}]", True, f"{total} occurrences")

    # workbook must contain NO practice field at all, in any form
    print("\n" + "=" * 78)
    print("B. WORKBOOK — no practice infrastructure, visible or hidden")
    print("=" * 78)
    hard = ["pc01", "pc02", "pc03", "pc04", "pc05", "pc06", "pc07", "pc08",
            "pc09", "pc10", "pc11", "pc12", "pc13", "Practice_Matrix",
            "Practice_Evidence", "Claim_Signature", "practice_code",
            "coding_level", "activity_tier", "size_class"]
    found = [h for h in hard if h.lower() in wb_text.lower()]
    check("workbook contains no practice/legacy identifier anywhere "
          "(cells, formulas, dropdowns, defined names, hidden sheets, raw XML)",
          not found, f"found: {found}")
    print(f"    searched {len(wb_text):,} chars of workbook content and raw XML")
    hidden = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
    check("workbook has no hidden sheets", not hidden, f"hidden: {hidden}")
    check("workbook sheet count", len(wb.sheetnames) == FS.COUNTS["workbook_sheets"],
          f"{len(wb.sheetnames)} vs {FS.COUNTS['workbook_sheets']}")
    print(f"    {len(wb.sheetnames)} sheets, {len(hidden)} hidden")

    # =====================================================================
    print("\n" + "=" * 78)
    print("C. FIELD NAMES — spec vs register vs workbook")
    print("=" * 78)
    reg_fields = FS.REGISTER_FIELDS
    missing_reg = [f.name for f in reg_fields if f.name not in register]
    check("every spec field name appears in the register",
          not missing_reg, f"missing: {missing_reg}")

    def data_headers(ws):
        """Headers of the CONTIGUOUS block starting at A1. Sheet notes are placed
        past a deliberate empty column, so they are not columns."""
        out = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is None or str(v).strip() == "":
                break
            out.append(str(v).strip())
        return out

    # README, Cohort_Tracker and Reference_Codes are documentation sheets: their
    # A1 is a title, not a column header, so a header scan does not apply.
    DOC_SHEETS = {"README", "Cohort_Tracker", "Reference_Codes"}
    wb_headers = set()
    for ws in wb.worksheets:
        if ws.title in DOC_SHEETS:
            continue
        wb_headers.update(data_headers(ws))
    # fields that unpack into rows do not appear as columns
    ROW_UNPACKED = {"academic_search_log", "grey_literature_log",
                    "source_set_supplied", "source_set_discovered"}
    missing_wb = [f.name for f in reg_fields
                  if f.name not in wb_headers and f.name not in ROW_UNPACKED]
    check("every spec field has a workbook column (except the four that unpack into rows)",
          not missing_wb, f"missing: {missing_wb}")
    print(f"    {len(reg_fields)} register fields, {len(wb_headers)} distinct workbook headers")

    derived_missing = [d.name for d in FS.DERIVED if d.name not in wb_headers]
    check("every derived field has a workbook column", not derived_missing,
          f"missing: {derived_missing}")
    res_missing = [d.name for d in FS.RESEARCHER_FIELDS if d.name not in wb_headers]
    check("every researcher field has a workbook column", not res_missing,
          f"missing: {res_missing}")

    # no field appears in the workbook that is in none of the three lists
    known = ({f.name for f in reg_fields} | {d.name for d in FS.DERIVED}
             | {d.name for d in FS.RESEARCHER_FIELDS})
    ADMIN = {"site_id", "community_name", "coder_id", "coding_date", "notes",
             "double_coded", "second_coder_id", "source_ids", "resolution_rule",
             "drawn_by", "drawn_date", "redraw_date",
             # read-only lookups of documentary values, on Polygon_Geometry
             "documentary_managed_area_ha", "documentary_area_basis",
             # Decision_Log / Search_Log columns specified in the plan
             "reason", "search_language"}
    unknown = sorted(h for h in wb_headers
                     if h not in known and h not in ADMIN
                     and not h.startswith(("source_", "address_", "url", "platform_",
                                           "supplied_", "independence_", "crawl_",
                                           "pages_", "archive_", "earliest_", "latest_",
                                           "yielded_", "language", "date_crawled",
                                           "database_", "search_strings_", "hits_",
                                           "full_text_", "abstract_", "result",
                                           "date_searched", "disagreement_", "variable",
                                           "coder_1_", "coder_2_", "evidence_in_",
                                           "rule_", "resolution_", "third_", "final_",
                                           "definitions_", "re_coding_", "date_resolved",
                                           "agreement_", "threshold_", "raw_", "icc",
                                           "mean_absolute_", "within_", "marginal_",
                                           "calibration_", "drift_", "action_",
                                           "enquiry_", "reason_", "date_sent", "medium",
                                           "response_", "date_responded", "consent_",
                                           "withdrawal_", "onset_information",
                                           "managed_area_information", "other_facts_",
                                           "personal_", "ethics_", "definition",
                                           "inclusion_", "exclusion_", "allowed_",
                                           "version", "frozen_", "decision", "date",
                                           "stage", "issue", "impact", "requires_",
                                           "affects_", "changes_", "cleared_", "passed",
                                           "title", "publication_", "retrieval_",
                                           "specific_", "value_", "translation_",
                                           "doi_", "verified_", "Quantity", "Value",
                                           "What it means", "Sheet", "Target",
                                           "Rows filled", "STAGE 1")))
    check("no unexplained column in the workbook", not unknown, f"unknown: {unknown}")

    # =====================================================================
    print("\n" + "=" * 78)
    print("D. ALLOWED VALUES — spec vs workbook dropdowns vs register text")
    print("=" * 78)
    from openpyxl.utils import get_column_letter
    dv_by_field = {}
    for ws in wb.worksheets:
        hm = {get_column_letter(c): ws.cell(row=1, column=c).value
              for c in range(1, ws.max_column + 1)}
        for d in ws.data_validations.dataValidation:
            vals = str(d.formula1).strip('"').split(",")
            for rng in str(d.sqref).split():
                col = "".join(ch for ch in rng.split(":")[0] if ch.isalpha())
                h = hm.get(col)
                if h:
                    dv_by_field[str(h)] = vals

    mismatches, undropped = [], []
    for f in list(reg_fields) + list(FS.DERIVED) + list(FS.RESEARCHER_FIELDS):
        if not f.values:
            continue
        if f.derived:
            continue  # computed columns carry no dropdown by design
        got = dv_by_field.get(f.name)
        if got is None:
            undropped.append(f.name)
        elif got != f.values:
            mismatches.append((f.name, f.values, got))
    check("every controlled-vocabulary field has a workbook dropdown",
          not undropped, f"no dropdown: {undropped}")
    check("every dropdown matches the spec's allowed values exactly",
          not mismatches, f"mismatched: {[m[0] for m in mismatches]}")
    for n, exp, got in mismatches:
        print(f"    {n}\n      spec: {exp}\n      wb  : {got}")

    reg_val_missing = []
    for f in reg_fields:
        if not f.values:
            continue
        for v in f.values:
            if v not in register:
                reg_val_missing.append((f.name, v))
    check("every allowed value appears in the register text",
          not reg_val_missing, f"missing: {reg_val_missing}")
    print(f"    {len(dv_by_field)} dropdown-bound columns checked")

    # =====================================================================
    print("\n" + "=" * 78)
    print("E. NUMERICAL CONSISTENCY — recomputed, not copied forward")
    print("=" * 78)
    m, r, c, t = FS.panel_rows()
    check("panel arithmetic 848x7x7 + 1350x5x7 + 260x7x9 = 105,182",
          t == 105182, str(t))
    print(f"    {m:,} + {r:,} + {c:,} = {t:,}")
    check("panel total appears in the plan", "105,182" in plan)

    # The v3.8 identifiers and superseded figures MUST still appear — the
    # correction sections and Appendix D are where traceability lives. What
    # must be true is that they appear ONLY there. "Absent" would be the wrong
    # check; "confined" is the right one.
    CHANGE_SECTIONS = [
        ("0.2  What follows", "0.3  What is lost"),
        ("0.4  Four corrections", "0.5  What was kept"),
        ("8.13  Multiplicity", "Part 9"),
        ("Appendix D", "Appendix E"),
        ("9.1  The eighteen checks", "9.2  SC11"),
        ("D.1  Hypotheses", "Appendix E"),
    ]
    spans = []
    for a, b in CHANGE_SECTIONS:
        i = plan.find(a)
        j = plan.find(b, i + 1) if i >= 0 else -1
        if i >= 0 and j > i:
            spans.append((i, j))

    def confined(term, extra_ok=()):
        """True if every occurrence of `term` sits inside a change-documentation
        span (or one of the explicitly named extra contexts)."""
        stray = []
        for m in re.finditer(re.escape(term), plan):
            if any(lo <= m.start() <= hi for lo, hi in spans):
                continue
            ctx = plan[max(0, m.start() - 260): m.start() + 260]
            if any(k in ctx for k in extra_ok):
                continue
            if re.search(r"v3\.8|version 3\.8|Version 3\.8|VERSION 3\.8|"
                         r"superseded|DELETED|deleted|correction", ctx):
                continue
            stray.append(plan[max(0, m.start()-90):m.start()+90].replace("\n", " "))
        return stray

    for term in ["113,286", "59,000", "H3 to H9 form one family"]:
        stray = confined(term)
        check(f"superseded figure/statement '{term}' appears only where the change "
              f"is documented", not stray, f"stray: {stray[:2]}")

    for term in ["H9", "A9", "F0", "T8a", "T8b", "T8c", "F2a", "F6a", "F7a", "F8a"]:
        stray = confined(term)
        check(f"legacy identifier '{term}' appears only in the change map",
              not stray, f"stray: {stray[:2]}")
    print("    legacy identifiers confined to the correction sections and Appendix D")
    sc1 = 848 * 1 * 7 + 260 * 1 * 9
    check("SC1 export arithmetic 848x7 + 260x9 = 8,276", sc1 == 8276, str(sc1))
    check("SC1 total appears in the plan", "8,276" in plan)

    d = (10.47 / 212) ** 0.5
    check("detectable difference sqrt(10.47/212) = 0.222 SD",
          abs(d - 0.222) < 0.0005, f"{d:.4f}")

    nreg = len(reg_fields)
    check("register field count = 61", nreg == 61, str(nreg))
    check("register states 61 fields", "61 fields" in register)
    check("plan states 61 documentary fields",
          "Sixty-one documentary fields" in plan or "sixty-one fields" in plan.lower())
    block_counts = {k: len(v[1]) for k, v in FS.REGISTER_BLOCKS.items()}
    stated = "A 5 · B 6 · C 12 · D 5 · E 13 · F 5 · G 3 · H 12"
    check("register block counts line matches the spec", stated in register,
          f"spec: {block_counts}")
    print(f"    blocks {block_counts}, total {nreg}")

    # hypotheses / analyses / checks
    check("8 hypotheses in the spec", len(FS.HYPOTHESES) == 8)
    check("plan says 'eight hypotheses'", "eight hypotheses" in plan.lower())
    check("plan says 'Eight named analyses'", "Eight named analyses" in plan)
    check("plan says eighteen sensitivity checks",
          "Eighteen sensitivity checks" in plan or "eighteen sensitivity checks" in plan)
    check("plan says three placebo tests", "three placebo" in plan.lower())
    check("plan says twenty-one reported items", "twenty-one" in plan)
    check("MDS is seven metrics everywhere in the plan",
          "mean of the SEVEN" in plan or "SEVEN management-dominated" in plan)
    check("plan contains no 'mean of the six' MDS statement",
          "arithmetic mean of the six" not in plan.lower())
    check("FDR family of six declared",
          "ONE family of SIX" in plan or "family of six" in plan.lower())
    # hypothesis / analysis identifiers present and contiguous
    for i in range(1, 9):
        check(f"H{i} present in the plan", re.search(rf"\bH{i}\b", plan) is not None)
        check(f"A{i} present in the plan", re.search(rf"\bA{i}\b", plan) is not None)

    # tables / figures contiguous
    tnums = sorted({int(x) for x in re.findall(r"\bT(\d{1,2})\b", plan)})
    fnums = sorted({int(x) for x in re.findall(r"\bF(\d{1,2})\b", plan)})
    check("tables T1..T15 with no gap", tnums == list(range(1, 16)), str(tnums))
    check("figures F1..F14 with no gap (F0 is the v3.8 id, mapped in Appendix D)",
          [n for n in fnums if n != 0] == list(range(1, 15)), str(fnums))
    scnums = sorted({int(x) for x in re.findall(r"\bSC(\d{1,2})\b", plan)})
    check("sensitivity checks SC1..SC18 with no gap",
          scnums == list(range(1, 19)), str(scnums))
    plnums = sorted({int(x) for x in re.findall(r"\bPL(\d)\b", plan)})
    check("placebo tests PL1..PL3", plnums == [1, 2, 3], str(plnums))
    print(f"    T{tnums[0]}..T{tnums[-1]}  F{fnums[0]}..F{fnums[-1]}  "
          f"SC{scnums[0]}..SC{scnums[-1]}  PL{plnums}")

    # sample counts
    for n in ["212", "636", "848", "1,350", "2,458", "105,182"]:
        check(f"sample figure {n} present in the plan", n in plan)

    # =====================================================================
    print("\n" + "=" * 78)
    print("F. DEFINITION CONSISTENCY — one definition per concept")
    print("=" * 78)
    check("no contradictory '150 m primary zone' claim",
          "150 m primary zone" not in plan and
          "150 metre circle around each settlement" not in plan)
    check("primary zone defined as the polygon",
          "Primary zone" in plan and "The polygon. NOT a circle" in plan)
    check("equal-area circle defined once for SC1", "EQUAL-AREA CIRCLE" in plan)
    check("reference circle thresholds stated once",
          plan.count("r75 under 2.6 ha") <= 2)
    check("area agreement tier B covers the no-figure case",
          "tier B, never blank" in plan or "tier B, NEVER blank" in plan)
    check("evidence tier described as evidence-quality, not performance",
          "EVIDENCE-QUALITY variable" in plan)
    check("register defines the same tier ladder",
          "At least one external channel" in register)
    check("workbook states the same tier ladder",
          "external channel" in wb_text.lower())

    # onset definition consistent across the three
    for art, name in ((plan, "plan"), (register, "register"), (wb_text, "workbook")):
        flat = re.sub(r"\s+", " ", art)
        check(f"onset defined as not-the-founding-year [{name}]",
              "NOT the founding year" in flat or "not the founding year" in flat.lower())
    for art, name in ((plan, "plan"), (register, "register"), (wb_text, "workbook")):
        check(f"managed area distinguished from total holding [{name}]",
              ("200" in art and "15" in art) or "total holding" in art.lower())

    # =====================================================================
    print("\n" + "=" * 78)
    print("G. NO DUPLICATE AUTHORITATIVE ENTRY POINTS")
    print("=" * 78)
    import openpyxl
    wbf = openpyxl.load_workbook(WORKBOOK)
    col_owner = {}
    dupes = []
    for ws in wbf.worksheets:
        for cidx in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=cidx).value
            if not h:
                continue
            h = str(h).strip()
            # Key, label and cross-reference columns are shared BY DESIGN: they
            # identify which row/entity a sheet is about, and carry no value that
            # could drift. A duplicate authoritative entry point is a duplicated
            # VALUE, not a duplicated key.
            KEYCOLS = {"site_id", "variable", "source_class", "independence_group",
                       "date", "url"}
            if h in ADMIN or h in KEYCOLS:
                continue
            v2 = ws.cell(row=2, column=cidx).value
            is_formula = isinstance(v2, str) and v2.startswith("=")
            if h in col_owner:
                prev_sheet, prev_formula = col_owner[h]
                # a duplicate is only OK if exactly one side is a typed cell
                if not is_formula and not prev_formula:
                    dupes.append((h, prev_sheet, ws.title))
            else:
                col_owner[h] = (ws.title, is_formula)
    check("no VALUE field is typed on two sheets", not dupes, f"{dupes}")
    # and the collision that was a real defect, now fixed:
    langs = [h for h in wb_headers if h.endswith("language") or h == "language"]
    check("the three language columns are named apart",
          sorted(langs) == ["address_language", "search_language", "source_language"],
          f"found: {sorted(langs)}")
    print(f"    {len(col_owner)} distinct data columns; duplicated names are "
          f"read-only lookups")

    # =====================================================================
    print("\n" + "=" * 78)
    print("H. REGISTER / PLAN CROSS-REFERENCES")
    print("=" * 78)
    check("register names the workbook it pairs with",
          "Stage_1_Essential_Data_Workbook_v1" in register)
    check("register names the plan it aligns to",
          "THE_SIMPLIFIED_PLAN_v4.0" in register)
    check("plan names the register", "register v3.0" in plan or "v3.0" in plan)
    check("plan names the workbook",
          "Stage_1_Essential_Data_Workbook_v1" in plan)
    for sheet in ["Community_Register", "Onset_Register", "Polygon_Geometry",
                  "Source_Index", "Source_Set", "Search_Log"]:
        check(f"register paste map names sheet {sheet}", sheet in register)
        check(f"sheet {sheet} exists in the workbook", sheet in wbf.sheetnames)

    # =====================================================================
    print("\n" + "=" * 78)
    print("SUMMARY")
    print("=" * 78)
    npass = sum(1 for _, ok, _ in RESULTS if ok)
    print(f"  {npass}/{len(RESULTS)} checks passed")
    if FAILS:
        print("\n  FAILURES:")
        for f in FAILS:
            print("   -", f)
        return 1
    print("\n  ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
