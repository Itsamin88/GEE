#!/usr/bin/env python3
"""
Builds Stage_1_Essential_Data_Workbook_v1.xlsx from field_spec.py.

Built from scratch. It is NOT a mutation of Stage_1_Documentary_Coding_Workbook_v6,
which is preserved untouched in originals/.

Design rules enforced here:
  * ONE authoritative entry point per variable. Where a value appears on a second
    sheet it is a read-only lookup formula, never a second typed cell.
  * Every derived quantity is a formula. Nothing that can be computed is typed.
  * No practice-code infrastructure anywhere: no practice sheet, no practice
    columns, no practice dropdowns, no practice named ranges, no hidden sheets.
  * Lookups run in one direction only: Polygon_Geometry reads Community_Register,
    Community_Register and Polygon_Geometry read Onset_Register for names.
    There is no cycle.
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from field_spec import V, COUNTS, panel_rows  # noqa: E402

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..",
                   "Stage_1_Essential_Data_Workbook_v1.xlsx")

HDR_FILL = PatternFill("solid", fgColor="1F3B4D")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
DERIVED_FILL = PatternFill("solid", fgColor="EDF3F7")
YELLOW_FILL = PatternFill("solid", fgColor="FFF3C4")
NOTE_FONT = Font(italic=True, color="5A6B7A", size=9)
TITLE_FONT = Font(bold=True, size=13, color="1F3B4D")
SUB_FONT = Font(bold=True, size=10, color="1F3B4D")
THIN = Border(bottom=Side(style="thin", color="C8D4DC"))

MAXROW = 402          # header + 400 data rows; 212 communities fits with room
DV_LAST = 402

wb = Workbook()
wb.remove(wb.active)


def sheet(name, headers, derived=(), notes=None, widths=None, freeze="B2"):
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = THIN
    ws.row_dimensions[1].height = 30
    for i, h in enumerate(headers, start=1):
        w = 18
        if widths and h in widths:
            w = widths[h]
        elif len(h) > 24:
            w = min(34, len(h) + 4)
        ws.column_dimensions[get_column_letter(i)].width = w
    if derived:
        for i, h in enumerate(headers, start=1):
            if h in derived:
                for r in range(2, MAXROW + 1):
                    ws.cell(row=r, column=i).fill = DERIVED_FILL
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    if notes:
        col = len(headers) + 2
        ws.cell(row=1, column=col, value=notes[0]).font = SUB_FONT
        ws.column_dimensions[get_column_letter(col)].width = 96
        for j, line in enumerate(notes[1:], start=2):
            c = ws.cell(row=j, column=col, value=line)
            c.font = NOTE_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def dv(ws, col_letters, values, last=DV_LAST):
    """Attach a list dropdown. Values are the exact allowed set."""
    formula = '"' + ",".join(values) + '"'
    assert len(formula) <= 255, f"dropdown too long for {col_letters}: {len(formula)}"
    d = DataValidation(type="list", formula1=formula, allow_blank=True,
                       showDropDown=False)
    d.error = "Use the dropdown. Free-typed variants corrupt every count."
    d.errorTitle = "Value not in the controlled vocabulary"
    ws.add_data_validation(d)
    for cl in col_letters:
        d.add(f"{cl}2:{cl}{last}")


def fill(ws, col_letter, formula_template, first=2, last=MAXROW):
    """Write a per-row formula down a column. {r} is substituted with the row."""
    for r in range(first, last + 1):
        ws[f"{col_letter}{r}"] = formula_template.format(r=r)


# ==========================================================================
# 1. README
# ==========================================================================
ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 104
ws.column_dimensions["C"].width = 22

readme = [
    ("T", "STAGE 1 — ESSENTIAL DATA WORKBOOK", ""),
    ("S", "Version 1.0 · aligned to THE_SIMPLIFIED_PLAN_v4.0 and "
          "WEB_SEARCH_FIELD_REGISTER_AND_CHATGPT_PROMPT_v3.0", ""),
    ("", "", ""),
    ("H", "WHAT THIS WORKBOOK IS", ""),
    ("R", "Purpose",
     "Stage 1 produces two things: the ESSENTIAL DOCUMENTARY DATA the rest of the study "
     "consumes, and the MEASUREMENT GEOMETRY the whole study is measured over. Nothing "
     "else. Every column here has a named downstream consumer in the plan; a column with "
     "no consumer was deleted rather than kept for completeness."),
    ("R", "What it replaces",
     "Stage_1_Documentary_Coding_Workbook_v6. That workbook is preserved unchanged. This "
     "one is rebuilt from scratch, not edited: the practice-code sheets are absent rather "
     "than hidden, and no practice column, dropdown, formula or named range survives "
     "anywhere in this file."),
    ("R", "What was removed and why",
     "Documentary practice codes (the thirteen pc* codes), the practice matrix, the "
     "practice evidence sheet and the claim-to-signature map are gone. Practice-level "
     "documentary information could not be established with sufficient completeness, "
     "consistency or independent verification across 212 communities to support "
     "quantitative inference. See DECISION_MEMO.md."),
    ("", "", ""),
    ("H", "THE ONE RULE", ""),
    ("R", "You are recording what sources SAY",
     "You are documenting what communities say, what independent sources say, and what can "
     "be established historically. You are NOT assessing ecological performance — the "
     "satellite pipeline does that, later, and it must not be influenced by anything here."),
    ("R", "Blindness",
     "Do not view satellite metrics, index values or any outcome data while coding or "
     "drawing. Historical high-resolution imagery is permitted and required for the "
     "polygon; derived vegetation quantities are not."),
    ("R", "Boundary with Stage 2",
     "Stage 1 produces coded facts. Stage 2 consumes them to build the matched control "
     "sample. Never record an eligibility decision here that Stage 2 should be making."),
    ("", "", ""),
    ("H", "SHEETS, AND WHAT EACH OWNS", ""),
    ("R", "Community_Register", "OWNS identity, eligibility, population, the documentary area "
     "figures, the five evidence channels, status, context and crawl provenance."),
    ("R", "Onset_Register", "OWNS every onset field and cohort candidacy. The priority block."),
    ("R", "Polygon_Geometry", "OWNS the drawn polygon and its reliability. READS the "
     "documentary area from Community_Register to compute the agreement tier."),
    ("R", "Source_Index", "One row per source, with its independence group and a mandatory "
     "verified_resolves check."),
    ("R", "Source_Set", "One row per web address per community."),
    ("R", "Search_Log", "One row per database per community, INCLUDING those that returned "
     "nothing. This is where 'evidence absent' is separated from 'not looked for'."),
    ("R", "Disagreement_Log", "Every double-coding disagreement and how it was resolved."),
    ("R", "Reliability_Report", "Agreement per variable. Pre-filled with the fourteen "
     "variables the plan requires."),
    ("R", "Enquiry_Record", "Only if you contact communities. Consent and ethics reference."),
    ("R", "Definitions_And_Freeze", "The definitions and agreement thresholds for every "
     "judgement-bearing variable, with ONE freeze date. Replaces the practice codebook."),
    ("R", "Calibration", "A coder proceeds to production only after meeting the threshold."),
    ("R", "Decision_Log", "Appendix A of the plan. One row per decision, with a date."),
    ("R", "Cohort_Tracker", "Decision D4, computed live."),
    ("R", "Reference_Codes", "The full vocabulary with the rule behind each value, and the "
     "list of quantities that must NEVER be searched for."),
    ("", "", ""),
    ("H", "HOW TO USE IT", ""),
    ("R", "Dark header row", "Every sheet has one, with an auto-filter. Type beneath it."),
    ("R", "Pale blue cells", "Computed. They contain formulas. Do not type into them — a typed "
     "value there is a second authoritative entry point and the two will drift."),
    ("R", "Yellow cells", "You must fill them: the freeze date on Definitions_And_Freeze."),
    ("R", "Dropdowns", "Every controlled-vocabulary column has one. Use it. 'Evidenced' and "
     "'evidenced' are different strings and will corrupt every count."),
    ("R", "No example rows", "Row 2 is empty and ready. The format is in Reference_Codes and "
     "in Definitions_And_Freeze, so there is no example row to forget to delete."),
    ("", "", ""),
    ("H", "SEQUENCE — do not start at IC001", ""),
    ("R", "1  Freeze", "Complete Definitions_And_Freeze, confirm every agreement threshold, "
     "and enter the freeze date. After that date the definitions may not change; a later "
     "change means the affected variable is re-coded IN FULL."),
    ("R", "2  Design", "Draw the 20 per cent double-coding subsample with a recorded seed, "
     "stratified by region and expected evidence tier. Record the seed in Decision_Log."),
    ("R", "3  Calibrate", "Train on communities NOT in the study sample. Record in "
     "Calibration. Proceed only after meeting the declared threshold."),
    ("R", "4  Code", "Fill Source_Set for a community before coding it, so you know what you "
     "are working from. Then Community_Register and Onset_Register."),
    ("R", "5  Draw", "Draw the polygon at every settlement, from STRUCTURE and never from "
     "colour. Translate it onto all three controls. Record in Polygon_Geometry."),
    ("R", "6  Decide", "At community 100, read Cohort_Tracker and make decision D4."),
    ("R", "7  Close", "Re-code the 20 per cent after at least four weeks. Compute reliability "
     "PER VARIABLE. Freeze and version the dataset."),
    ("", "", ""),
    ("H", "SEVEN RULES MOST OFTEN BROKEN", ""),
    ("R", "Founding year is not onset", "A community founded in 1985 that began ecological "
     "work in 1992 has an onset of 1992. Record both."),
    ("R", "Managed area is not total holding", "A community holding 200 ha and working 15 ha "
     "has a managed area of 15 ha. Record both, separately."),
    ("R", "Unknown is not dissolved", "A vanished website is not a vanished community. "
     "Dissolution requires POSITIVE evidence."),
    ("R", "Absence of evidence is not evidence of absence", "A field left as 'not found' after "
     "an exhaustive search and a field left as 'not found' because the run was truncated look "
     "identical. crawl_truncated is what tells them apart. Fill it honestly."),
    ("R", "A copy is not a second source", "Four addresses belonging to one community are one "
     "voice. Corroboration counts independence GROUPS, never URLs."),
    ("R", "Trace structure, never colour", "When drawing a polygon use field edges, tracks, "
     "terraces and changes in field pattern. Never greenness. A drawing bias does not cancel "
     "between the two arms, because the control polygon is a translation of the settlement's."),
    ("R", "The evidence tier is not a performance score", "evidence_tier records how well "
     "documented a community is. It does not record how much ecological work it does, and it "
     "must never be reported or interpreted as if it did."),
    ("", "", ""),
    ("H", "PROGRESS", ""),
]
r = 1
for kind, a, b in readme:
    if kind == "T":
        c = ws.cell(row=r, column=1, value=a); c.font = TITLE_FONT
    elif kind == "S":
        c = ws.cell(row=r, column=1, value=a); c.font = NOTE_FONT
    elif kind == "H":
        c = ws.cell(row=r, column=1, value=a); c.font = SUB_FONT
        c.fill = PatternFill("solid", fgColor="E3ECF2")
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="E3ECF2")
        ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="E3ECF2")
    elif kind == "R":
        ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=9)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=9)
    r += 1

prog_start = r
ws.cell(row=r, column=1, value="Sheet").font = Font(bold=True, size=9)
ws.cell(row=r, column=2, value="Target").font = Font(bold=True, size=9)
ws.cell(row=r, column=3, value="Rows filled").font = Font(bold=True, size=9)
r += 1
for sh, target in [("Community_Register", "212 communities"),
                   ("Onset_Register", "212 communities"),
                   ("Polygon_Geometry", "212 communities"),
                   ("Source_Set", "as required"),
                   ("Source_Index", "as required"),
                   ("Search_Log", "as required"),
                   ("Disagreement_Log", "as required")]:
    ws.cell(row=r, column=1, value=sh).font = Font(size=9)
    ws.cell(row=r, column=2, value=target).font = Font(size=9)
    ws.cell(row=r, column=3, value=f"=COUNTA({sh}!A2:A{MAXROW})")
    r += 1
ws.cell(row=r + 1, column=2,
        value="Counts read 0 until you begin and 212 when Stage 1 is complete. There is no "
              "example row to subtract.").font = NOTE_FONT

# ==========================================================================
# 2. Cohort_Tracker
# ==========================================================================
ws = wb.create_sheet("Cohort_Tracker")
ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 88
ws.cell(row=1, column=1, value="LONGITUDINAL COHORT TRACKER — decision D4").font = TITLE_FONT
ws.cell(row=2, column=1,
        value="Analysis A7 needs about 65 communities with onset in 2020 or 2021. Decide at "
              "community 100, not at the end.").font = NOTE_FONT
hdr = ["Quantity", "Value", "What it means"]
for i, h in enumerate(hdr, start=1):
    c = ws.cell(row=3, column=i, value=h); c.fill = HDR_FILL; c.font = HDR_FONT
rows = [
    ("Communities coded so far", f"=COUNTA(Community_Register!A2:A{MAXROW})",
     "Decision D4 is taken at 100."),
    ("cohort_candidate = core (2020-2021)", f'=COUNTIF(Onset_Register!Q2:Q{MAXROW},"core (2020-2021)")',
     "The pool of candidates."),
    ("  of those, confidence tier A or B", "PLACEHOLDER",
     "THE NUMBER THAT MATTERS. Tier C onsets are too uncertain to anchor a "
     "before-and-after window."),
    ("cohort_candidate = extension (2019)", f'=COUNTIF(Onset_Register!Q2:Q{MAXROW},"extension (2019)")',
     "Usable if the core falls short — costs one pre-onset year."),
    ("cohort_candidate = uncertain", f'=COUNTIF(Onset_Register!Q2:Q{MAXROW},"uncertain")',
     "Resolve these before deciding."),
    ("DECISION D4 STATUS", "PLACEHOLDER2",
     "Thresholds from the plan: counted on core AND tier A or B. Above 50 proceed; "
     "40 to 50 proceed but report as exploratory; below 40 drop the component."),
]
r = 4
for a, b, c in rows:
    ws.cell(row=r, column=1, value=a).font = Font(size=9, bold=(a.startswith("DECISION")))
    if b == "PLACEHOLDER":
        b = (f'=COUNTIFS(Onset_Register!Q2:Q{MAXROW},"core (2020-2021)",Onset_Register!O2:O{MAXROW},"A")'
             f'+COUNTIFS(Onset_Register!Q2:Q{MAXROW},"core (2020-2021)",Onset_Register!O2:O{MAXROW},"B")')
    if b == "PLACEHOLDER2":
        b = (f'=IF(COUNTA(Community_Register!A2:A{MAXROW})<100,'
             f'"KEEP CODING — decide at community 100",'
             f'IF($B$6>=50,"PROCEED — the component is well worth its four to six weeks",'
             f'IF($B$6>=40,"PROCEED BUT REPORT AS EXPLORATORY",'
             f'"DROP THE COMPONENT — an underpowered before-and-after test does more harm than omitting it")))')
    ws.cell(row=r, column=2, value=b)
    ws.cell(row=r, column=3, value=c).font = Font(size=9)
    ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
ws.freeze_panes = "A4"

# ==========================================================================
# 3. Community_Register  — 55 columns
# ==========================================================================
CR = [
    # identity + held coordinates
    "site_id", "community_name_official", "alternative_names", "country", "admin_region",
    "latitude", "longitude", "coordinate_agreement",
    # eligibility
    "e1_network_listing", "e1_pathway", "e1_self_identification",
    "e2_settlement_type", "e2_evidence_note", "e8_setting_at_onset",
    # population
    "population_value", "population_lower", "population_upper", "population_source_date",
    # documentary area
    "managed_area_ha", "managed_area_lower_ha", "managed_area_upper_ha",
    "managed_area_basis", "managed_area_source_class", "documentary_area_note",
    "total_holding_ha", "area_type", "parcel_structure",
    # evidence verification
    "v1_self_documentation", "v2_external_documentation", "v3_substantive_affiliation",
    "v4_visual_documentation", "v5_continuity_evidence", "channel_count", "evidence_tier",
    # status
    "status_current", "status_evidence", "last_listing_year", "dissolution_year",
    "delisting_reason",
    # context
    "external_funding_or_programme", "protected_area_status", "notable_context",
    # provenance
    "pages_opened_count", "documents_opened", "source_classes_found", "search_languages",
    "negative_consultations", "independence_groups", "stages_completed", "crawl_truncated",
    # coding admin
    "coder_id", "coding_date", "double_coded", "second_coder_id", "notes",
]
ws = sheet("Community_Register", CR,
           derived={"channel_count", "evidence_tier"},
           widths={"site_id": 10, "latitude": 11, "longitude": 11, "notes": 40,
                   "e1_self_identification": 34, "e2_evidence_note": 34,
                   "documentary_area_note": 34, "status_evidence": 30,
                   "negative_consultations": 34, "external_funding_or_programme": 30,
                   "notable_context": 28},
           notes=[
               "ONE ROW PER COMMUNITY. THE SINGLE PASTE TARGET.",
               "Field names are identical to the web-search register v3.0, so a returned "
               "record pastes straight in.",
               "",
               "channel_count and evidence_tier are COMPUTED from v1..v5. Do not type them. "
               "In v6 they were typed and could disagree with the channels behind them.",
               "",
               "evidence_tier is an EVIDENCE-QUALITY variable. It records how well documented "
               "a community's ecological work is, not how much of it there is. It is used as a "
               "named confounder in analysis A4, as a sample-description variable, and behind "
               "sensitivity check SC18. It must never be reported as an ecological-performance "
               "measure, and it must never be disaggregated by KIND of activity — doing that "
               "would rebuild a practice score by another name.",
               "",
               "The tier rule, ordered by INDEPENDENCE and then by amount: Fail = fewer "
               "than two channels; C = two or more channels but NO external one (V2, V3); "
               "B = an external channel and two channels in total; A = an external channel and "
               "three or more. v2.4 ordered by count first, which made its tier C unreachable "
               "and ranked three self-documented channels above two including a thesis.",
               "",
               "managed_area_ha is the worked area, NEVER the total holding. Both are recorded "
               "separately. Polygon_Geometry reads managed_area_ha from column S of this sheet "
               "to compute the agreement tier; it is entered here and nowhere else.",
               "",
               "population_value appears once. In v2.4 of the register the same quantity was "
               "collected twice, as e3_population_value and population_value, which is two "
               "authoritative entry points for one number.",
               "",
               "crawl_truncated = yes is not a failure. A truncated run that says so is useful "
               "data; a truncated run that reads as a complete one is worse than no run at all.",
           ])
dv(ws, ["H"], V["coordinate_agreement"])
dv(ws, ["J"], V["e1_pathway"])
dv(ws, ["L"], V["e2_settlement_type"])
dv(ws, ["N"], V["e8_setting_at_onset"])
dv(ws, ["V"], V["managed_area_basis"])
dv(ws, ["W"], V["source_class"])
dv(ws, ["Z"], V["area_type"])
dv(ws, ["AA"], V["parcel_structure"])
dv(ws, ["AB", "AC", "AD", "AE", "AF"], V["yes_no"])
dv(ws, ["AI"], V["status_current"])
dv(ws, ["AM"], V["delisting_reason"])
dv(ws, ["AO"], V["protected_area_status"])
dv(ws, ["AX"], V["yes_no"])          # crawl_truncated
dv(ws, ["BA"], V["yes_no"])          # double_coded
# channel_count (AG) and evidence_tier (AH)
fill(ws, "AG", '=IF(COUNTA(AB{r}:AF{r})=0,"",COUNTIF(AB{r}:AF{r},"yes"))')
# evidence_tier. Ordered by INDEPENDENCE of the evidence, then by how much of
# it there is. The v2.4 ladder was ordered by count first, which made tier C
# ("2 community-originated channels") unreachable — the only channel that is
# neither external nor visual nor continuity is V1, and a community cannot have
# two of V1. It also ranked three self-documented channels above two including
# a thesis, which inverts what the tier is for.
fill(ws, "AH",
     '=IF($AG{r}="","",'
     'IF($AG{r}<2,"Fail",'
     'IF(OR($AC{r}="yes",$AD{r}="yes"),IF($AG{r}>=3,"A","B"),"C")))')
# guard: 3+ channels but no external documentation cannot be tier A
ws.cell(row=1, column=len(CR) + 2)

# --------------------------------------------------------------------------
# Guarded cross-sheet lookup.
# INDEX over an EMPTY cell returns 0, not "". Unguarded, a site with no
# documentary managed area produced area_ratio = 0 and area_agreement_tier = C
# — "the two disagree by more than a factor of two" — when in fact there is no
# documentary figure to disagree with. The plan requires tier B there. The
# inner IF is what makes an absent value read as absent.
# --------------------------------------------------------------------------
LOOKUP_TMPL = (
    '=IF($A{{r}}="","",IFERROR('
    'IF(INDEX(Community_Register!${col}$2:${col}$' + str(MAXROW) + ','
    'MATCH($A{{r}},Community_Register!$A$2:$A$' + str(MAXROW) + ',0))="","",'
    'INDEX(Community_Register!${col}$2:${col}$' + str(MAXROW) + ','
    'MATCH($A{{r}},Community_Register!$A$2:$A$' + str(MAXROW) + ',0))),""))'
)

# ==========================================================================
# 4. Onset_Register — 21 columns
# ==========================================================================
ORG = [
    "site_id", "community_name", "date_formal_founding", "founding_decade",
    "date_land_acquisition", "date_intervention_onset",
    "onset_lower_bound", "onset_upper_bound", "onset_band_width_years",
    "onset_evidence_rank", "onset_evidence_description", "onset_conflicting_sources",
    "resolution_rule", "onset_proxy_flag", "onset_confidence_tier",
    "onset_first_or_major", "cohort_candidate", "source_ids",
    "coder_id", "coding_date", "notes",
]
ws = sheet("Onset_Register", ORG,
           derived={"community_name", "founding_decade", "onset_band_width_years"},
           widths={"site_id": 10, "onset_evidence_description": 38,
                   "onset_conflicting_sources": 34, "resolution_rule": 34, "notes": 34},
           notes=[
               "THE PRIORITY BLOCK. More analyses depend on onset than on any other "
               "documentary field, and it is the hardest field to establish.",
               "",
               "Three candidate dates are retained and only ONE of them is the study's onset. "
               "date_formal_founding distinguishes founding from onset — the study's central "
               "dating rule. date_land_acquisition is the hardest available LOWER bound: a "
               "community cannot intervene on land it does not hold. "
               "date_intervention_onset is the onset.",
               "",
               "date_first_residence and domain_onsets were dropped in v4.0. Neither had a "
               "downstream consumer: habitation neither bounds nor dates the intervention, and "
               "with practice codes removed there is no per-domain analysis for a per-domain "
               "onset to feed. The onset is the EARLIEST deliberate ecological intervention of "
               "any kind.",
               "",
               "Where sources of equal rank disagree, take the EARLIER year as the value and "
               "let the gap become the lower and upper bounds. Record the disagreement in "
               "onset_conflicting_sources whatever you do with it, and the rule you applied in "
               "resolution_rule. A conflict is never resolved silently.",
               "",
               "onset_proxy_flag = yes means a founding year was used as a substitute. Those "
               "sites leave analysis A4 and stay in every other analysis.",
               "",
               "onset_confidence_tier gates cohort admission: A7 admits tier A and B only.",
           ])
dv(ws, ["J"], V["onset_evidence_rank"])
dv(ws, ["N"], V["yes_no"])
dv(ws, ["O"], V["onset_confidence_tier"])
dv(ws, ["P"], V["onset_first_or_major"])
dv(ws, ["Q"], V["cohort_candidate"])
fill(ws, "B", LOOKUP_TMPL.format(col="B"))
fill(ws, "D", '=IF(OR($C{r}="",NOT(ISNUMBER($C{r}))),"",FLOOR($C{r},10))')
fill(ws, "I", '=IF(OR($G{r}="",$H{r}="",NOT(ISNUMBER($G{r})),NOT(ISNUMBER($H{r}))),"",$H{r}-$G{r})')

# ==========================================================================
# 5. Polygon_Geometry — 23 columns
# ==========================================================================
PG = [
    "site_id", "community_name", "polygon_area_ha", "polygon_file_id",
    "polygon_imagery_date", "polygon_imagery_source", "polygon_confidence",
    "below_minimum_flag", "reference_circle", "equal_area_circle_radius_m",
    "polygon_redrawn", "redraw_area_ha", "polygon_iou",
    "documentary_managed_area_ha", "documentary_area_basis",
    "area_ratio", "area_agreement_tier", "agreement_note",
    "controls_translated", "drawn_by", "drawn_date", "redraw_date", "notes",
]
ws = sheet("Polygon_Geometry", PG,
           derived={"community_name", "below_minimum_flag", "reference_circle",
                    "equal_area_circle_radius_m", "documentary_managed_area_ha",
                    "documentary_area_basis", "area_ratio", "area_agreement_tier"},
           widths={"site_id": 10, "agreement_note": 40, "notes": 36,
                   "equal_area_circle_radius_m": 22, "documentary_managed_area_ha": 24},
           notes=[
               "THE MEASUREMENT GEOMETRY. This sheet, not the documentary area, defines the "
               "ground every metric is computed over.",
               "",
               "YOU supply: polygon_area_ha, polygon_file_id, the imagery date and source, "
               "polygon_confidence, the redraw fields, agreement_note and controls_translated. "
               "Everything else on this sheet computes itself.",
               "",
               "The four documentary columns are READ-ONLY LOOKUPS from Community_Register. "
               "They cannot be typed and cannot drift. Enter the documentary area once, there.",
               "",
               "reference_circle is matched to the polygon's AREA — r75 under 2.6 ha, r110 to "
               "5.2, r150 to 9.9, r210 to 19.8, r300 above. Rescaling a 2-hectare polygon "
               "against a 7-hectare reference distribution biases every score at that site, "
               "and nothing downstream would reveal it.",
               "",
               "equal_area_circle_radius_m is sqrt(area/pi) in metres. It is the alternative "
               "measurement geometry for sensitivity check SC1: a circle of the SAME AREA as "
               "the polygon, so that SC1 isolates the effect of SHAPE alone. It is computed "
               "here rather than by hand.",
               "",
               "area_agreement_tier: A within 30 per cent, B where there is no documentary "
               "figure or a gap of 30 to 100 per cent, C beyond a factor of two. A site with a "
               "polygon and no documentary figure is tier B, never blank. The polygon stands in "
               "every case — the tier records how well corroborated it is, not which number wins.",
               "",
               "A gap beyond a factor of two usually has one cause: the source quoted the TOTAL "
               "HOLDING. The second commonest is a genuinely non-contiguous holding whose "
               "detached parcels the polygon excludes. parcel_structure on Community_Register "
               "is what separates the two. Record what you found in agreement_note.",
               "",
               "MINIMUM 1.0 HECTARE. Below it the site falls back to a 75 m circle and is "
               "flagged; below_minimum_flag says so automatically.",
           ])
dv(ws, ["F"], V["polygon_imagery_source"])
dv(ws, ["G"], V["polygon_confidence"])
dv(ws, ["K", "S"], V["yes_no"])
M = str(MAXROW)
fill(ws, "B", LOOKUP_TMPL.format(col="B"))
fill(ws, "H", '=IF($C{r}="","",IF($C{r}<1,"yes","no"))')
fill(ws, "I", '=IF($C{r}="","",IF($C{r}<2.6,"r75",IF($C{r}<5.2,"r110",'
              'IF($C{r}<9.9,"r150",IF($C{r}<19.8,"r210","r300")))))')
fill(ws, "J", '=IF($C{r}="","",ROUND(SQRT($C{r}*10000/PI()),1))')
fill(ws, "N", LOOKUP_TMPL.format(col="S"))
fill(ws, "O", LOOKUP_TMPL.format(col="V"))
fill(ws, "P", '=IF(OR($C{r}="",$N{r}="",NOT(ISNUMBER($N{r}))),"",$N{r}/$C{r})')
fill(ws, "Q", '=IF($C{r}="","",IF(OR($N{r}="",NOT(ISNUMBER($N{r}))),"B",'
              'IF(ABS($P{r}-1)<=0.3,"A",IF(OR($P{r}>2,$P{r}<0.5),"C","B"))))')

# ==========================================================================
# 6. Source_Index — 17 columns
# ==========================================================================
SI = ["source_id", "site_id", "source_class", "source_type", "title",
      "url_or_reference", "publication_or_snapshot_date", "retrieval_date",
      "full_text_or_abstract", "specific_passage_or_page", "value_supported",
      "source_language", "translation_used", "doi_or_identifier", "verified_resolves",
      "independence_group", "notes"]
ws = sheet("Source_Index", SI,
           widths={"title": 38, "url_or_reference": 40, "value_supported": 38, "notes": 30},
           notes=[
               "ONE ROW PER SOURCE.",
               "",
               "verified_resolves is MANDATORY for every S1 and S2 source. Open the DOI or the "
               "repository record yourself. A fabricated academic citation looks like the "
               "strongest evidence in the record and will be believed.",
               "",
               "independence_group is a short id (G1, G2, G3) shared by every source deriving "
               "from the same underlying statement. Assign it as you read, not afterwards. A "
               "community's website, its own Facebook page and a directory listing copied from "
               "it are ONE group, however many URLs. A thesis, a grant record and a newspaper "
               "are three.",
               "",
               "The test: could this source be wrong in the same way as that one, for the same "
               "reason? If yes, same group.",
               "",
               "source_language is the language THIS SOURCE is written in. It is deliberately "
               "not called 'language': Source_Set records the language of an ADDRESS and "
               "Search_Log the language a SEARCH was run in, and three different quantities "
               "sharing one column name is how a merge silently produces nonsense.",
           ])
dv(ws, ["C"], V["source_class"])
dv(ws, ["I"], V["full_text_or_abstract"])
dv(ws, ["M", "O"], V["yes_no"])

# ==========================================================================
# 7. Source_Set — 17 columns
# ==========================================================================
SS = ["site_id", "address_id", "url", "platform_type", "supplied_or_discovered",
      "independence_group", "crawl_status", "pages_opened", "archive_checked",
      "archive_earliest_snapshot", "earliest_dated_item", "latest_dated_item",
      "yielded_fields", "address_language", "date_crawled", "coder_id", "notes"]
ws = sheet("Source_Set", SS,
           widths={"url": 44, "yielded_fields": 36, "notes": 30},
           notes=[
               "ONE ROW PER WEB ADDRESS PER COMMUNITY. Fill it BEFORE coding the community.",
               "",
               "Most communities have several addresses and they are not equivalent. A "
               "community's current site says what it wants said today; an abandoned domain, an "
               "old album or a directory listing captured in 2013 says what was true THEN. This "
               "study is about DATING, so the old material is usually worth more.",
               "",
               "crawl_status must never be left as 'not attempted' when a community is closed. "
               "An address that silently vanished from a search assistant's output was not "
               "crawled, whatever the rest of the response implies.",
               "",
               "'blocked' is a real and reportable result. A guess about the content of a "
               "platform you could not open is fabrication.",
           ])
dv(ws, ["D"], V["platform_type"])
dv(ws, ["E"], V["supplied_or_discovered"])
dv(ws, ["G"], V["crawl_status"])
dv(ws, ["I"], V["yes_no"])

# ==========================================================================
# 8. Search_Log — 12 columns
# ==========================================================================
SL = ["site_id", "database_or_source", "database_type", "search_strings_used",
      "search_language", "hits_returned", "full_text_opened", "abstract_only",
      "result", "date_searched", "coder_id", "notes"]
ws = sheet("Search_Log", SL,
           widths={"search_strings_used": 44, "notes": 32},
           notes=[
               "ONE ROW PER DATABASE PER COMMUNITY — INCLUDING EVERY DATABASE THAT RETURNED "
               "NOTHING.",
               "",
               "This is the sheet that separates the three states the study must never "
               "confuse: EVIDENCE FOUND, EVIDENCE ABSENT after an adequately exhaustive "
               "search, and SEARCH INCOMPLETE. A row saying 'none found' in a named database "
               "is a finding. A missing row is not.",
               "",
               "Stages 5 and 6 — academic and grey literature — are now the study's "
               "HIGHEST-PRIORITY evidence route, because almost every rank-1 onset record "
               "lives there: a thesis that dates the fieldwork, a grant report that dates the "
               "project, a planning permit.",
               "",
               "Expect most academic searches to return nothing. In this population a handful "
               "of communities are well studied and the large majority have never been written "
               "about at all. Finding none is the normal and correct outcome.",
           ])
dv(ws, ["C"], V["database_type"])
dv(ws, ["I"], V["search_result"])

# ==========================================================================
# 9. Disagreement_Log — 14 columns
# ==========================================================================
DL = ["disagreement_id", "site_id", "variable", "coder_1_value", "coder_2_value",
      "evidence_in_question", "rule_invoked", "resolution_type", "third_coder",
      "final_value", "definitions_amended", "re_coding_required", "date_resolved", "notes"]
ws = sheet("Disagreement_Log", DL,
           widths={"evidence_in_question": 40, "rule_invoked": 40, "notes": 30},
           notes=["Every double-coding disagreement and its resolution.",
                  "",
                  "If definitions_amended = yes, the affected variable is re-coded IN FULL "
                  "across every community already done. An amendment applied only to later "
                  "communities makes the variable mean two different things in one column."])
dv(ws, ["H"], V["resolution_type"])
dv(ws, ["K", "L"], V["yes_no"])

# ==========================================================================
# 10. Reliability_Report — 16 columns, pre-filled with 14 variables
# ==========================================================================
RR = ["variable", "variable_type", "agreement_method", "threshold_declared",
      "raw_agreement", "agreement_coefficient", "icc", "mean_absolute_difference",
      "within_uncertainty_band", "marginal_distribution_notes", "calibration_start",
      "calibration_midpoint", "calibration_end", "drift_result",
      "action_if_below_threshold", "notes"]
ws = sheet("Reliability_Report", RR,
           widths={"variable": 26, "agreement_method": 40, "marginal_distribution_notes": 40,
                   "action_if_below_threshold": 40, "notes": 40},
           notes=["PRE-FILLED with the fourteen variables the plan requires. Report reliability "
                  "PER VARIABLE — never one figure for the whole exercise.",
                  "",
                  "v6 of this workbook listed twenty-four variables, thirteen of which were "
                  "practice codes. Those are gone."])
REL = [
    ("e2_settlement_type", "categorical", "chance-corrected agreement",
     "The highest-consequence categorical in Stage 1: a coding error here removes a community "
     "from the sample entirely."),
    ("date_intervention_onset", "date", "ICC + mean absolute difference + agreement within band",
     "Report MAD in YEARS beside the coefficient."),
    ("onset_evidence_rank", "ordered categorical", "weighted chance-corrected agreement",
     "Weights disagreements by distance between ranks."),
    ("onset_confidence_tier", "ordered categorical", "weighted chance-corrected agreement",
     "Load-bearing for cohort admission and decision D4."),
    ("onset_proxy_flag", "categorical", "chance-corrected agreement",
     "Determines which sites leave analysis A4."),
    ("cohort_candidate", "categorical", "chance-corrected agreement",
     "Determines which communities enter analysis A7."),
    ("evidence_tier", "ordered categorical", "weighted chance-corrected agreement",
     "Report the marginal distribution: tiers are usually skewed. An evidence-quality "
     "variable, not a performance measure."),
    ("managed_area_ha", "continuous", "ICC + mean absolute difference",
     "Report MAD in HECTARES. It corroborates the drawn polygon rather than setting geometry, "
     "so its reliability bounds the corroboration and not the measurement."),
    ("managed_area_basis", "categorical", "chance-corrected agreement",
     "Determines how much weight a disagreement between the stated and drawn areas carries."),
    ("population_value", "continuous", "ICC + mean absolute difference",
     "A Stage 2 matching criterion, so an error propagates into the control sample."),
    ("status_current", "categorical", "chance-corrected agreement",
     "Report how often 'unknown' was coded. Unknown is not dissolved."),
    ("polygon_area_ha", "continuous", "ICC + mean absolute difference, from the 20% redraw",
     "Report MAD in HECTARES. This is the reliability figure for the study's PRIMARY "
     "MEASUREMENT GEOMETRY."),
    ("polygon_iou (redraw overlap)", "continuous", "mean overlap, and the count below 0.80",
     "Shared area divided by combined area between the two drawings. Below a mean of 0.80, "
     "report SC1 (equal-area circles) as the primary specification instead."),
    ("reference_circle", "ordered categorical", "weighted chance-corrected agreement",
     "Derived from polygon area, so its reliability follows the polygon's."),
]
for i, (v, t, m, note) in enumerate(REL, start=2):
    ws.cell(row=i, column=1, value=v).font = Font(size=9, bold=True)
    ws.cell(row=i, column=2, value=t).font = Font(size=9)
    ws.cell(row=i, column=3, value=m).font = Font(size=9)
    ws.cell(row=i, column=4, value="declared before coding").font = Font(size=9)
    ws.cell(row=i, column=15,
            value="Revise the definition, re-code this variable in full, report both figures"
            ).font = Font(size=9)
    c = ws.cell(row=i, column=16, value=note)
    c.font = Font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
dv(ws, ["B"], V["variable_type"], last=len(REL) + 1)

# ==========================================================================
# 11. Enquiry_Record — 19 columns
# ==========================================================================
ER = ["enquiry_id", "site_id", "reason_for_contact", "date_sent", "medium",
      "response_status", "date_responded", "consent_to_answer",
      "consent_to_identification", "consent_to_location_disclosure",
      "withdrawal_request", "onset_information", "managed_area_information",
      "other_facts_supplied", "source_class", "personal_identifiers_retained",
      "ethics_reference", "coder_id", "notes"]
ws = sheet("Enquiry_Record", ER,
           widths={"reason_for_contact": 34, "onset_information": 30, "notes": 30},
           notes=["Decision D1 recommends coding from PUBLISHED SOURCES ONLY, and contacting "
                  "only those communities whose onset remains genuinely undeterminable. That is "
                  "usually a small number and it keeps the ethics application narrow.",
                  "",
                  "If you contact anyone, every row here needs an ethics reference and a "
                  "consent state. Non-response is NEVER exclusion: communities that answer "
                  "unsolicited academic enquiries differ systematically from those that do not."])
dv(ws, ["E"], V["enquiry_medium"])
dv(ws, ["F"], V["response_status"])
dv(ws, ["H", "I", "J", "K", "P"], V["consent"])
dv(ws, ["O"], V["source_class"])

# ==========================================================================
# 12. Definitions_And_Freeze — 11 columns, pre-filled, ONE freeze date
# ==========================================================================
DF = ["variable", "definition", "inclusion_example_1", "inclusion_example_2",
      "exclusion_example_1", "allowed_values", "agreement_threshold",
      "threshold_rationale", "version", "frozen_date", "notes"]
ws = sheet("Definitions_And_Freeze", DF,
           derived={"frozen_date"},
           widths={"variable": 26, "definition": 52, "inclusion_example_1": 38,
                   "inclusion_example_2": 38, "exclusion_example_1": 38,
                   "allowed_values": 40, "threshold_rationale": 46, "notes": 30},
           notes=[
               "THE DEFINITIONS FREEZE. This sheet replaces the practice codebook of workbook "
               "v6, which defined thirteen practice codes that no longer exist.",
               "",
               "It carries the definition, the worked examples and the declared agreement "
               "threshold for every JUDGEMENT-BEARING variable that survives — the variables "
               "where two careful readers of the same page could reasonably differ.",
               "",
               "Enter ONE date in the yellow cell at M3. Every frozen_date fills itself from "
               "it. After that date the definitions may not change. A later change means the "
               "affected variable is re-coded IN FULL, and the change is logged in "
               "Decision_Log with changes_results assessed honestly.",
               "",
               "Enter the freeze date only when all of these are true:",
               "  1. Every definition and example is complete.",
               "  2. Every agreement_threshold is confirmed.",
               "  3. Calibration passed and is recorded in Calibration.",
               "  4. This file is committed to your repository.",
           ])
DEFS = [
    ("date_intervention_onset",
     "The year the first deliberate action to alter vegetation, soil, water or land cover for "
     "ecological purposes is documented. NOT the founding year, not the year of land purchase "
     "and not the year of first residence.",
     "A dated grant record funding a planting scheme", "A permit for earthworks, dated",
     "A statement of intention to begin work, with no evidence work began",
     "year, or 'not found'", 0.80,
     "Objectivity tier 1 — a dated record either exists or it does not."),
    ("onset_evidence_rank",
     "The strength of the evidence behind the onset year, on the declared five-point scale.",
     "Rank 1: a dated independent record — permit, grant award or report, project report, "
     "registry entry, academic paper or thesis",
     "Rank 2: a dated archived snapshot describing the work as ALREADY UNDER WAY",
     "Rank 5 is not an onset at all: a directory founding year used as a proxy. Set "
     "onset_proxy_flag = yes",
     "1 / 2 / 3 / 4 / 5", 0.75,
     "Objectivity tier 2 — the scale is explicit but rank 2 against rank 3 turns on whether a "
     "snapshot describes work under way or work intended."),
    ("onset_confidence_tier",
     "How precisely the onset year is known: A precise, B plus or minus one year, C uncertain "
     "beyond that.",
     "A: a dated permit naming the month",
     "B: two independent sources agreeing within one year",
     "C: a retrospective community statement giving a decade",
     "A / B / C", 0.80,
     "Objectivity tier 1 — it follows mechanically from the band width, once the band is set."),
    ("onset_proxy_flag",
     "Whether the recorded onset is a founding year standing in for an unknown intervention "
     "year.",
     "Yes: the only date found is a directory founding year",
     "Yes: the community says 'founded 1985' and nothing about when work began",
     "No: the community dates its first planting, whatever its founding year",
     "yes / no", 0.85,
     "Objectivity tier 1 — it restates what the evidence description already says."),
    ("cohort_candidate",
     "Whether the onset falls in the longitudinal cohort window.",
     "core: onset 2020 or 2021", "extension: onset 2019",
     "uncertain: the onset band straddles the boundary — resolve before decision D4",
     "core (2020-2021) / extension (2019) / no / uncertain", 0.85,
     "Objectivity tier 1 — it follows from the onset year and its band."),
    ("e2_settlement_type",
     "What kind of entity the site is, for eligibility criterion E2.",
     "village-scale permanent residence: permanent residents on a shared holding",
     "campus: an institution with a residential ecological programme",
     "retreat centre: visitors rather than residents, however ecological the land management",
     "village-scale permanent residence / retreat centre / campus / business / single "
     "household / urban co-housing / unclear", 0.80,
     "Objectivity tier 1 by evidence, but the highest-consequence code in Stage 1: it decides "
     "whether a community is in the study at all."),
    ("e8_setting_at_onset",
     "Whether the site was rural or peri-urban rather than urban at the time of onset.",
     "rural: dispersed settlement, agricultural surroundings",
     "peri-urban: at the edge of a built-up area, still with worked land",
     "urban: inside a continuously built-up area — the community is ineligible",
     "rural / peri-urban / urban / unclear", 0.80,
     "Objectivity tier 1 — settlement pattern at onset is a matter of record and imagery."),
    ("managed_area_ha",
     "The land the community actively works ecologically, in hectares, as stated by a source. "
     "NOT the total landholding.",
     "'15 hectares under cultivation' in a thesis site description",
     "A grant application stating the worked area it was funded against",
     "'our 200 hectares' where the community visibly works twelve — that is "
     "total_holding_ha",
     "hectares, or 'not found'", 0.75,
     "Objectivity tier 2 — the figure is stated, but deciding whether a stated figure refers to "
     "worked land or to the whole holding is a judgement."),
    ("managed_area_basis",
     "How the documentary area figure was arrived at.",
     "measured: a thesis, survey, land registry entry, grant application or scaled site plan",
     "stated: the community gives a figure without saying how it was arrived at",
     "inferred: you derived it from a stated number of beds, plots or hectares under a named "
     "crop — record how in documentary_area_note",
     "measured / stated / inferred / not found", 0.75,
     "Objectivity tier 2 — measured against stated turns on whether the source says how it "
     "knows."),
    ("population_value",
     "Permanent residents only.",
     "'twenty-eight adults and nine children live here permanently'",
     "A census or registry figure for the settlement",
     "Visitor, volunteer, student or course-participant numbers",
     "integer, or 'not found'", 0.80,
     "Objectivity tier 1 — the number is stated; the judgement is only whom it counts."),
    ("v1_self_documentation",
     "The community describes PARTICULAR ACTIONS it has taken on the land, not aims it holds.",
     "'we planted 400 trees along the north boundary in 2014'",
     "A dated work log or project page describing completed work",
     "'we are committed to regenerating this land' — an aim, not an action",
     "yes / no", 0.75,
     "Objectivity tier 2 — specificity against aspiration is exactly where two readers differ."),
    ("v2_external_documentation",
     "An account of the WORK by someone other than the community: academic paper, thesis, "
     "project record, certification, grant award, or media coverage of the work.",
     "A thesis whose author visited and described the land management",
     "A LIFE or LEADER grant record naming the community and the project",
     "A news piece about the community that describes only its philosophy",
     "yes / no", 0.80,
     "Objectivity tier 1 — an external source either exists and describes the work, or does "
     "not. This is the channel that separates evidence tier A."),
    ("v3_substantive_affiliation",
     "Membership of a body that ASSESSES practice, rather than one that merely lists members.",
     "An organic certifier with a published client list and inspection regime",
     "A participatory guarantee system with documented criteria",
     "A directory that lists any community that submits a form",
     "yes / no", 0.75,
     "Objectivity tier 2 — 'assesses' against 'lists' needs the exclusion example to separate."),
    ("v4_visual_documentation",
     "Dated photographs, site plans, design drawings or maps showing the land or its structures.",
     "A dated photograph of a completed pond, terrace or planted block",
     "A site plan or design drawing with a date",
     "An undated photograph, or one whose date is inferred from its position in a feed",
     "yes / no", 0.75,
     "Objectivity tier 2 — the judgement is whether the date is genuinely attached."),
    ("v5_continuity_evidence",
     "The work is described consistently across YEARS, rather than announced once.",
     "Archived snapshots from 2011, 2016 and 2022 all describing the same worked land",
     "A run of dated posts or reports across several years",
     "A single announcement, however detailed",
     "yes / no", 0.75,
     "Objectivity tier 2 — how many years and how consistent is a matter of degree."),
    ("status_current",
     "The present state of the community.",
     "active: the site is operating and inhabited",
     "dormant: activity reduced or suspended, but not dissolved",
     "dissolved requires POSITIVE evidence. A vanished website is 'unknown', never 'dissolved'",
     "active / dormant / transformed / relocated / dissolved / unknown", 0.80,
     "Objectivity tier 1 by rule, but the unknown-is-not-dissolved discipline is the whole "
     "difficulty."),
    ("polygon_confidence",
     "How clearly the managed boundary could be read from the imagery.",
     "clear: field edges, tracks and terrace lines define the boundary unambiguously",
     "moderate: most of the boundary is structural, part is inferred from field pattern",
     "poor: the boundary rests substantially on judgement — report these separately",
     "clear / moderate / poor", 0.70,
     "Objectivity tier 3 — it is a self-assessment of a judgement, and is reported beside the "
     "redraw overlap rather than instead of it."),
]
for i, (v, d, e1, e2, x1, av, th, rat) in enumerate(DEFS, start=2):
    ws.cell(row=i, column=1, value=v).font = Font(size=9, bold=True)
    for col, val in ((2, d), (3, e1), (4, e2), (5, x1), (6, av)):
        c = ws.cell(row=i, column=col, value=val)
        c.font = Font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=7, value=th).font = Font(size=9)
    c = ws.cell(row=i, column=8, value=rat)
    c.font = Font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=i, column=9, value="1.0").font = Font(size=9)
    ws.cell(row=i, column=10, value='=IF($M$3="","NOT YET FROZEN",TEXT($M$3,"YYYY-MM-DD"))')
    ws.cell(row=i, column=10).font = Font(size=9)
ws.cell(row=2, column=13, value="FREEZE DATE (YYYY-MM-DD)").font = SUB_FONT
fc = ws.cell(row=3, column=13)
fc.fill = YELLOW_FILL
fc.border = Border(outline=True, left=Side("medium"), right=Side("medium"),
                   top=Side("medium"), bottom=Side("medium"))
ws.column_dimensions["M"].width = 26
ws.cell(row=4, column=13,
        value="All definitions freeze together. Enter one date here.").font = NOTE_FONT
ws.cell(row=5, column=13,
        value="Thresholds are chance-corrected agreement coefficients. 0.70 is the floor.").font = NOTE_FONT

# ==========================================================================
# 13. Calibration — 12 columns
# ==========================================================================
CAL = ["calibration_round", "calibration_date", "coder_id", "calibration_set_size", "variable",
       "agreement_with_reference", "threshold_required", "passed", "drift_vs_initial",
       "action_taken", "cleared_for_production", "notes"]
ws = sheet("Calibration", CAL,
           widths={"notes": 44},
           notes=["A coder proceeds to production coding only after meeting the declared "
                  "threshold on a calibration set drawn from communities NOT in the study "
                  "sample.",
                  "",
                  "Run the round again at the midpoint and at the end. Drift between rounds is "
                  "a real and reportable finding about the coding, not a nuisance."])
dv(ws, ["A"], V["calibration_round"])
dv(ws, ["H", "K"], V["yes_no"])

# ==========================================================================
# 14. Decision_Log — 12 columns
# ==========================================================================
DEC = ["decision_id", "date", "stage", "issue", "decision", "rule_or_plan_section",
       "reason", "impact", "requires_recode", "affects_frozen_output",
       "changes_results", "notes"]
ws = sheet("Decision_Log", DEC,
           widths={"issue": 40, "decision": 40, "reason": 44, "notes": 30},
           notes=["Appendix A of the plan. One spreadsheet tab, one row per decision, WITH A "
                  "DATE. This is the entirety of what remains of the old integrity "
                  "architecture, and it takes a few minutes a week.",
                  "",
                  "Log the decisions you later REVERSED as well. A log showing only what "
                  "survived is a tidied story, and a reader who suspects tidying discounts all "
                  "of it. The reversals are what make it credible.",
                  "",
                  "Seed values belong here too: the reference-pool draw, the semi-urban split, "
                  "the double-coding subsample, and any permutation or bootstrap seed. Record "
                  "each on the day you set it."])
dv(ws, ["C"], V["stage"])
dv(ws, ["I", "J"], V["yes_no"])
dv(ws, ["K"], V["changes_results"])

# ==========================================================================
# 15. Reference_Codes
# ==========================================================================
ws = wb.create_sheet("Reference_Codes")
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 74
ws.column_dimensions["C"].width = 74
ws.freeze_panes = "A2"
ws.cell(row=1, column=1, value="REFERENCE CODES — the full vocabulary, with the rule behind "
                               "each value").font = TITLE_FONT

REF = [
    ("H", "Source classes (S1 strongest ... S8 weakest)", ""),
    ("S1", "ACADEMIC — peer-reviewed papers, THESES AND DISSERTATIONS, conference papers, preprints",
     "Where most RANK 1 onset evidence lives"),
    ("S2", "INSTITUTIONAL — government, NGO, certification, land registry, GRANT AND FUNDING RECORDS",
     "A grant record is dated, independent and public"),
    ("S3", "External network or directory profile", "Externally hosted, usually self-submitted — often a COPY of S4"),
    ("S4", "The community's own current published material", "Richest and least independent"),
    ("S5", "Archived snapshots of community material", "Uniquely valuable for DATING"),
    ("S6", "Journalism and documentary media", "Independent but often uncritical"),
    ("S7", "Community social media and member accounts", "Supporting evidence only; never the sole basis"),
    ("S8", "Direct communication with the community", "Authoritative on facts the community knows; needs ethics clearance"),
    ("H", "Onset evidence ranks and their typical uncertainty bands", ""),
    ("1", "Dated independent record — permit, GRANT AWARD OR REPORT, project report, registry entry, ACADEMIC PAPER OR THESIS",
     "0 to +/-1 year"),
    ("2", "Dated archived snapshot describing work ALREADY under way", "+/-1 to +/-3 years; firm UPPER bound"),
    ("3", "The community's own dated retrospective statement", "+/-2 to +/-5 years; drifts to round numbers"),
    ("4", "Undated community statement, onset inferred from context", "+/-5 years or wider; flagged"),
    ("5", "Directory founding year used as a proxy", "NOT an onset. Set onset_proxy_flag = yes. Excluded from analysis A4"),
    ("H", "Onset confidence tier — gates cohort admission", ""),
    ("A", "Precise year, rank 1 or 2 evidence", "Usable in the cohort without qualification"),
    ("B", "Plus or minus one year", "Usable in the cohort"),
    ("C", "Uncertain beyond +/-1 year", "NOT usable to anchor a before-and-after window"),
    ("H", "Cohort candidacy", ""),
    ("core (2020-2021)", "Onset in 2020 or 2021", "3 pre-onset years and 4 post-onset years inside 2017-2025"),
    ("extension (2019)", "Onset in 2019", "Only 2 pre-years, and 2017 is half a single-satellite year"),
    ("no", "Onset outside 2019-2021", "The large majority"),
    ("uncertain", "Onset band straddles the boundary", "Resolve before decision D4"),
    ("H", "Evidence verification channels — HOW WELL DOCUMENTED, not how much is done", ""),
    ("V1", "Community self-documentation of SPECIFIC work", "Specificity distinguishes this from an aspiration"),
    ("V2", "External documentation of the work itself — INCLUDING A THESIS OR GRANT RECORD",
     "Independent of the community's account. The channel that separates tier A"),
    ("V3", "Affiliation with a body applying substantive criteria", "Record the body and its criteria"),
    ("V4", "Visual documentation: dated photos, plans, design drawings", "The date must be genuinely attached"),
    ("V5", "Continuity across years in snapshots or publications", "Distinguishes sustained practice from one announcement"),
    ("H", "Evidence tiers — computed from V1..V5, counting INDEPENDENCE GROUPS", ""),
    ("EXTERNAL CHANNELS", "V2 and V3 are the external channels",
     "Both rest on somebody other than the community: a researcher, a funder, a certifier. V1, V4 and V5 are the community's own voice"),
    ("A", "At least one EXTERNAL channel, and 3 or more channels in total",
     "Externally documented and corroborated. The strongest available assurance"),
    ("B", "At least one EXTERNAL channel, and 2 channels in total",
     "Externally documented but thin"),
    ("C", "NO external channel, and 2 or more channels",
     "Community-documented only. However many forms the community's own material takes, it is one voice"),
    ("Fail", "Fewer than 2 channels, or aims without specific actions", "Not eligible"),
    ("Why this ladder and not v2.4's", "v2.4 ordered the tiers by CHANNEL COUNT first and independence second",
     "That made tier C unreachable — the only channel that is neither external nor visual nor continuity is V1, and no community can have two of V1 — and it ranked three self-documented channels above two including a thesis. v4.0 orders by independence first. Every tier is now reachable"),
    ("SC18 restriction", "Sensitivity check SC18 restricts to evidence_tier A or B — sites with at least one external channel",
     "Tier A alone is reported as a stricter variant"),
    ("WHAT THIS TIER IS NOT", "It is NOT a measure of ecological activity, quality or performance.",
     "It measures how well a community documents itself, which correlates with size, age and "
     "organisational capacity. Use it as a confounder, a sample descriptor and a restriction. "
     "Never as an outcome, and never disaggregated by kind of activity."),
    ("H", "Status and persistence", ""),
    ("active", "Operating and inhabited", ""),
    ("dormant", "Reduced or suspended activity, not dissolved", ""),
    ("transformed", "Continues under a materially different form or purpose", ""),
    ("relocated", "The community moved to a different site", ""),
    ("dissolved", "The community ceased to exist", "Requires POSITIVE evidence"),
    ("unknown", "Cannot be determined", "UNKNOWN IS NOT DISSOLVED"),
    ("H", "Area types", ""),
    ("actively managed", "Land the community actually works ecologically",
     "CORROBORATES the drawn polygon. It does not set geometry"),
    ("total holding only", "Source gives the whole holding and no managed extent",
     "Record it as total_holding_ha. NEVER substitute one for the other"),
    ("both recorded", "Both figures available and recorded separately", ""),
    ("not stated", "Neither figure available", "area_agreement_tier becomes B; the polygon still gives full geometry"),
    ("H", "Managed-area basis — how the documentary figure was arrived at", ""),
    ("measured", "Someone actually measured it", "Thesis, survey, land registry, grant application, scaled site plan"),
    ("stated", "The community gives a figure without saying how", "Commonest case; frequently the TOTAL HOLDING"),
    ("inferred", "You derived it from something else", "Number of beds, plots, hectares under a named crop. Say how in the note"),
    ("not found", "No figure anywhere", "A complete and correct answer. Tier B, and the polygon is unaffected"),
    ("H", "Reference circle — matched to the POLYGON's area, filled by formula", ""),
    ("r75", "polygon under 2.6 ha", "75 m radius, 1.77 ha. Also the fallback where a polygon is under the 1.0 ha minimum"),
    ("r110", "2.6 to 5.2 ha", "110 m radius, 3.80 ha"),
    ("r150", "5.2 to 9.9 ha", "150 m radius, 7.07 ha"),
    ("r210", "9.9 to 19.8 ha", "210 m radius, 13.85 ha"),
    ("r300", "over 19.8 ha", "300 m radius, 28.27 ha"),
    ("H", "Area agreement tier — drawn polygon against the documentary figure", ""),
    ("A", "Documentary area within 30% of the polygon area", "The two agree. SC16 restricts to these"),
    ("B", "NO documentary figure, OR the two differ by 30 to 100%", "The polygon stands and both numbers are kept"),
    ("C", "They differ by more than a factor of two",
     "The polygon still stands. Investigate: usually the source quoted the TOTAL HOLDING, "
     "sometimes the holding is genuinely non-contiguous"),
    ("H", "Platform type — Source_Set", ""),
    ("own website", "The community's current primary site", "Richest, least independent. Start here"),
    ("secondary or former website", "An older domain, a project microsite, a second-language site",
     "Old domains hold the OLDEST material and are the best dating source on the open web"),
    ("Facebook", "Page or group", "About tab carries a creation date; albums and events are dated"),
    ("Instagram", "Profile", "Rarely readable without an account. Report BLOCKED rather than guessing"),
    ("YouTube", "Channel", "Sort videos OLDEST FIRST. An upload date is a dated record"),
    ("Vimeo", "Channel", "Same use as YouTube"),
    ("blog platform", "WordPress, Blogspot, Medium, Substack",
     "The RSS feed and sitemap give the full dated post list in one request"),
    ("directory listing", "GEN, FIC, NuMundo, WWOOF, Workaway, national networks",
     "Self-submitted, so usually the SAME independence group as the website. Its archive gives last_listing_year"),
    ("crowdfunding", "Kickstarter, GoFundMe, Ulule, Betterplace and similar",
     "A dated campaign page describing a specific project — often rank 1 or 2 onset evidence"),
    ("LinkedIn", "Organisation page", "Carries a founded year"),
    ("booking or hosting", "Airbnb, Hostelworld, retreat and volunteer platforms",
     "Often describes the LAND in detail where the website does not"),
    ("news outlet", "A local or national outlet's page about the community", "S6. Dated, independent, frequently uncritical"),
    ("other", "Anything else", "Say what it is in notes"),
    ("H", "Crawl status — Source_Set", ""),
    ("crawled", "Opened and worked through to the depth the protocol asks for", ""),
    ("partial", "Opened, but the enumeration was cut short", "Say why in notes and set crawl_truncated = yes"),
    ("blocked", "Requires an account, or refused automated reading", "A REPORTED block is data. A guess about its content is fabrication"),
    ("dead link", "The address no longer resolves", "Go to the Wayback Machine — a dead domain is often the richest one"),
    ("not attempted", "Not reached in this run", "Every address must end at something other than this before the community is closed"),
    ("H", "Independence group — Source_Index and Source_Set", ""),
    ("What it is", "A short id (G1, G2, G3) shared by every source deriving from the same underlying statement",
     "Assign it as you read, not afterwards"),
    ("The test", "Could this source be wrong in the same way as that one, for the same reason?", "If yes, same group"),
    ("Same group", "A website, its own Facebook page, and a directory listing copied from it", "One voice, however many URLs"),
    ("Different groups", "The community's own account; a thesis; a grant record; a newspaper",
     "These can contradict each other, which is what makes agreement worth something"),
    ("What it changes", "channel_count, evidence_tier, and any claim that sources corroborate", "All count GROUPS, never URLs"),
    ("H", "The three search outcomes — never represent them identically", ""),
    ("EVIDENCE FOUND", "A source states the value", "Record the value and the source id"),
    ("EVIDENCE ABSENT", "An adequately exhaustive search found nothing",
     "Record 'not found' AND the negative consultations in Search_Log. This is a finding"),
    ("SEARCH INCOMPLETE", "The run was truncated, blocked or cut short",
     "Set crawl_truncated = yes and say in stages_completed where it stopped. This is NOT a finding"),
    ("H", "WHAT NOT TO SEARCH FOR — these come from Google Earth Engine or from your own drawing", ""),
]
for code, meaning, note in [
    ("VM1-VM14, PM1-PM3, FC1-FC4, CA", "All vegetation, provisioning, flag and contour metrics",
     "Sentinel-2 and SRTM, computed in the pipeline"),
    ("VCI, VCI-P/S/T/C, PCI, MDS, LCC", "The index and its components", "Derived from the above"),
    ("built_fraction, tree_cover_pct", "Land-cover fractions", "Dynamic World"),
    ("elevation_m, slope_deg, terrain class", "Terrain", "SRTM"),
    ("water_dist_m", "Distance to permanent water", "Global Surface Water — a matching criterion, but satellite-derived"),
    ("koppen_group, biome", "Stratifiers", "Beck et al. / RESOLVE ecoregions"),
    ("rainfall, driest month, drought year", "Climate", "CHIRPS"),
    ("n_clear", "Clear-observation count", "Extraction output"),
    ("control_distance_km", "How far each control sits from its settlement", "A Stage 2 MATCHING output, not documentary coding"),
    ("polygon_area_ha, polygon_iou", "The drawn geometry and its reliability",
     "YOUR measurement, on Polygon_Geometry. Never estimated from imagery by a search assistant"),
    ("reference_circle, equal_area_circle_radius_m", "Derived from the polygon area",
     "Computed by formula. NEVER derived from Earth Engine — that would set the measurement "
     "zone from the very signal measured inside it, which is circular and breaks blindness"),
]:
    REF.append((code, meaning, note))

r = 2
for a, b, c in REF:
    if a == "H":
        cell = ws.cell(row=r, column=1, value=b)
        cell.font = SUB_FONT
        for col in (1, 2, 3):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="E3ECF2")
    else:
        ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=9)
        for col, val in ((2, b), (3, c)):
            cc = ws.cell(row=r, column=col, value=val)
            cc.font = Font(size=9)
            cc.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ==========================================================================
# Order the sheets and save
# ==========================================================================
ORDER = ["README", "Cohort_Tracker", "Community_Register", "Onset_Register",
         "Polygon_Geometry", "Source_Index", "Source_Set", "Search_Log",
         "Disagreement_Log", "Reliability_Report", "Enquiry_Record",
         "Definitions_And_Freeze", "Calibration", "Decision_Log", "Reference_Codes"]
wb._sheets = [wb[n] for n in ORDER]

wb.save(OUT)
print("wrote", os.path.normpath(OUT))
print("sheets:", len(wb.sheetnames), wb.sheetnames)
assert len(wb.sheetnames) == COUNTS["workbook_sheets"], len(wb.sheetnames)
print("Community_Register columns:", len(CR))
print("Onset_Register columns:", len(ORG))
print("Polygon_Geometry columns:", len(PG))
m, rf, c, t = panel_rows()
print("panel rows check:", t)
